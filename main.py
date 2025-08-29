import os
import sys
import time
import json
import threading
import traceback
from datetime import datetime
import random
import configparser
import uuid
import psutil
import ctypes
from PySide6.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit, QTreeWidget,
    QTreeWidgetItem, QGroupBox, QMessageBox, QMenu, QDialog,
    QTextEdit, QFileDialog, QComboBox, QCheckBox, QTableWidget, QDialogButtonBox,
    QTableWidgetItem, QFormLayout, QDateEdit, QListWidget, QListWidgetItem,
    QDateTimeEdit, QCalendarWidget, QHeaderView)
from PySide6.QtCore import Qt, QTimer, Signal, QObject, QDateTime, QDate, QTime, QThread, QMetaObject, Q_ARG
from styles import StyleSheet, apply_stylesheet
from wechat import (
    SimpleWeChatInfo, WeChatMessageMonitor, start_new_wechat,
    send_message_to_wxid, send_image_to_wxid,
    parse_special_message, add_wechat_friend, get_wechat_resources,
    add_friend_by_phone, ContactInfoMonitor, 
    send_message_simple, send_image_simple,
    RemarkModifier, get_group_members, get_all_group_members,
    OpenProcess, CloseHandle, get_wechat_base, detect_wechat_processes,
    get_wechat_service
)
from aizhuli_combined import AIAssistantTab, get_last_yuanbao_sender

def _check_single_instance(name: str = "Global\\WeChatManagerAppMutex") -> bool:
    """检查程序是否已运行"""
    try:
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        handle = kernel32.CreateMutexW(None, False, name)
        if handle and ctypes.get_last_error() == 183:
            kernel32.CloseHandle(handle)
            return False
        return True
    except Exception:
        return True

def _show_already_running_message():
    """显示程序已运行的消息"""
    try:
        ctypes.windll.user32.MessageBoxW(None, "程序已在运行，无需重复打开。", "AI全功能营销系统", 0x40)
    except Exception:
        pass

def setup_tree_columns(tree, specs=None, default_size=150, stretch_last=True):
    """设置树型组件的列宽度"""
    header = tree.header()
    header.setDefaultSectionSize(default_size)
    header.setStretchLastSection(stretch_last)
    
    col_count = getattr(tree, 'columnCount', lambda: header.count() if hasattr(header, 'count') else 0)()
    
    for i in range(col_count):
        header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
    
    if col_count > 0:
        header.resizeSection(0, 50)
        for i in range(1, col_count):
            header.resizeSection(i, 150)

class TaskTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.selected_contacts = []
        self.scheduled_tasks = []
        self.task_timer = QTimer()
        self.task_timer.timeout.connect(self.check_scheduled_tasks)
        self.task_timer.start(1000)

        self.selected_wxids = set()
        self.init_ui()



    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        button_group = QGroupBox("")
        button_layout = QHBoxLayout()
        self.mass_all_contacts_btn = QPushButton("群发所有联系人")
        self.mass_all_contacts_btn.clicked.connect(self.mass_all_contacts)
        button_layout.addWidget(self.mass_all_contacts_btn, alignment=Qt.AlignLeft)
        self.mass_all_groups_btn = QPushButton("群发所有群列表")
        self.mass_all_groups_btn.clicked.connect(self.mass_all_groups)
        button_layout.addWidget(self.mass_all_groups_btn, alignment=Qt.AlignLeft)
        self.custom_select_btn = QPushButton("自定义选择群发")
        self.custom_select_btn.clicked.connect(self.custom_select_mass)
        button_layout.addWidget(self.custom_select_btn, alignment=Qt.AlignLeft)
        button_layout.addStretch()

        button_group.setLayout(button_layout)
        main_layout.addWidget(button_group)

        schedule_group = QGroupBox("定时发送设置")
        schedule_layout = QVBoxLayout()
        send_mode_layout = QHBoxLayout()
        self.scheduled_check = QCheckBox("定时发送")
        self.scheduled_check.stateChanged.connect(self.on_send_mode_changed)
        send_mode_layout.addWidget(self.scheduled_check)
        send_mode_layout.addSpacing(15)
        send_mode_layout.addWidget(QLabel("发送时间:"))
        self.schedule_datetime = QDateTimeEdit()
        self.schedule_datetime.setDateTime(QDateTime.currentDateTime().addSecs(300))
        self.schedule_datetime.setCalendarPopup(True)
        self.schedule_datetime.setDisplayFormat("yyyy-MM-dd HH:mm")
        self.schedule_datetime.setEnabled(False)
        send_mode_layout.addWidget(self.schedule_datetime)
        self.select_time_btn = QPushButton("选择时间")
        self.select_time_btn.clicked.connect(self.show_time_dialog)
        self.select_time_btn.setEnabled(False)
        send_mode_layout.addWidget(self.select_time_btn)
        send_mode_layout.addStretch()
        schedule_layout.addLayout(send_mode_layout)
        delay_layout = QHBoxLayout()
        delay_layout.addWidget(QLabel("对象切换间隔:"))
        self.min_delay = QLineEdit("1")
        self.min_delay.setFixedWidth(25)
        self.min_delay.setMaximumHeight(25)
        delay_layout.addWidget(self.min_delay)
        delay_layout.addWidget(QLabel("至"))
        self.max_delay = QLineEdit("3")
        self.max_delay.setFixedWidth(25)
        self.max_delay.setMaximumHeight(25)
        delay_layout.addWidget(self.max_delay)
        delay_layout.addWidget(QLabel("秒"))
        delay_layout.addStretch()
        schedule_layout.addLayout(delay_layout)
        schedule_group.setLayout(schedule_layout)
        main_layout.addWidget(schedule_group)

        content_group = QGroupBox("发送内容")
        content_layout = QVBoxLayout()
        text_layout = QHBoxLayout()
        text_layout.addWidget(QLabel("消息文本:"))
        self.message_text = QTextEdit()
        self.message_text.setMaximumHeight(80)
        self.message_text.setPlaceholderText("请输入要发送的文本消息...")
        text_layout.addWidget(self.message_text)
        content_layout.addLayout(text_layout)
        media_layout = QHBoxLayout()
        media_layout.addWidget(QLabel("图片/文件/视频路径:"))
        self.image_path = QLineEdit()
        self.image_path.setPlaceholderText("请输入图片/文件/视频路径...")
        media_layout.addWidget(self.image_path)
        self.select_image_btn = QPushButton("选择文件")
        self.select_image_btn.clicked.connect(self.select_image)
        media_layout.addWidget(self.select_image_btn)
        content_layout.addLayout(media_layout)
        content_group.setLayout(content_layout)
        main_layout.addWidget(content_group)

        selected_group = QGroupBox("已选中联系人明细，选中联系人双击可删除")
        selected_layout = QVBoxLayout()
        self.selected_contacts_list = QListWidget()
        self.selected_contacts_list.itemDoubleClicked.connect(self.remove_selected_contact)
        selected_layout.addWidget(self.selected_contacts_list)
        actions_layout = QHBoxLayout()
        clear_selection_btn = QPushButton("清空选择")
        clear_selection_btn.clicked.connect(self.clear_selection)
        actions_layout.addWidget(clear_selection_btn)
        self.confirm_btn = QPushButton("确定")
        self.confirm_btn.clicked.connect(self.confirm_create_task)
        actions_layout.addWidget(self.confirm_btn)
        actions_layout.addStretch()
        selected_layout.addLayout(actions_layout)
        selected_group.setLayout(selected_layout)
        main_layout.addWidget(selected_group)

        task_group = QGroupBox("定时任务列表")
        task_layout = QVBoxLayout()
        self.task_table = QTableWidget()
        self.task_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.task_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.task_table.setColumnCount(6)
        self.task_table.setHorizontalHeaderLabels([
            "序号", "任务名称", "发送时间", "目标数量", "状态", "操作"
        ])
        self.task_table.verticalHeader().setVisible(False)
        header = self.task_table.horizontalHeader()
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        for i in range(self.task_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.resizeSection(0, 50)
        for i in range(1, self.task_table.columnCount()):
            header.resizeSection(i, 150)
        task_layout.addWidget(self.task_table)
        task_buttons = QHBoxLayout()
        self.delete_selected_btn = QPushButton("删除选中任务")
        self.delete_selected_btn.clicked.connect(self.delete_selected_tasks)
        task_buttons.addWidget(self.delete_selected_btn)
        task_buttons.addStretch()
        task_layout.addLayout(task_buttons)
        task_group.setLayout(task_layout)
        main_layout.addWidget(task_group)

    def on_send_mode_changed(self):
        is_scheduled = self.scheduled_check.isChecked()
        self.schedule_datetime.setEnabled(is_scheduled)
        self.select_time_btn.setEnabled(is_scheduled)
        if is_scheduled:
            self.show_time_dialog()

    def show_time_dialog(self):
        """显示时间选择对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("选择发送时间")
        dialog.setMinimumWidth(350)
        layout = QVBoxLayout(dialog)
        
        # 日历组件
        calendar = QCalendarWidget()
        calendar.setMinimumDate(QDate.currentDate())
        calendar.setSelectedDate(QDateTime.currentDateTime().date())
        layout.addWidget(calendar)

        # 时间输入
        time_container = QWidget()
        time_layout = QHBoxLayout(time_container)
        time_layout.setContentsMargins(10, 5, 10, 5)

        time_layout.addWidget(QLabel("时间:"))

        hour_input = QLineEdit(str(datetime.now().hour))
        hour_input.setFixedSize(60, 30)
        time_layout.addWidget(hour_input)
        
        time_layout.addWidget(QLabel(":"))
        
        minute_input = QLineEdit(str(datetime.now().minute))
        minute_input.setFixedSize(60, 30)
        time_layout.addWidget(minute_input)
        
        time_layout.addStretch()
        layout.addWidget(time_container)
        
        preview_label = QLabel()
        layout.addWidget(preview_label)

        def update_preview():
            """更新预览时间"""
            try:
                selected_date = calendar.selectedDate()
                hour = int(hour_input.text())
                minute = int(minute_input.text())
                selected_time = QTime(hour, minute)
                selected_datetime = QDateTime(selected_date, selected_time)
                preview_label.setText(f"预计发送时间: {selected_datetime.toString('yyyy年MM月dd日 HH:mm')}")
            except (ValueError, TypeError):
                preview_label.setText("请输入有效的时间")
        
        # 绑定事件
        calendar.selectionChanged.connect(update_preview)
        hour_input.textChanged.connect(update_preview)
        minute_input.textChanged.connect(update_preview)
        update_preview()

        # 按钮布局
        button_layout = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._apply_selected_time(calendar, hour_input, minute_input)
    
    def _apply_selected_time(self, calendar, hour_input, minute_input):
        """应用选中的时间"""
        try:
            selected_date = calendar.selectedDate()
            hour = int(hour_input.text())
            minute = int(minute_input.text())
            selected_time = QTime(hour, minute)
            selected_datetime = QDateTime(selected_date, selected_time)

            if selected_datetime <= QDateTime.currentDateTime():
                selected_datetime = QDateTime.currentDateTime().addSecs(300)

            self.schedule_datetime.setDateTime(selected_datetime)
        except (ValueError, TypeError):
            pass

    def select_image(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择文件",
            "",
            "所有文件 (*.*);;图片 (*.png *.jpg *.jpeg *.gif *.bmp);;视频 (*.mp4 *.avi *.mov *.mkv)"
        )
        if file_path:
            self.image_path.setText(file_path)

    def mass_all_contacts(self):
        if not self.check_wechat_login():
            QMessageBox.warning(self, "警告", "请先登录微信")
            return
        contacts = self.get_all_contacts()
        if not contacts:
            QMessageBox.warning(self, "警告", "未找到联系人")
            return
        self.selected_contacts = contacts
        self.selected_wxids = {contact.get('wxid') for contact in contacts if contact.get('wxid')}
        self.update_selected_contacts_list()
        QMessageBox.information(self, "提示", f"已选择 {len(contacts)} 个联系人，请填写发送内容后点击确定按钮")

    def mass_all_groups(self):
        if not self.check_wechat_login():
            QMessageBox.warning(self, "警告", "请先登录微信")
            return
        groups = self.get_all_groups()
        if not groups:
            QMessageBox.warning(self, "警告", "未找到群组")
            return
        self.selected_contacts = groups
        self.selected_wxids = {group.get('wxid') for group in groups if group.get('wxid')}
        self.update_selected_contacts_list()
        QMessageBox.information(self, "提示", f"已选择 {len(groups)} 个群组，请填写发送内容后点击确定按钮")

    def custom_select_mass(self):
        if self.parent:
            self.parent.notebook.setCurrentIndex(0)
            QMessageBox.information(self, "提示", "请在主界面选择要群发的联系人，然后右键选择'定时消息'")

    def get_all_contacts(self):
        """获取所有联系人（不包括群聊）"""
        if not (hasattr(self.parent, 'all_contacts') and self.parent.all_contacts):
            return []
        return [contact for contact in self.parent.all_contacts 
                if '@chatroom' not in contact.get('wxid', '')]

    def get_all_groups(self):
        """获取所有群组"""
        if not (hasattr(self.parent, 'all_contacts') and self.parent.all_contacts):
            return []
        return [contact for contact in self.parent.all_contacts 
                if '@chatroom' in contact.get('wxid', '')]

    def check_wechat_login(self):
        """检查微信是否登录"""
        try:
            wechat_info = SimpleWeChatInfo()
            accounts = wechat_info.run()
            return len(accounts) > 0
        except Exception:
            return False

    def update_selected_contacts_list(self):
        self.selected_contacts_list.clear()
        for contact in self.selected_contacts:
            nickname = contact.get('nickname', '未知')
            wxid = contact.get('wxid', '')
            item_text = f"{nickname} ({wxid})"
            item = QListWidgetItem(item_text)
            item.setData(Qt.ItemDataRole.UserRole, wxid)
            self.selected_contacts_list.addItem(item)

    def remove_selected_contact(self, item):
        """移除选中的联系人"""
        wxid = item.data(Qt.ItemDataRole.UserRole)
        if wxid:
            self.selected_contacts = [c for c in self.selected_contacts if c.get('wxid') != wxid]
            self.selected_wxids.discard(wxid)
            self.update_selected_contacts_list()

    def clear_selection(self):
        """清空所有选择"""
        self.selected_contacts = []
        self.selected_wxids.clear()
        self.selected_contacts_list.clear()

    def confirm_create_task(self):
        """确认创建任务"""
        if not self.selected_contacts:
            QMessageBox.information(self, "提示", "请先选择联系人或群组")
            return
            
        message_text = self.message_text.toPlainText().strip()
        image_path = str(self.image_path.text()).strip().strip('"')
        
        if not message_text and not image_path:
            QMessageBox.information(self, "提示", "请输入文本或选择图片")
            return
            
        task_name = "自定义群发" if len(self.selected_contacts) > 1 else self.selected_contacts[0].get('nickname', '任务')
        
        if not self.scheduled_check.isChecked():
            reply = QMessageBox.question(
                self, "确认发送",
                "您没有勾选定时发送，将立即发送消息。\n\n是否继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Yes:
                pid = self._get_current_pid()
                if pid:
                    try:
                        min_delay = int(self.min_delay.text())
                        max_delay = int(self.max_delay.text())
                    except (ValueError, TypeError):
                        min_delay, max_delay = 1, 3
                    
                    threading.Thread(
                        target=self.send_messages, 
                        args=(pid, list(self.selected_contacts), message_text, image_path, min_delay, max_delay), 
                        daemon=True
                    ).start()
                    
                    self.selected_contacts = []
                    self.selected_wxids.clear()
                    self.update_selected_contacts_list()
                    self.message_text.clear()
                    self.image_path.clear()
        else:
            if self.schedule_datetime.dateTime() <= QDateTime.currentDateTime():
                self.schedule_datetime.setDateTime(QDateTime.currentDateTime().addSecs(300))
            
            current_pid = self._get_current_pid()
            self.create_scheduled_task(task_name, list(self.selected_contacts), current_pid)
            self.selected_contacts = []
            self.selected_wxids.clear()
            self.update_selected_contacts_list()
            self.message_text.clear()
            self.image_path.clear()

    def _get_current_pid(self):
        if self.parent and hasattr(self.parent, 'current_account_pid'):
            pid = getattr(self.parent, 'current_account_pid', None)
            if pid:
                return pid
        wechat_info = SimpleWeChatInfo()
        accounts = wechat_info.run()
        return accounts[0]['pid'] if accounts else None

    def create_scheduled_task(self, task_name, contacts, pid=None):
        """创建定时任务"""
        try:
            schedule_time = self.schedule_datetime.dateTime()
            message_text = self.message_text.toPlainText()
            image_path = str(self.image_path.text()).strip().strip('"')
            try:
                min_delay = int(self.min_delay.text())
                max_delay = int(self.max_delay.text())
            except (ValueError, TypeError):
                min_delay, max_delay = 1, 3
            
            task = {
                'id': int(time.time()),
                'name': task_name,
                'schedule_time': schedule_time.toSecsSinceEpoch(),
                'contacts': contacts,
                'message_text': message_text,
                'image_path': image_path,
                'min_delay': min_delay,
                'max_delay': max_delay,
                'status': '等待中',
                'pid': pid
            }
            
            self.scheduled_tasks.append(task)
            self.refresh_task_table_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"创建定时任务失败: {str(e)}")

    def add_selected_contacts(self, contacts):
        added = False
        for c in contacts:
            wxid = c.get('wxid')
            if not wxid or wxid in self.selected_wxids:
                continue
            self.selected_contacts.append(c)
            self.selected_wxids.add(wxid)
            added = True
        if added:
            self.update_selected_contacts_list()

    def update_task_table(self):
        """更新任务表格"""
        prev_sort = self.task_table.isSortingEnabled()
        self.task_table.setSortingEnabled(False)
        self.task_table.setRowCount(0)
        
        for i, task in enumerate(self.scheduled_tasks):
            self.task_table.insertRow(i)
            self._populate_task_row(i, task)
            
        self.task_table.setSortingEnabled(prev_sort)
    
    def _populate_task_row(self, row, task):
        self.task_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
        self.task_table.setItem(row, 1, QTableWidgetItem(task['name']))
        
        schedule_time = datetime.fromtimestamp(task['schedule_time'])
        time_str = schedule_time.strftime('%Y-%m-%d %H:%M:%S')
        self.task_table.setItem(row, 2, QTableWidgetItem(time_str))
        
        self.task_table.setItem(row, 3, QTableWidgetItem(str(len(task['contacts']))))
        
        raw_status = task.get('status', '等待中')
        display_status = str(raw_status).strip() if raw_status else '等待中'
        if (display_status == '完成' or display_status.startswith('完成(') or
            display_status == '执行中' or display_status.startswith('执行中(')):
            display_status = '已完成'
            task['status'] = '已完成'
        self.task_table.setItem(row, 4, QTableWidgetItem(display_status))
        
        if task.get('status') in ('等待中', '已完成'):
            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.clicked.connect(lambda: self.edit_task(row))
            
            del_btn = QPushButton("删除")
            del_btn.clicked.connect(lambda: self.delete_task(task['id']))
            
            layout.addWidget(edit_btn)
            layout.addWidget(del_btn)
            layout.addStretch()
            
            self.task_table.setCellWidget(row, 5, cell_widget)

    def refresh_task_table_async(self):
        QTimer.singleShot(0, self.update_task_table)

    def delete_task(self, task_id):
        self.scheduled_tasks = [task for task in self.scheduled_tasks if task['id'] != task_id]
        self.update_task_table()

    def delete_selected_tasks(self):
        rows = sorted({idx.row() for idx in self.task_table.selectedIndexes()}, reverse=True)
        if not rows:
            QMessageBox.information(self, "提示", "请先选择要删除的任务")
            return
        ids = []
        for r in rows:
            if 0 <= r < len(self.scheduled_tasks):
                ids.append(self.scheduled_tasks[r]['id'])
        if not ids:
            return
        self.scheduled_tasks = [t for t in self.scheduled_tasks if t['id'] not in ids]
        self.update_task_table()

    def check_scheduled_tasks(self):
        current_time = int(time.time())
        for task in self.scheduled_tasks:
            if task['status'] == '等待中' and current_time >= task['schedule_time']:
                task['status'] = '执行中'
                self.refresh_task_table_async()
                threading.Thread(target=self.execute_task, args=(task,), daemon=True).start()

    def edit_task(self, row_index: int):
        if row_index < 0 or row_index >= len(self.scheduled_tasks):
            return
        task = self.scheduled_tasks[row_index]
        dialog = QDialog(self)
        dialog.setWindowTitle("编辑任务")
        layout = QVBoxLayout(dialog)
        form = QFormLayout()
        name_edit = QLineEdit(task.get('name', ''))
        form.addRow("任务名称:", name_edit)
        time_edit = QDateTimeEdit()
        time_edit.setDisplayFormat("yyyy年MM月dd日 HH:mm")
        time_edit.setCalendarPopup(True)
        time_edit.setDateTime(QDateTime.fromSecsSinceEpoch(int(task.get('schedule_time', int(time.time())+300))))
        form.addRow("发送时间:", time_edit)
        min_delay_edit = QLineEdit(str(task.get('min_delay', 1)))
        max_delay_edit = QLineEdit(str(task.get('max_delay', 3)))
        delay_row = QHBoxLayout()
        delay_row.addWidget(min_delay_edit)
        delay_row.addWidget(QLabel("至"))
        delay_row.addWidget(max_delay_edit)
        delay_row.addWidget(QLabel("秒"))
        delay_container = QWidget()
        delay_container.setLayout(delay_row)
        form.addRow("对象切换间隔:", delay_container)

        text_edit = QTextEdit(task.get('message_text', ''))
        form.addRow("消息文本:", text_edit)

        img_row = QHBoxLayout()
        img_edit = QLineEdit(task.get('image_path', ''))
        img_btn = QPushButton("选择图片")
        def pick_img():
            path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "Images (*.png *.jpg *.jpeg *.gif *.bmp)")
            if path:
                img_edit.setText(path)
        img_btn.clicked.connect(pick_img)
        img_row.addWidget(img_edit)
        img_row.addWidget(img_btn)
        img_container = QWidget()
        img_container.setLayout(img_row)
        form.addRow("图片/文件/视频路径:", img_container)

        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(btns)

        def on_accept():
            try:
                name = name_edit.text().strip() or task.get('name', '任务')
                dt = time_edit.dateTime()
                if dt <= QDateTime.currentDateTime():
                    dt = QDateTime.currentDateTime().addSecs(300)
                mn = int(min_delay_edit.text()) if min_delay_edit.text().isdigit() else 0
                mx = int(max_delay_edit.text()) if max_delay_edit.text().isdigit() else mn
                if mx < mn:
                    mx = mn

                task['name'] = name
                task['schedule_time'] = int(dt.toSecsSinceEpoch())
                task['min_delay'] = mn
                task['max_delay'] = mx
                task['message_text'] = text_edit.toPlainText()
                task['image_path'] = str(img_edit.text()).strip().strip('"')
                self.update_task_table()
                dialog.accept()
            except Exception as e:
                QMessageBox.warning(dialog, "错误", f"保存失败: {e}")
        btns.accepted.connect(on_accept)
        btns.rejected.connect(dialog.reject)
        dialog.exec()

    def execute_task(self, task):
        """执行定时任务"""
        try:
            contacts = task['contacts']
            message_text = task['message_text']
            image_path = task['image_path']
            min_delay = task['min_delay']
            max_delay = task['max_delay']

            # 获取进程ID
            pid = self._get_task_pid(task)
            if not pid:
                task['status'] = '失败'
                self.refresh_task_table_async()
                return

            # 执行发送
            success_count = self.send_messages(pid, contacts, message_text, image_path, min_delay, max_delay)
            
            task['status'] = '已完成'
            self.refresh_task_table_async()
            
        except Exception:
            task['status'] = '失败'
            self.refresh_task_table_async()
    
    def _get_task_pid(self, task):
        pid = task.get('pid')
        if not pid and self.parent and hasattr(self.parent, 'current_account_pid'):
            pid = getattr(self.parent, 'current_account_pid', None)
        if not pid:
            wechat_info = SimpleWeChatInfo()
            accounts = wechat_info.run()
            if accounts:
                pid = accounts[0]['pid']
        return pid

    def start_mass_send(self, contacts):
        """开始群发消息"""
        message_text = self.message_text.toPlainText()
        image_path = str(self.image_path.text()).strip().strip('"')

        if not message_text.strip() and not image_path.strip():
            return

        pid = self._get_current_pid()
        if not pid:
            return

        try:
            min_delay = int(self.min_delay.text())
            max_delay = int(self.max_delay.text())
        except (ValueError, TypeError):
            min_delay, max_delay = 1, 3

        threading.Thread(
            target=self.send_messages, 
            args=(pid, contacts, message_text, image_path, min_delay, max_delay), 
            daemon=True
        ).start()

    def send_messages(self, pid, contacts, message_text, image_path, min_delay, max_delay):
        """发送消息给多个联系人"""
        success_count = 0
        
        for contact in contacts:
            wxid = contact.get('wxid')
            if not wxid:
                continue
                
            try:
                # 发送文本消息
                if message_text.strip():
                    send_message_simple(pid, wxid, message_text)
                
                # 发送图片或文件
                if image_path.strip() and os.path.exists(image_path):
                    send_image_simple(pid, wxid, image_path)
                
                success_count += 1
                
                # 随机延迟
                if min_delay > 0 or max_delay > 0:
                    delay = random.randint(min_delay, max_delay)
                    time.sleep(delay)
                    
            except Exception:
                continue
                
        return success_count


class DataManager:
    def __init__(self):
        self.messages_file = os.path.join("config", "messages.ini")
        os.makedirs("config", exist_ok=True)
        
        self.message_config = configparser.ConfigParser()
        if os.path.exists(self.messages_file):
            self.message_config.read(self.messages_file, encoding='utf-8')
        
        self.account_data_cache = {}

    def save_account_data(self, account_info, contacts, friends, groups):
        wxid = account_info.get('wxid')
        if not wxid:
            return False

        self.account_data_cache[wxid] = {
            'account_info': account_info,
            'contacts': contacts,
            'friends': friends,
            'groups': groups,
            'last_update': int(time.time())
        }
        return True

    def update_account_remark(self, wxid, friend_wxid, new_remark):
        account_data = self.account_data_cache.get(wxid)
        if not account_data:
            return False
            
        for friend in account_data['friends']:
            if friend.get('wxid') == friend_wxid:
                friend['remarks'] = new_remark
                break

        for contact in account_data['contacts']:
            if contact.get('wxid') == friend_wxid:
                contact['remarks'] = new_remark
                break

        return True

    def load_account_data(self, wxid):
        return self.account_data_cache.get(wxid)

    def save_message(self, message):
        try:
            message_id = str(uuid.uuid4())
            timestamp = message.get('timestamp', int(time.time()))
            time_str = datetime.fromtimestamp(timestamp).strftime('%Y%m%d%H%M%S')

            if 'Messages' not in self.message_config:
                self.message_config['Messages'] = {}

            self.message_config['Messages'][message_id] = time_str
            message_section = f"Message_{message_id}"
            
            self.message_config[message_section] = {
                'timestamp': str(timestamp),
                'wxid': message.get('wxid', ''),
                'content': message.get('content', ''),
                'account_wxid': message.get('account', {}).get('wxid', ''),
                'account_nickname': message.get('account', {}).get('nickname', ''),
                'member_id': message.get('member_id', '')
            }

            with open(self.messages_file, 'w', encoding='utf-8') as f:
                self.message_config.write(f)
            return True
            
        except Exception:
            return False

    def load_messages(self, limit=0):
        if not os.path.exists(self.messages_file):
            return []

        try:
            self.message_config = configparser.ConfigParser()
            self.message_config.read(self.messages_file, encoding='utf-8')

            if 'Messages' not in self.message_config:
                return []

            message_ids = [(msg_id, time_str) for msg_id, time_str in self.message_config['Messages'].items()]
            message_ids.sort(key=lambda x: x[1], reverse=True)

            if limit > 0:
                message_ids = message_ids[:limit]

            messages = []
            for message_id, _ in message_ids:
                section_name = f"Message_{message_id}"
                if section_name in self.message_config:
                    message = self._parse_message_section(section_name)
                    if message:
                        messages.append(message)

            return messages
        except Exception:
            return []
    
    def _parse_message_section(self, section_name):
        try:
            section = self.message_config[section_name]
            message = {
                'timestamp': int(section['timestamp']),
                'wxid': section['wxid'],
                'content': section['content'],
                'account': {
                    'wxid': section['account_wxid'],
                    'nickname': section['account_nickname']
                }
            }
            
            if section.get('member_id'):
                message['member_id'] = section['member_id']
                
            return message
        except (KeyError, ValueError):
            return None

    def cleanup_old_messages(self, max_days=30):
        if not os.path.exists(self.messages_file):
            return 0

        try:
            cutoff_time = int(time.time()) - (max_days * 24 * 60 * 60)
            
            if 'Messages' not in self.message_config:
                return 0

            message_ids = list(self.message_config['Messages'].keys())
            deleted_count = 0

            for message_id in message_ids:
                section_name = f"Message_{message_id}"
                if section_name in self.message_config:
                    timestamp = int(self.message_config[section_name]['timestamp'])
                    if timestamp < cutoff_time:
                        self.message_config.remove_section(section_name)
                        del self.message_config['Messages'][message_id]
                        deleted_count += 1

            if deleted_count > 0:
                with open(self.messages_file, 'w', encoding='utf-8') as f:
                    self.message_config.write(f)

            return deleted_count
        except Exception:
            return 0

    def load_all_accounts(self):
        return self.account_data_cache.copy()

class MessageReceiver(QObject):
    message_received = Signal(dict)

    def __init__(self):
        super().__init__()
        self.use_chinese = True

class MessageMonitorManager:
    def __init__(self, message_receiver):
        self.message_receiver = message_receiver
        self.monitors = {}
        self.is_running = False

    def start_monitor_for_account(self, account):
        pid = account['pid']

        if pid in self.monitors:
            return

        try:
            monitor = WeChatMessageMonitor(pid)

            def message_callback(msg_data):
                msg_data['account'] = {
                    'nickname': account['nickname'],
                    'wxid': account['wxid'],
                    'pid': pid
                }

                self.message_receiver.message_received.emit(msg_data)

            monitor.set_callback(message_callback)

            monitor.start()

            self.monitors[pid] = monitor

            return True
        except Exception as e:
            return False

    def start_monitor_all(self):
        if self.is_running:
            return

        wechat_info = SimpleWeChatInfo()
        accounts = wechat_info.run()

        if not accounts:
            return

        success_count = 0
        for account in accounts:
            if self.start_monitor_for_account(account):
                success_count += 1

        if success_count > 0:
            self.is_running = True
        else:
            pass
    def stop_monitor_all(self):
        if not self.is_running:
            return

        for pid, monitor in list(self.monitors.items()):
            try:
                monitor.stop()
            except Exception as e:
                pass
        self.monitors.clear()
        self.is_running = False
    def get_contact_name(self, wxid):
        return wxid

class SendMessageDialog(QDialog):
    def __init__(self, friend_name, parent=None, wxid=None, pid=None):
        super().__init__(parent)
        self.friend_name = friend_name
        self.wxid = wxid
        self.pid = pid
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("发送消息")
        self.resize(400, 250)

        layout = QVBoxLayout(self)

        id_layout = QHBoxLayout()
        id_label = QLabel("当前选中的微信ID:")
        id_layout.addWidget(id_label)

        id_value = QLabel(self.wxid if self.wxid else "未知")
        id_layout.addWidget(id_value)

        layout.addLayout(id_layout)

        name_layout = QHBoxLayout()
        name_label = QLabel("好友昵称:")
        name_layout.addWidget(name_label)

        name_value = QLabel(self.friend_name)
        name_layout.addWidget(name_value)
        layout.addLayout(name_layout)
        tab_widget = QTabWidget()
        layout.addWidget(tab_widget)

        text_tab = QWidget()
        text_layout = QVBoxLayout(text_tab)

        content_label = QLabel("需要发送的内容:")
        text_layout.addWidget(content_label)
        self.content_edit = QTextEdit()
        text_layout.addWidget(self.content_edit)
        text_send_button = QPushButton("发送文本消息")
        text_send_button.clicked.connect(self.send_text)
        text_layout.addWidget(text_send_button)

        image_tab = QWidget()
        image_layout = QVBoxLayout(image_tab)

        image_path_label = QLabel("图片/文件/视频路径:")
        image_layout.addWidget(image_path_label)

        image_path_layout = QHBoxLayout()

        self.image_path_edit = QTextEdit()
        self.image_path_edit.setMinimumHeight(self.content_edit.sizeHint().height())
        image_path_layout.addWidget(self.image_path_edit)

        select_image_button = QPushButton("选择路径")
        select_image_button.clicked.connect(self.select_image_file)
        image_path_layout.addWidget(select_image_button)

        image_layout.addLayout(image_path_layout)

        image_send_button = QPushButton("发送图片/文件/视频消息")
        image_send_button.clicked.connect(self.send_image)
        image_layout.addWidget(image_send_button)

        tab_widget.addTab(text_tab, "发送文本")
        tab_widget.addTab(image_tab, "发送图片/文件/视频")

        button_layout = QHBoxLayout()

        cancel_button = QPushButton("取消")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)

    def select_image_file(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        file_dialog.setNameFilter(
            "图片文件 (*.jpg *.jpeg *.png *.bmp *.gif *.pdf)")

        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.image_path_edit.setPlainText(selected_files[0])


    def send_text(self):
        content = self.content_edit.toPlainText()
        if not content.strip():
            QMessageBox.warning(self, "警告", "消息内容不能为空", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        parent = self.parent()
        if parent and isinstance(parent, WeChatManagerApp):
            success = parent.send_message_to_wxid(self.wxid, content, pid=self.pid)
            if success:
                self.accept()
            else:
                QMessageBox.warning(self, "发送失败", "消息发送失败，请检查微信状态或重试", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        else:
            QMessageBox.warning(self, "发送失败", "无法获取发送接口", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def send_image(self):
        image_path = self.image_path_edit.toPlainText()

        try:
            p = str(image_path).strip()
            if p.startswith('"') and p.endswith('"') and len(p) >= 2:
                p = p[1:-1]
            else:
                p = p.strip('"')
        except Exception:
            p = image_path
        if not p.strip():
            QMessageBox.warning(self, "警告", "请选择要发送的图片", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        if not os.path.isfile(p):
            QMessageBox.warning(self, "警告", f"图片文件不存在: {p}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        parent = self.parent()
        if parent and isinstance(parent, WeChatManagerApp):
            success = parent.send_image_to_wxid(self.wxid, p, pid=self.pid)
            if success:
                self.accept()
            else:
                QMessageBox.warning(self, "发送失败", "图片发送失败，请检查微信状态或重试", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        else:
            QMessageBox.warning(self, "发送失败", "无法获取发送接口", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

class WeChatManagerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AI全功能营销系统")
        self.resize(1000, 800)

        self.data_manager = DataManager()
        self.wechat_service = get_wechat_service()
        self.wechat_info = SimpleWeChatInfo()  # 初始化wechat_info属性

        self.message_receiver = MessageReceiver()
        self.monitor_manager = MessageMonitorManager(self.message_receiver)
        self.message_receiver.message_received.connect(self.on_message_received)

        self.notebook = QTabWidget()
        self.setCentralWidget(self.notebook)

        self.account_tab = QWidget()
        self.notebook.addTab(self.account_tab, "主界面")

        self.add_friend_tab = QWidget()
        self.notebook.addTab(self.add_friend_tab, "添加好友")

        self.auto_reply_tab = QWidget()
        self.notebook.addTab(self.auto_reply_tab, "自动回复")

        self.task_tab = TaskTab(self)
        self.notebook.addTab(self.task_tab, "定时群发")

        self.ai_assistant_tab = AIAssistantTab()
        self.notebook.addTab(self.ai_assistant_tab, "AI助手")

        self.friend_id_input = QLineEdit()
        self.greeting_input = QLineEdit()
        self.scene_combo = QComboBox()
        self.add_friend_result = QTextEdit()
        self.friend_tree = QTreeWidget()
        self.group_tree = QTreeWidget()
        self.members_tree = QTreeWidget()
        self.friend_count_label = QLabel("好友总数: 0")
        self.group_count_label = QLabel("群总数: 0")
        self.members_count_label = QLabel("成员总数: 0")
        self.account_tree = QTreeWidget()
        self.search_entry = QLineEdit()
        self.group_search_entry = QLineEdit()
        self.detect_button = QPushButton("检测已登录微信")
        self.open_button = QPushButton("打开新微信")

        self.rules_data = []
        self.auto_reply_message_counter = 0
        self.auto_reply_history_data = []
        self.data_changed = False
        self.data_save_timer = QTimer()
        self.data_save_timer.setSingleShot(True)
        self.data_save_timer.timeout.connect(self.save_rules_data)

        self.opening_wechat = False

        self.wechat_pid = None

        self.specific_friend_wxids = set()
        self.specific_group_wxids = set()

        self.init_account_tab()
        self.init_add_friend_tab()
        self.init_auto_reply_tab()

        self.all_contacts = []

        self.statusBar().showMessage("正在初始化...", 3000)

        QTimer.singleShot(500, self.detect_wechat_accounts)

        QTimer.singleShot(1000, self.start_message_monitoring)

        QTimer.singleShot(2500, self.load_rules_data)

        self.startup_timestamp = int(time.time())

    def closeEvent(self, event):
        """确保应用关闭时干净地停止所有监控、线程与定时器。"""
        try:
            try:
                if hasattr(self, 'monitor_manager') and self.monitor_manager:
                    try:
                        if hasattr(self.monitor_manager, 'is_running'):
                            self.monitor_manager.is_running = True
                    except Exception:
                        pass
                    self.monitor_manager.stop_monitor_all()
            except Exception as e:
                pass
            try:
                contact_monitors = getattr(self, 'contact_monitors', {}) or {}
                for pid, mon in list(contact_monitors.items()):
                    try:
                        if mon and hasattr(mon, 'is_active') and mon.is_active():
                            mon.stop()
                    except Exception as ie:
                        pass
                if isinstance(contact_monitors, dict):
                    contact_monitors.clear()
            except Exception as e:
                pass
            # 3) 停止所有定时器（包含 monitor_check_timer、data_save_timer、任务定时器等）
            try:
                for name in ('monitor_check_timer', 'data_save_timer'):
                    t = getattr(self, name, None)
                    if t:
                        try:
                            t.stop()
                            t.deleteLater()
                        except Exception:
                            pass
                try:
                    if hasattr(self, 'notebook'):
                        for i in range(self.notebook.count()):
                            w = self.notebook.widget(i)
                            if hasattr(w, 'task_timer') and getattr(w, 'task_timer'):
                                try:
                                    w.task_timer.stop()
                                    w.task_timer.deleteLater()
                                except Exception:
                                    pass
                except Exception:
                    pass
                for t in self.findChildren(QTimer):
                    try:
                        t.stop()
                        t.deleteLater()
                    except Exception:
                        pass
            except Exception as e:
                pass
            try:
                if hasattr(self, 'is_running'):
                    self.is_running = False
                if hasattr(self, 'is_paused'):
                    self.is_paused = True
            except Exception:
                pass
                pass
        except Exception as e:
            pass
        finally:
            event.accept()

    def init_account_tab(self):
        layout = QVBoxLayout(self.account_tab)
        layout.setSpacing(0)
        top_layout = QHBoxLayout()

        account_group = QWidget()
        account_layout = QVBoxLayout(account_group)
        account_layout.setContentsMargins(2, 2, 2, 2)

        self.account_tree.setHeaderLabels(["序号", "昵称:双击列表获取联系人", "微信ID", "手机号"])
        self.account_tree.setAlternatingRowColors(True)
        self.account_tree.setStyleSheet(StyleSheet.TREE)

        setup_tree_columns(
            self.account_tree,
            specs=[
                (QHeaderView.ResizeMode.Interactive, 50),    # 序号
                (QHeaderView.ResizeMode.Interactive, 160),  #  昵称
                (QHeaderView.ResizeMode.Interactive, 150),  # 微信ID
                (QHeaderView.ResizeMode.Interactive, 150),  # 手机号
            ],

        )

        self.account_tree.itemDoubleClicked.connect(
            self.on_account_double_click)

        account_layout.addWidget(self.account_tree)

        button_group = QWidget()
        button_layout = QVBoxLayout(button_group)

        button_row1 = QHBoxLayout()

        self.detect_button.clicked.connect(self.detect_wechat_accounts)
        button_row1.addWidget(self.detect_button)

        self.open_button.clicked.connect(self.open_new_wechat_instance)
        button_row1.addWidget(self.open_button)

        button_layout.addLayout(button_row1)

        button_row2 = QHBoxLayout()

        self.check_zombie_fans_button = QPushButton("检测僵尸粉")
        self.check_zombie_fans_button.clicked.connect(self.check_zombie_fans)
        button_row2.addWidget(self.check_zombie_fans_button)

        self.contact_service_button = QPushButton("联系客服")
        self.contact_service_button.clicked.connect(self.show_contact_service_dialog)
        button_row2.addWidget(self.contact_service_button)

        button_layout.addLayout(button_row2)

        top_layout.addWidget(account_group, 3)
        top_layout.addWidget(button_group, 1)

        friend_group = QWidget()
        friend_layout = QVBoxLayout(friend_group)
        friend_layout.setContentsMargins(2, 2, 2, 2)

        search_layout = QHBoxLayout()

        search_layout.addWidget(self.friend_count_label)

        search_label = QLabel("搜索好友:")
        self.search_entry.textChanged.connect(self.search_friends)
        search_button = QPushButton("搜索")
        search_button.clicked.connect(self.search_friends)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_entry)
        search_layout.addWidget(search_button)

        export_button = QPushButton("导出联系人信息")
        export_button.clicked.connect(self.export_contacts)
        search_layout.addWidget(export_button)

        search_layout.addStretch(1)
        friend_layout.addLayout(search_layout)

        self.friend_tree.setHeaderLabels(
        ["序号", "好友昵称", "微信ID", "备注：列表鼠标右键有功能", "标签", "手机号"])
        setup_tree_columns(
            self.friend_tree,
            specs=[
                (QHeaderView.ResizeMode.Interactive, 50),    # 序号
                (QHeaderView.ResizeMode.Interactive, 160),  # 好友昵称
                (QHeaderView.ResizeMode.Interactive, 150),  # 微信ID
                (QHeaderView.ResizeMode.Interactive, 150),  # 备注
                (QHeaderView.ResizeMode.Interactive, 100),  # 标签
                (QHeaderView.ResizeMode.Interactive, 120),  # 手机号
            ],
            default_size=150,
            stretch_last=True,
        )

        self.friend_tree.setAlternatingRowColors(True)

        self.friend_tree.setSelectionMode(QTreeWidget.SelectionMode.ExtendedSelection)

        self.friend_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.friend_tree.customContextMenuRequested.connect(
            self.show_friend_context_menu)

        friend_layout.addWidget(self.friend_tree)

        groups_layout = QHBoxLayout()

        group_group = QWidget()
        group_layout = QVBoxLayout(group_group)
        group_layout.setContentsMargins(2, 2, 2, 2)

        group_search_layout = QHBoxLayout()

        group_search_layout.addWidget(self.group_count_label)

        group_search_label = QLabel("搜索群聊:")
        self.group_search_entry.textChanged.connect(self.search_groups)
        group_search_button = QPushButton("搜索")
        group_search_button.clicked.connect(self.search_groups)
        group_search_layout.addWidget(group_search_label)
        group_search_layout.addWidget(self.group_search_entry)
        group_search_layout.addWidget(group_search_button)

        group_search_layout.addStretch(1)
        group_layout.addLayout(group_search_layout)

        self.group_tree.setHeaderLabels(["序号", "群聊名称", "群聊ID：列表鼠标右键有功能"])
        setup_tree_columns(
            self.group_tree,
            specs=[],
            default_size=150,
            stretch_last=True,
        )

        self.group_tree.setAlternatingRowColors(True)
        self.group_tree.setStyleSheet(StyleSheet.TREE)

        self.group_tree.setSelectionMode(QTreeWidget.SelectionMode.ExtendedSelection)

        self.group_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.group_tree.customContextMenuRequested.connect(
            self.show_group_context_menu)

        group_layout.addWidget(self.group_tree)

        members_group = QWidget()
        members_layout = QVBoxLayout(members_group)
        members_layout.setContentsMargins(2, 2, 2, 2)

        members_layout.addWidget(self.members_count_label)

        self.members_tree.setHeaderLabels(["序号", "群聊名称", "群成员昵称", "群成员ID"])

        setup_tree_columns(
            self.members_tree,
            specs=[],
            default_size=150,
            stretch_last=True,
        )

        self.members_tree.setAlternatingRowColors(True)
        self.members_tree.setStyleSheet(StyleSheet.TREE)

        self.members_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.members_tree.customContextMenuRequested.connect(
            self.show_members_context_menu)

        members_layout.addWidget(self.members_tree)

        groups_layout.setSpacing(2)
        groups_layout.addWidget(group_group, 1)
        groups_layout.addWidget(members_group, 1)

        layout.addLayout(top_layout, 2)
        layout.addWidget(friend_group, 4)
        layout.addLayout(groups_layout, 4)

    def init_add_friend_tab(self):
        add_friend_layout = QVBoxLayout(self.add_friend_tab)

        controls_layout = QHBoxLayout()

        delay_layout = QHBoxLayout()
        delay_label = QLabel("延时(分钟):")
        delay_label.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {StyleSheet.TEXT_SECONDARY};")
        self.min_delay_input = QLineEdit("5")
        self.min_delay_input.setMaximumWidth(60)
        self.min_delay_input.setStyleSheet(StyleSheet.LINE_EDIT)
        delay_to_label = QLabel("至")
        delay_to_label.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {StyleSheet.TEXT_SECONDARY};")
        self.max_delay_input = QLineEdit("10")
        self.max_delay_input.setMaximumWidth(60)
        self.max_delay_input.setStyleSheet(StyleSheet.LINE_EDIT)

        delay_layout.addWidget(delay_label)
        delay_layout.addWidget(self.min_delay_input)
        delay_layout.addWidget(delay_to_label)
        delay_layout.addWidget(self.max_delay_input)

        freq_label = QLabel("检测到添加频繁自动停止")
        freq_label.setStyleSheet(f"font-size: 14px; color: {StyleSheet.ERROR_COLOR};")
        delay_layout.addWidget(freq_label)

        controls_layout.addLayout(delay_layout)

        controls_layout.addStretch(1)

        add_friend_layout.addLayout(controls_layout)

        buttons_main_layout = QVBoxLayout()
        buttons_main_layout.setAlignment(Qt.AlignLeft)

        buttons_layout1 = QHBoxLayout()
        buttons_layout1.setAlignment(Qt.AlignLeft)

        self.import_button = QPushButton("导入手机号添加")
        self.import_button.clicked.connect(self.import_from_xls)
        buttons_layout1.addWidget(self.import_button)

        self.add_group_members_button = QPushButton("添加群成员添加")
        self.add_group_members_button.clicked.connect(self.add_group_members)
        buttons_layout1.addWidget(self.add_group_members_button)

        self.export_button = QPushButton("导出数据")
        self.export_button.clicked.connect(self.export_data)
        buttons_layout1.addWidget(self.export_button)

        self.clear_button = QPushButton("清空列表")
        self.clear_button.clicked.connect(self.clear_table)
        buttons_layout1.addWidget(self.clear_button)

        buttons_main_layout.addLayout(buttons_layout1)

        buttons_layout2 = QHBoxLayout()
        buttons_layout2.setAlignment(Qt.AlignLeft)

        self.start_button = QPushButton("开始")
        self.start_button.clicked.connect(self.start_process)
        buttons_layout2.addWidget(self.start_button)

        self.pause_button = QPushButton("暂停")
        self.pause_button.clicked.connect(self.pause_process)
        self.pause_button.setEnabled(False)
        buttons_layout2.addWidget(self.pause_button)

        self.resume_button = QPushButton("继续")
        self.resume_button.clicked.connect(self.resume_process)
        self.resume_button.setEnabled(False)
        buttons_layout2.addWidget(self.resume_button)

        self.stop_button = QPushButton("结束")
        self.stop_button.clicked.connect(self.stop_process)
        self.stop_button.setEnabled(False)
        buttons_layout2.addWidget(self.stop_button)

        buttons_main_layout.addLayout(buttons_layout2)

        add_friend_layout.addLayout(buttons_main_layout)

        self.add_friend_table = QTableWidget()
        self.add_friend_table.setColumnCount(7)
        self.add_friend_table.setHorizontalHeaderLabels(["序号", "手机号", "招呼语", "状态/微信号", "v3信息", "联系人昵称", "备注"])
        self.add_friend_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.add_friend_table.setSelectionMode(QTableWidget.SingleSelection)

        self.add_friend_table.verticalHeader().setVisible(False)

        try:
            self.add_friend_table.itemChanged.disconnect()
        except Exception:
            pass
        self.add_friend_table.itemChanged.connect(lambda *_: self.save_add_friend_data())
        self.add_friend_table.setStyleSheet(StyleSheet.TABLE)

        header = self.add_friend_table.horizontalHeader()
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        for i in range(self.add_friend_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.resizeSection(0, 50)
        for i in range(1, self.add_friend_table.columnCount()):
            header.resizeSection(i, 150)

        add_friend_layout.addWidget(self.add_friend_table)

        self.add_friend_status = QLabel("准备就绪")
        add_friend_layout.addWidget(self.add_friend_status)

        self.contact_monitors = {}
        self.is_running = False
        self.is_paused = False
        self.phone_queue = []
        self.current_index = 0

        try:
            import sys as _sys_for_path
        except Exception:
            _sys_for_path = sys
        if getattr(_sys_for_path, 'frozen', False):
            base_dir = os.path.dirname(_sys_for_path.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        config_dir = os.path.join(base_dir, "config")
        os.makedirs(config_dir, exist_ok=True)
        self.data_file = os.path.join(config_dir, "wechat_tool_data.json")

        self.load_add_friend_data()

        QTimer.singleShot(500, self.start_monitoring)


    def _init_simple_add_friend_tab(self):
        layout = QVBoxLayout(self.add_friend_tab)

        add_friend_group = QGroupBox("添加好友")
        add_friend_layout = QVBoxLayout(add_friend_group)

        id_layout = QHBoxLayout()
        id_label = QLabel("好友ID/手机号:")
        self.friend_id_input.setPlaceholderText("输入好友的微信ID或手机号")
        id_layout.addWidget(id_label)
        id_layout.addWidget(self.friend_id_input)
        add_friend_layout.addLayout(id_layout)

        greeting_layout = QHBoxLayout()
        greeting_label = QLabel("招呼语:")
        self.greeting_input.setPlaceholderText("输入添加好友时的招呼语")
        self.greeting_input.setText("您好，我想添加您为好友")
        greeting_layout.addWidget(greeting_label)
        greeting_layout.addWidget(self.greeting_input)
        add_friend_layout.addLayout(greeting_layout)

        add_button_layout = QHBoxLayout()
        add_button = QPushButton("添加好友")
        add_button.clicked.connect(self.add_friend)
        add_button_layout.addWidget(add_button)
        add_friend_layout.addLayout(add_button_layout)

        self.add_friend_result.setReadOnly(True)
        add_friend_layout.addWidget(self.add_friend_result)

        layout.addWidget(add_friend_group)

    def init_auto_reply_tab(self):
        layout = QVBoxLayout(self.auto_reply_tab)

        switch_group = QGroupBox("功能开关")
        switch_layout = QVBoxLayout()

        rule_reply_layout = QHBoxLayout()
        self.rule_reply_switch = QCheckBox("启用规则回复：")
        self.rule_reply_switch.stateChanged.connect(self.on_rule_reply_switch)
        rule_reply_layout.addWidget(self.rule_reply_switch)

        self.reply_friend_switch = QCheckBox("回复所有联系人")
        self.reply_friend_switch.stateChanged.connect(self.on_reply_friend_switch)
        rule_reply_layout.addWidget(self.reply_friend_switch)

        self.reply_group_switch = QCheckBox("回复所有群消息")
        self.reply_group_switch.stateChanged.connect(self.on_reply_group_switch)
        rule_reply_layout.addWidget(self.reply_group_switch)

        self.specific_friend_switch = QCheckBox("指定好友回复")
        self.specific_friend_switch.stateChanged.connect(self.on_specific_friend_switch)
        rule_reply_layout.addWidget(self.specific_friend_switch)

        self.specify_friend_btn = QPushButton("选择指定好友")
        self.specify_friend_btn.clicked.connect(self.on_click_specify_friend)
        rule_reply_layout.addWidget(self.specify_friend_btn)

        self.specific_group_switch = QCheckBox("指定群回复")
        self.specific_group_switch.stateChanged.connect(self.on_specific_group_switch)
        rule_reply_layout.addWidget(self.specific_group_switch)

        self.specify_group_btn = QPushButton("选择指定群")
        self.specify_group_btn.clicked.connect(self.on_click_specify_group)
        rule_reply_layout.addWidget(self.specify_group_btn)

        rule_reply_layout.addStretch()
        switch_layout.addLayout(rule_reply_layout)

        friend_box = QGroupBox()
        friend_box.setTitle("")
        friend_v = QVBoxLayout()
        self.specific_friend_list = QListWidget()
        self.specific_friend_list.itemDoubleClicked.connect(self.remove_selected_specific_friend)
        friend_v.addWidget(self.specific_friend_list)
        friend_btn_row = QHBoxLayout()
        clear_friend_btn = QPushButton("清空好友/列表双击选中好友删除")
        clear_friend_btn.clicked.connect(self.clear_specific_friends)
        friend_btn_row.addWidget(clear_friend_btn)
        friend_btn_row.addStretch()
        friend_v.addLayout(friend_btn_row)
        friend_box.setLayout(friend_v)

        group_box = QGroupBox()
        group_box.setTitle("")
        group_v = QVBoxLayout()
        self.specific_group_list = QListWidget()
        self.specific_group_list.itemDoubleClicked.connect(self.remove_selected_specific_group)
        group_v.addWidget(self.specific_group_list)
        group_btn_row = QHBoxLayout()
        clear_group_btn = QPushButton("清空群/列表双击选中群删除")
        clear_group_btn.clicked.connect(self.clear_specific_groups)
        group_btn_row.addWidget(clear_group_btn)
        group_btn_row.addStretch()
        group_v.addLayout(group_btn_row)
        group_box.setLayout(group_v)

        selected_layout = QHBoxLayout()
        selected_layout.addWidget(friend_box)
        selected_layout.addWidget(group_box)
        switch_layout.addLayout(selected_layout)

        ai_reply_layout = QHBoxLayout()
        self.ai_reply_switch = QCheckBox("启用AI回复")
        self.ai_reply_switch.setChecked(False)
        self.ai_reply_switch.stateChanged.connect(self.on_ai_reply_switch)
        ai_reply_layout.addWidget(self.ai_reply_switch)

        self.yuanbao_reply_switch = QCheckBox("元宝客服回复")
        self.yuanbao_reply_switch.stateChanged.connect(self.on_yuanbao_reply_switch)
        self.yuanbao_reply_switch.setEnabled(False)
        ai_reply_layout.addWidget(self.yuanbao_reply_switch)

        self.model_reply_switch = QCheckBox("大模型回复")
        self.model_reply_switch.stateChanged.connect(self.on_model_reply_switch)
        self.model_reply_switch.setEnabled(False)
        ai_reply_layout.addWidget(self.model_reply_switch)

        ai_reply_layout.addStretch()
        switch_layout.addLayout(ai_reply_layout)

        new_friend_layout = QHBoxLayout()
        self.new_friend_reply_switch = QCheckBox("启用新好友回复")
        self.new_friend_reply_switch.stateChanged.connect(self.on_new_friend_reply_switch)
        new_friend_layout.addWidget(self.new_friend_reply_switch)
        new_friend_layout.addStretch()
        switch_layout.addLayout(new_friend_layout)

        match_layout = QHBoxLayout()
        match_layout.addWidget(QLabel("匹配类型："))

        self.fuzzy_match_switch = QCheckBox("模糊匹配")
        self.fuzzy_match_switch.stateChanged.connect(self.on_fuzzy_match_switch)
        match_layout.addWidget(self.fuzzy_match_switch)

        self.exact_match_switch = QCheckBox("精准匹配")
        self.exact_match_switch.stateChanged.connect(self.on_exact_match_switch)
        match_layout.addWidget(self.exact_match_switch)

        match_layout.addStretch()
        switch_layout.addLayout(match_layout)

        interval_layout = QHBoxLayout()
        interval_layout.addWidget(QLabel("收到回复间隔时间："))

        self.min_interval = QLineEdit("2")
        self.min_interval.setFixedWidth(25)
        self.min_interval.setMaximumHeight(25)
        interval_layout.addWidget(self.min_interval)

        interval_layout.addWidget(QLabel("秒到"))

        self.max_interval = QLineEdit("5")
        self.max_interval.setFixedWidth(25)
        self.max_interval.setMaximumHeight(25)
        interval_layout.addWidget(self.max_interval)

        interval_layout.addWidget(QLabel("秒"))
        interval_layout.addStretch()
        switch_layout.addLayout(interval_layout)

        switch_group.setLayout(switch_layout)
        layout.addWidget(switch_group)

        rules_group = QGroupBox("自动回复规则")
        rules_layout = QVBoxLayout()

        self.rules_table = QTableWidget()
        self.rules_table.setColumnCount(4)
        self.rules_table.setHorizontalHeaderLabels([
            "序号", "启用", "关键词", "回复内容"
        ])

        self.rules_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.rules_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.rules_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.rules_table.verticalHeader().setVisible(False)
        self.rules_table.setAlternatingRowColors(True)

        self.rules_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.rules_table.customContextMenuRequested.connect(self.show_context_menu)

        header = self.rules_table.horizontalHeader()
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        for i in range(self.rules_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.resizeSection(0, 50)
        header.resizeSection(1, 50)  
        for i in range(2, self.rules_table.columnCount()):
            header.resizeSection(i, 150)

        self.rules_table.itemChanged.connect(self.on_data_changed)

        rules_layout.addWidget(self.rules_table)

        button_layout = QHBoxLayout()
        add_rule_btn = QPushButton("添加规则")
        import_rules_btn = QPushButton("导入规则")
        export_rules_btn = QPushButton("导出规则")
        ai_reply_rules_btn = QPushButton("大模型回复规则")
        doc_training_btn = QPushButton("上传文档资料大模型训练回复话术（开发中）")

        add_rule_btn.clicked.connect(self.add_reply_rule)
        import_rules_btn.clicked.connect(self.import_rules)
        export_rules_btn.clicked.connect(self.export_rules)
        ai_reply_rules_btn.clicked.connect(self.show_ai_reply_settings)
        doc_training_btn.clicked.connect(self.show_doc_training_dialog)

        button_layout.addWidget(add_rule_btn)
        button_layout.addWidget(import_rules_btn)
        button_layout.addWidget(export_rules_btn)
        button_layout.addWidget(ai_reply_rules_btn)
        button_layout.addWidget(doc_training_btn)

        button_layout.addStretch()

        rules_layout.addLayout(button_layout)
        rules_group.setLayout(rules_layout)
        layout.addWidget(rules_group)

        self.update_specific_selected_lists()

        history_group = QGroupBox("接收消息")
        history_layout = QVBoxLayout()

        self.auto_reply_history_table = QTableWidget()
        self.auto_reply_history_table.setColumnCount(6)
        self.auto_reply_history_table.setHorizontalHeaderLabels([
            "序号", "本微信昵称", "好友/群昵称", "好友ID", "消息内容", "接收时间"
        ])

        header = self.auto_reply_history_table.horizontalHeader()
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        for i in range(self.auto_reply_history_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.resizeSection(0, 50)
        for i in range(1, self.auto_reply_history_table.columnCount()):
            header.resizeSection(i, 150)

        self.auto_reply_history_table.verticalHeader().setVisible(False)
        self.auto_reply_history_table.setAlternatingRowColors(True)

        history_layout.addWidget(self.auto_reply_history_table)

        history_btn_layout = QHBoxLayout()

        load_history_btn = QPushButton("获取历史信息")
        load_history_btn.clicked.connect(self.show_load_history_dialog)
        history_btn_layout.addWidget(load_history_btn)

        clear_btn = QPushButton("清空记录")
        export_btn = QPushButton("导出记录")

        clear_btn.clicked.connect(self.clear_reply_history)
        export_btn.clicked.connect(self.export_reply_history)

        history_btn_layout.addWidget(clear_btn)
        history_btn_layout.addWidget(export_btn)

        history_btn_layout.addStretch()
        history_layout.addLayout(history_btn_layout)

        history_group.setLayout(history_layout)
        layout.addWidget(history_group)

    def detect_wechat_accounts(self):
        self.detect_button.setEnabled(False)

        try:
            self.account_tree.clear()
            accounts = self.wechat_service.get_all_accounts(force_refresh=True)

            if accounts:
                for i, account in enumerate(accounts):
                    item = QTreeWidgetItem([
                        str(i + 1),
                        account["nickname"],
                        account["wxid"],
                        account["phone"]
                    ])
                    self.account_tree.addTopLevelItem(item)
            else:
                QMessageBox.information(self, "提示", "未能获取到微信账号信息", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"检测微信账号失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        finally:
            self.detect_button.setEnabled(True)

    def open_new_wechat_instance(self):
        if self.opening_wechat:
            QMessageBox.information(self, "提示", "正在启动微信，请稍候...", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        try:
            self.opening_wechat = True
            self.open_button.setEnabled(False)
            self.open_button.setText("正在启动...")

            start_new_wechat()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开新微信失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        finally:
            self.opening_wechat = False
            self.open_button.setEnabled(True)
            self.open_button.setText("打开新微信")

    def search_friends(self):
        search_text = self.search_entry.text().lower()

        self.friend_tree.clear()

        if not hasattr(self, 'all_contacts') or not self.all_contacts:
            return

        friends = [
            contact for contact in self.all_contacts if "@chatroom" not in contact["wxid"]]

        if not search_text:
            for i, friend in enumerate(friends, 1):
                phone = friend.get("phone", "")
                if phone == "未知":
                    phone = ""

                item = QTreeWidgetItem([
                    str(i),
                    friend["nickname"],
                    friend["wxid"],
                    friend.get("remarks", ""),
                    friend.get("tag", ""),
                    phone
                ])
                self.friend_tree.addTopLevelItem(item)

            self.friend_count_label.setText(f"好友总数: {len(friends)}")
            return

        matched_friends = []
        for friend in friends:
            nickname = friend["nickname"].lower() if friend.get("nickname") else ""
            remarks = friend.get("remarks", "").lower() if friend.get("remarks") else ""
            tag = friend.get("tag", "").lower() if friend.get("tag") else ""
            phone = friend.get("phone", "").lower() if friend.get("phone") else ""
            if (search_text in nickname or search_text in remarks or
                search_text in tag or search_text in phone):
                    pass
                    matched_friends.append(friend)

        for i, friend in enumerate(matched_friends, 1):
            phone = friend.get("phone", "")
            if phone == "未知":
                phone = ""
            item = QTreeWidgetItem([
                str(i),
                friend["nickname"],
                friend["wxid"],
                friend.get("remarks", ""),
                friend.get("tag", ""),
                phone
            ])
            self.friend_tree.addTopLevelItem(item)

        self.friend_count_label.setText(
            f"匹配好友: {len(matched_friends)}/{len(friends)}")

    def search_groups(self):
        search_text = self.group_search_entry.text().lower()

        self.group_tree.clear()

        if not hasattr(self, 'all_contacts') or not self.all_contacts:
            return

        groups = [contact for contact in self.all_contacts if "@chatroom" in contact["wxid"]
                  and contact["nickname"] and contact["nickname"].strip()]

        if not search_text:
            for i, group in enumerate(groups, 1):
                item = QTreeWidgetItem([
                    str(i),
                    group["nickname"],
                    group["wxid"]
                ])
                self.group_tree.addTopLevelItem(item)

            self.group_count_label.setText(f"群总数: {len(groups)}")
            return

        matched_groups = []
        for group in groups:
            nickname = group["nickname"].lower() if group["nickname"] else ""

            if search_text in nickname:
                matched_groups.append(group)

        for i, group in enumerate(matched_groups, 1):
            item = QTreeWidgetItem([
                str(i),
                group["nickname"],
                group["wxid"]
            ])
            self.group_tree.addTopLevelItem(item)

        self.group_count_label.setText(
            f"匹配群聊: {len(matched_groups)}/{len(groups)}")

    def get_wechat_groups(self):
        self.auto_fetch_contacts()

    def refresh_group_list(self):
        self.auto_fetch_contacts()

    def refresh_friend_list(self):
        self.auto_fetch_contacts()

    def show_modify_remark_dialog(self, wxid, nickname, current_remark):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"修改好友备注 - {nickname}")
        dialog.setMinimumWidth(400)

        layout = QVBoxLayout(dialog)

        info_label = QLabel(f"正在修改 {nickname} 的备注")
        layout.addWidget(info_label)

        wxid_layout = QHBoxLayout()
        wxid_label = QLabel("微信ID:")
        wxid_value = QLabel(wxid)
        wxid_layout.addWidget(wxid_label)
        wxid_layout.addWidget(wxid_value)
        layout.addLayout(wxid_layout)

        current_remark_layout = QHBoxLayout()
        current_remark_label = QLabel("当前备注:")
        current_remark_value = QLabel(current_remark if current_remark else "无")
        current_remark_layout.addWidget(current_remark_label)
        current_remark_layout.addWidget(current_remark_value)
        layout.addLayout(current_remark_layout)

        new_remark_layout = QHBoxLayout()
        new_remark_label = QLabel("新备注:")
        new_remark_input = QLineEdit()
        new_remark_input.setText(current_remark)
        new_remark_layout.addWidget(new_remark_label)
        new_remark_layout.addWidget(new_remark_input)
        layout.addLayout(new_remark_layout)

        button_layout = QHBoxLayout()
        cancel_button = QPushButton("取消")
        confirm_button = QPushButton("确认修改")
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(confirm_button)
        layout.addLayout(button_layout)

        status_label = QLabel("")
        layout.addWidget(status_label)

        cancel_button.clicked.connect(dialog.reject)
        confirm_button.clicked.connect(lambda: self.modify_friend_remark(
            wxid, new_remark_input.text(), dialog, status_label))

        dialog.exec()

    def modify_friend_remark(self, wxid, new_remark, dialog, status_label):
        if not new_remark.strip():
            status_label.setText("备注不能为空")
            status_label.setStyleSheet(f"color: {StyleSheet.ERROR_COLOR}")
            return

        try:
            wechat_pids = self.wechat_info.find_all_wechat_processes()
            if not wechat_pids:
                status_label.setText("未找到微信进程，请确保微信已启动")
                status_label.setStyleSheet(f"color: {StyleSheet.ERROR_COLOR}")
                return

            status_label.setText("正在修改备注，请稍候...")
            status_label.setStyleSheet(f"color: {StyleSheet.INFO_COLOR}")
            QApplication.processEvents()

            remark_modifier = RemarkModifier()

            selected_pid = getattr(self, 'current_account_pid', None)
            if selected_pid not in (wechat_pids or []):
                selected_pid = wechat_pids[0]

            result = remark_modifier.modify_remark(selected_pid, wxid, new_remark)

            if result:
                status_label.setText("备注修改成功！")
                status_label.setStyleSheet(f"color: {StyleSheet.SUCCESS_COLOR}")

                wechat_info = SimpleWeChatInfo()
                accounts = wechat_info.run()
                if accounts:
                    matched = next((acc for acc in accounts if acc.get('pid') == selected_pid), None) or accounts[0]

                    self.data_manager.update_account_remark(
                        matched.get('wxid', ''),
                        wxid,
                        new_remark
                    )

                QTimer.singleShot(1000, dialog.accept)

                QTimer.singleShot(1500, self.refresh_friend_list)
            else:
                status_label.setText("备注修改失败，请检查微信ID是否正确")
                status_label.setStyleSheet("color: red")

        except Exception as e:
            status_label.setText(f"修改备注失败: {str(e)}")
            status_label.setStyleSheet("color: red")

    def modify_friend_remark_silent(self, wxid: str, new_remark: str) -> bool:
        try:
            if not wxid or not new_remark or not new_remark.strip():
                print(f"备注修改失败：参数无效 wxid={wxid}, remark={new_remark}")
                return False
            
            wechat_pids = self.wechat_info.find_all_wechat_processes()
            if not wechat_pids:
                print("备注修改失败：未找到微信进程")
                return False
            
            # 优先使用当前选中的微信账号PID
            selected_pid = None
            if hasattr(self, 'selected_wechat_accounts') and self.selected_wechat_accounts:
                current_account = self.selected_wechat_accounts[getattr(self, 'current_account_index', 0)]
                selected_pid = current_account.get('pid')
                print(f"使用当前账号PID: {selected_pid} ({current_account.get('nickname', '未知')})")
            
            if not selected_pid or selected_pid not in wechat_pids:
                selected_pid = wechat_pids[0]
                print(f"回退到第一个微信进程PID: {selected_pid}")
            
            modifier = RemarkModifier()
            print(f"开始修改备注：wxid={wxid}, remark={new_remark.strip()}, pid={selected_pid}")
            
            ok = modifier.modify_remark(selected_pid, wxid, new_remark.strip())
            
            if ok:
                print(f"备注修改成功：{wxid} -> {new_remark.strip()}")
                try:
                    accounts = SimpleWeChatInfo().run()
                    if accounts:
                        matched = next((acc for acc in accounts if acc.get('pid') == selected_pid), None) or accounts[0]
                        self.data_manager.update_account_remark(
                            matched.get('wxid', ''),
                            wxid,
                            new_remark.strip()
                        )
                except Exception as e:
                    print(f"更新数据管理器失败: {e}")
            else:
                print(f"备注修改失败：RemarkModifier.modify_remark返回False")
            
            return ok
        except Exception as e:
            print(f"备注修改异常: {e}")
            return False

    def resolve_sender_nickname(self, sender_wxid: str) -> str:
        try:
            if not sender_wxid:
                return ''
            if hasattr(self, 'add_friend_table') and self.add_friend_table is not None:
                rows = self.add_friend_table.rowCount()
                cols = self.add_friend_table.columnCount()
                for r in range(rows):
                    matched_row = False
                    for c in range(cols):
                        item = self.add_friend_table.item(r, c)
                        if item and item.text().strip() == sender_wxid:
                            matched_row = True
                            break
                    if matched_row:
                        nick_item = self.add_friend_table.item(r, 5)
                        if nick_item:
                            return nick_item.text().strip()
            return ''
        except Exception:
            return ''

    def parse_nickname_from_message(self, content: str) -> str:
        try:
            if not content:
                return ''
            text = content.replace(' ', '')

            import re
            patterns = [
                r"你已添加了[（(]([^）)]+)[）)]，?现在可以开始聊天了",
                r"已添加了[（(]([^）)]+)[）)]，?现在可以开始聊天了",
            ]
            for pat in patterns:
                m = re.search(pat, text)
                if m:
                    return m.group(1).strip()
            return ''
        except Exception:
            return ''

    def find_row_by_nickname_or_wxid(self, nickname: str, wxid: str) -> int:
        try:
            if not hasattr(self, 'add_friend_table') or self.add_friend_table is None:
                return -1
            rows = self.add_friend_table.rowCount()
            cols = self.add_friend_table.columnCount()

            if nickname:
                for row in range(rows):
                    nick_item = self.add_friend_table.item(row, 5)
                    table_nickname = nick_item.text().strip() if nick_item else ''
                    if table_nickname == nickname:
                        return row

            if wxid:
                for row in range(rows):
                    for c in range(cols):
                        item = self.add_friend_table.item(row, c)
                        if item and item.text().strip() == wxid:
                            return row
            return -1
        except Exception:
            return -1

    def handle_auto_remark_on_acceptance(self, message_data: dict) -> None:
        try:
            content = (message_data.get('original_content') or message_data.get('content') or '').strip()
            normalized = content.replace(' ', '')

            print(f"处理好友验证消息：{content}")
            
            # 检查多种好友验证通过的消息格式
            is_friend_accepted = any([
                '我通过了你的朋友验证请求' in normalized,
                '你已添加了' in normalized,
                '已通过你的朋友验证请求' in normalized,  # 新增格式
                '通过了你的朋友验证请求' in normalized,   # 新增格式
                '现在可以开始聊天了' in normalized and ('验证请求' in normalized or '朋友验证' in normalized)  # 更通用的匹配
            ])
            
            if not is_friend_accepted:
                print("消息不是好友验证通过消息，跳过")
                return
                
            sender_nickname = (message_data.get('sender_nickname') or '').strip()
            sender_wxid = (message_data.get('sender_wxid') or '').strip()
            
            print(f"发送者信息：wxid={sender_wxid}, nickname={sender_nickname}")
            
            if not sender_wxid:
                print("发送者wxid为空，跳过")
                return

            if not sender_nickname:
                parsed = self.parse_nickname_from_message(content)
                if parsed:
                    sender_nickname = parsed
                if not sender_nickname:
                    resolved = self.resolve_sender_nickname(sender_wxid)
                    if resolved:
                        sender_nickname = resolved
                        
            if not hasattr(self, 'add_friend_table') or self.add_friend_table is None:
                print("添加好友表格不存在，跳过")
                return
                
            matched_row = self.find_row_by_nickname_or_wxid(sender_nickname, sender_wxid)
            print(f"匹配到的行号：{matched_row}")
            
            if matched_row < 0:
                print("未找到匹配的添加好友记录")
                return
                
            remark_item = self.add_friend_table.item(matched_row, 6)
            remark_text = remark_item.text().strip() if remark_item else ''
            
            print(f"备注内容：'{remark_text}'")
            
            if not remark_text:
                print("没有修改备注")
                return
                
            print(f"开始修改备注：{sender_wxid} -> {remark_text}")
            
            if self.modify_friend_remark_silent(sender_wxid, remark_text):
                print("备注修改成功，更新状态")
                try:
                    status_item = self.add_friend_table.item(matched_row, 3)
                    prev = status_item.text() if status_item else ''
                    label = '已改备注'
                    new_status = prev if label in prev else (f"{prev} {label}" if prev else label)
                    self.add_friend_table.setItem(matched_row, 3, QTableWidgetItem(new_status))
                    self.save_add_friend_data()
                    print(f"状态已更新为：{new_status}")
                except Exception as e:
                    print(f"更新状态失败: {e}")
            else:
                print("备注修改失败")
                
        except Exception as e:
            print(f"处理好友验证消息异常: {e}")
    def on_apply_remark_by_accepted_nickname(self):
        try:
            if not hasattr(self, 'add_friend_table') or self.add_friend_table is None:
                QMessageBox.warning(self, "提示", "未找到添加好友表")
                return
            from PySide6.QtWidgets import QInputDialog
            nickname, ok = QInputDialog.getText(self, "按通过昵称修改备注", "请输入系统消息中的通过昵称：")
            if not ok or not nickname.strip():
                return
            self.apply_remark_for_accepted_nickname(nickname.strip())
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败：{e}")

    def _extract_wxid_from_row(self, row: int) -> str:
        try:
            if not hasattr(self, 'add_friend_table') or self.add_friend_table is None:
                return ''
            cols = self.add_friend_table.columnCount()
            for c in range(cols):
                item = self.add_friend_table.item(row, c)
                text = item.text().strip() if item else ''
                if not text:
                    continue
                if text.startswith('wxid_') or text.startswith('v1_') or ('@' in text):
                    return text
            return ''
        except Exception:
            return ''

    def apply_remark_for_accepted_nickname(self, accepted_nickname: str):
        try:
            matched = 0
            rows = self.add_friend_table.rowCount()
            for row in range(rows):
                nick_item = self.add_friend_table.item(row, 5)
                table_nickname = nick_item.text().strip() if nick_item else ''
                if table_nickname != accepted_nickname:
                    continue
                remark_item = self.add_friend_table.item(row, 6)
                remark_text = remark_item.text().strip() if remark_item else ''
                if not remark_text:
                    continue
                wxid = self._extract_wxid_from_row(row)
                if not wxid:
                    continue
                if self.modify_friend_remark_silent(wxid, remark_text):
                    status_item = self.add_friend_table.item(row, 3)
                    prev = status_item.text() if status_item else ''
                    label = '已改备注'
                    new_status = prev if label in prev else (f"{prev} {label}" if prev else label)
                    self.add_friend_table.setItem(row, 3, QTableWidgetItem(new_status))
                    matched += 1
            if matched > 0:
                self.save_add_friend_data()
                QMessageBox.information(self, "完成", f"已根据通过昵称修改 {matched} 条备注")
            else:
                QMessageBox.information(self, "提示", "未找到匹配昵称且备注非空的记录")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"执行失败：{e}")

    def on_apply_remark_batch_by_table(self):
        try:
            if not hasattr(self, 'add_friend_table') or self.add_friend_table is None:
                QMessageBox.warning(self, "提示", "未找到添加好友表")
                return
            changed = 0
            rows = self.add_friend_table.rowCount()
            for row in range(rows):
                nick_item = self.add_friend_table.item(row, 5)
                remark_item = self.add_friend_table.item(row, 6)
                nickname = nick_item.text().strip() if nick_item else ''
                remark_text = remark_item.text().strip() if remark_item else ''
                if not nickname or not remark_text:
                    continue
                wxid = self._extract_wxid_from_row(row)
                if not wxid:
                    continue
                if self.modify_friend_remark_silent(wxid, remark_text):
                    status_item = self.add_friend_table.item(row, 3)
                    prev = status_item.text() if status_item else ''
                    label = '已改备注'
                    new_status = prev if label in prev else (f"{prev} {label}" if prev else label)
                    self.add_friend_table.setItem(row, 3, QTableWidgetItem(new_status))
                    changed += 1
            if changed > 0:
                self.save_add_friend_data()
                QMessageBox.information(self, "完成", f"已按表批量修改 {changed} 条备注")
            else:
                QMessageBox.information(self, "提示", "未找到需要修改的记录")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"执行失败：{e}")
        except Exception as e:
            pass
    def on_account_double_click(self, item, _):
        try:
            index = self.account_tree.indexOfTopLevelItem(item)
            if index < 0:
                return

            wechat_info = SimpleWeChatInfo()
            accounts = wechat_info.run()

            if not accounts or index >= len(accounts):
                QMessageBox.warning(self, "获取失败", "无法获取选中账号信息", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            selected_pid = accounts[index]['pid']

            try:
                self.current_account_pid = selected_pid
            except Exception:
                pass

            if hasattr(self, 'all_contacts'):
                self.all_contacts = []

            self.friend_tree.clear()
            self.group_tree.clear()

            self.statusBar().showMessage(f"正在加载账号 {accounts[index]['nickname']} 的联系人数据，请稍候...")

            try:
                def update_progress(current, total, contact):
                    if total > 0:
                        percent = int(current * 100 / total)
                        if contact:
                            self.statusBar().showMessage(f"正在读取联系人数据 {percent}%: {contact.get('nickname', '')}")
                        else:
                            self.statusBar().showMessage(f"读取联系人完成 {percent}%")
                    QApplication.processEvents()

                resources = get_wechat_resources(selected_pid, update_progress)

                if resources:
                    filtered_contacts = [contact for contact in resources['contacts']
                                        if contact.get('nickname') and contact.get('nickname').strip()]

                    filtered_friends = [friend for friend in resources['friends']
                                      if friend.get('nickname') and friend.get('nickname').strip()]

                    filtered_groups = [group for group in resources['groups']
                                     if group.get('nickname') and group.get('nickname').strip()
                                     and '@chatroom' in group.get('wxid', '')]

                    self.all_contacts = filtered_contacts

                    self.friend_tree.clear()
                    self.friend_count_label.setText(f"好友总数: {len(filtered_friends)}")

                    prev_sorting = self.friend_tree.isSortingEnabled()
                    self.friend_tree.setSortingEnabled(False)
                    self.friend_tree.setUpdatesEnabled(False)

                    for i, friend in enumerate(filtered_friends, 1):
                        phone = friend.get("phone", "")
                        if phone == "未知":
                            phone = ""

                        item = QTreeWidgetItem([
                            str(i),
                            friend["nickname"],
                            friend["wxid"],
                            friend.get("remarks", ""),
                            friend.get("tag", ""),
                            phone
                        ])
                        self.friend_tree.addTopLevelItem(item)

                        if i % 200 == 0:
                            QApplication.processEvents()

                    self.friend_tree.setUpdatesEnabled(True)
                    self.friend_tree.setSortingEnabled(prev_sorting)

                    self.group_tree.clear()
                    self.group_count_label.setText(f"群总数: {len(filtered_groups)}")

                    prev_group_sorting = self.group_tree.isSortingEnabled()
                    self.group_tree.setSortingEnabled(False)
                    self.group_tree.setUpdatesEnabled(False)

                    for i, group in enumerate(filtered_groups, 1):
                        item = QTreeWidgetItem([
                            str(i),
                            group["nickname"],
                            group["wxid"]
                        ])
                        self.group_tree.addTopLevelItem(item)

                        if i % 200 == 0:
                            QApplication.processEvents()

                    self.group_tree.setUpdatesEnabled(True)
                    self.group_tree.setSortingEnabled(prev_group_sorting)

                    pass
                else:
                    self.statusBar().showMessage("加载联系人数据失败!", 5000)
                    QMessageBox.warning(self, "加载失败",
                                      "无法获取联系人信息，请确保微信正常登录",
                                      QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            except Exception as e:
                self.statusBar().showMessage(f"获取联系人失败: {str(e)}", 5000)
                QMessageBox.critical(self, "错误", f"获取联系人失败: {str(e)}",
                                  QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理双击事件失败: {str(e)}",
                               QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def show_friend_context_menu(self, position):
        selected_items = self.friend_tree.selectedItems()
        if not selected_items:
            return

        menu = QMenu()

        send_message_action = menu.addAction("发送消息")

        modify_remark_action = menu.addAction("修改备注")

        scheduled_message_action = menu.addAction("定时消息")

        specify_reply_action = menu.addAction("指定回复")

        action = menu.exec(self.friend_tree.viewport().mapToGlobal(position))

        wxid = selected_items[0].text(2)
        nickname = selected_items[0].text(1)
        current_remark = selected_items[0].text(3)

        if action == send_message_action:
            dialog = SendMessageDialog(nickname, self, wxid)
            dialog.exec()
            try:
                if hasattr(self, 'yuanbao_reply_switch'):
                    self.yuanbao_reply_switch.setEnabled(True)
                if hasattr(self, 'model_reply_switch'):
                    self.model_reply_switch.setEnabled(True)
            except Exception:
                pass
        elif action == modify_remark_action:
            self.show_modify_remark_dialog(wxid, nickname, current_remark)
        elif action == scheduled_message_action:
            self.handle_scheduled_message_for_friends(selected_items)
        elif action == specify_reply_action:
            try:
                for item in selected_items:
                    wxid_item = item.text(2)
                    if wxid_item:
                        self.specific_friend_wxids.add(wxid_item)
                self.update_specific_selected_lists()

                self.notebook.setCurrentWidget(self.auto_reply_tab)
                self.statusBar().showMessage(f"已添加{len(selected_items)}个好友到指定回复", 3000)
            except Exception as e:
                pass
    def show_group_context_menu(self, position):
        selected_items = self.group_tree.selectedItems()
        if not selected_items:
            return

        menu = QMenu()

        send_message_action = menu.addAction("发送消息")

        get_members_action = menu.addAction("获取群成员")

        add_group_members_action = menu.addAction("添加群成员")

        scheduled_message_action = menu.addAction("定时消息")

        specify_reply_action = menu.addAction("指定回复")

        action = menu.exec(self.group_tree.viewport().mapToGlobal(position))

        wxid = selected_items[0].text(2)
        nickname = selected_items[0].text(1)

        if action == send_message_action:
            dialog = SendMessageDialog(nickname, self, wxid)
            dialog.exec()
        elif action == get_members_action:
            self.get_group_members(wxid, nickname)
        elif action == add_group_members_action:
            self.add_group_members_from_context(wxid, nickname)
        elif action == scheduled_message_action:
            self.handle_scheduled_message_for_groups(selected_items)
        elif action == specify_reply_action:
            try:
                for item in selected_items:
                    group_wxid = item.text(2)
                    if group_wxid:
                        self.specific_group_wxids.add(group_wxid)
                self.update_specific_selected_lists()

                self.notebook.setCurrentWidget(self.auto_reply_tab)
                self.statusBar().showMessage(f"已添加{len(selected_items)}个群到指定回复", 3000)
            except Exception as e:
                pass
    def get_group_members(self, group_id, group_name):
        try:
            pid = self._get_wechat_pid()
            if not pid:
                QMessageBox.warning(self, "获取失败", "未找到微信进程，请确保微信已启动", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            self.members_tree.clear()
            def progress_callback(current, total, member):
                pass

            members = get_group_members(pid, group_id, progress_callback)

            for i, member in enumerate(members, 1):
                nickname = member.get("nickname", "无昵称")
                wxid = member.get("wxid", "")

                item = QTreeWidgetItem()
                item.setText(0, str(i))
                item.setText(1, group_name)
                item.setText(2, nickname)
                item.setText(3, wxid)
                self.members_tree.addTopLevelItem(item)

            self.members_count_label.setText(f"成员总数: {len(members)}")

            import gc
            gc.collect()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"获取群成员失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def _get_wechat_pid(self):
        wechat_info = SimpleWeChatInfo()
        accounts = wechat_info.run()

        if not accounts:
            QMessageBox.warning(self, "发送失败", "未找到已登录的微信账号", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return None

        try:
            selected_pid = getattr(self, 'current_account_pid', None)
            if selected_pid:
                for acc in accounts:
                    if acc.get('pid') == selected_pid:
                        return selected_pid
        except Exception:
            pass

        if len(accounts) == 1:
            return accounts[0]['pid']

        return accounts[0]['pid']

    def _handle_send_failure_simple(self, current_pid, target_id, content, is_group=False):
        try:
            retry_result = send_message_simple(current_pid, target_id, content)
            if retry_result:
                pass
            else:
                pass
        except Exception as e:
            pass
    def send_message_to_wxid(self, wxid, content, pid=None):
        try:
            current_pid = pid if pid else self._get_wechat_pid()
            if not current_pid:
                return False
            return send_message_to_wxid(current_pid, wxid, content)
        except Exception as e:
            QMessageBox.warning(self, "发送失败", f"发送消息时出错: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return False

    def send_image_to_wxid(self, wxid, image_path, pid=None):
        try:
            current_pid = pid if pid else self._get_wechat_pid()
            if not current_pid:
                return False

            return send_image_to_wxid(current_pid, wxid, image_path)
        except Exception as e:
            QMessageBox.warning(self, "发送失败", f"发送图片时出错: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return False

    def show_members_context_menu(self, position):
        selected_items = self.members_tree.selectedItems()
        if not selected_items:
            return

        menu = QMenu()

        copy_id_action = menu.addAction("复制群成员ID")

        add_friend_action = menu.addAction("添加为好友")

        action = menu.exec(self.members_tree.viewport().mapToGlobal(position))

        member_wxid = selected_items[0].text(3)
        member_nickname = selected_items[0].text(2)

        if action == copy_id_action:
            clipboard = QApplication.clipboard()
            clipboard.setText(member_wxid)
            QMessageBox.information(self, "复制成功", f"已复制群成员ID: {member_wxid}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
        elif action == add_friend_action:
            self.friend_id_input.setText(member_wxid)
            self.greeting_input.setText(f"您好，我是通过群聊认识您的，想添加您为好友")

            self.notebook.setCurrentWidget(self.add_friend_tab)

            reply = QMessageBox.question(self, "添加好友",
                                      f"确定要添加 {member_nickname} 为好友吗？\nID: {member_wxid}",
                                      QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    idx = self.scene_combo.findData("14")
                    if idx != -1:
                        self.scene_combo.setCurrentIndex(idx)
                except Exception:
                    pass

                self.add_friend()

    def detect_wechat_for_add_friend(self):
        try:
            wechat_pids = [p.pid for p in psutil.process_iter() if 'WeChat.exe' in p.name()]

            if wechat_pids:
                self.wechat_pid = wechat_pids[0]
                self.log_add_friend(f"找到微信进程: {wechat_pids}")

                try:
                    wechat_info = SimpleWeChatInfo()
                    accounts = wechat_info.run()
                    if accounts:
                        account = accounts[0]
                        self.log_add_friend(f"当前登录账号: {account['nickname']} ({account['wxid']})")
                except Exception as e:
                    self.log_add_friend(f"获取账号信息失败: {e}")
            else:
                self.wechat_pid = None
                self.log_add_friend("未找到微信进程，请确保微信已启动")
        except Exception as e:
            self.log_add_friend(f"检测微信进程失败: {e}")
    def log_add_friend(self, message):
        self.add_friend_result.append(message)

    def add_friend(self):
        pid = self._get_wechat_pid()
        if not pid:
            QMessageBox.warning(self, "错误", "未找到微信进程，请确保微信已启动", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        self.wechat_pid = pid

        friend_id = self.friend_id_input.text().strip()
        greeting = self.greeting_input.text().strip()
        scene = self.scene_combo.currentData() or "3"

        if not friend_id:
            QMessageBox.warning(self, "错误", "请输入好友ID或手机号", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            return

        if not greeting:
            greeting = "您好，我想添加您为好友"

        self.add_friend_result.clear()
        self.log_add_friend(f"开始添加好友 {friend_id}...")
        self.log_add_friend(f"招呼语: {greeting}")
        self.log_add_friend(f"场景类型: {self.scene_combo.currentText()} (值={scene})")

        try:
            start_time = time.time()
            result = add_wechat_friend(self.wechat_pid, friend_id, greeting, scene)
            end_time = time.time()

            self.log_add_friend(f"\n操作完成，耗时 {end_time - start_time:.2f} 秒")

            if result:
                self.log_add_friend(f"已成功发送添加好友请求给 {friend_id}")
                QMessageBox.information(self, "添加成功", f"已成功发送添加好友请求给 {friend_id}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
            else:
                self.log_add_friend(f"添加好友 {friend_id} 失败，请检查ID是否正确")
                QMessageBox.warning(self, "添加失败", "添加好友失败，请检查ID是否正确", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

        except Exception as e:
            self.log_add_friend(f"添加好友失败: {e}")
            self.log_add_friend(str(e))
            QMessageBox.critical(self, "错误", f"添加好友失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def auto_fetch_contacts(self):
        try:
            wechat_pids = self.wechat_info.find_all_wechat_processes()
            if not wechat_pids:
                return

            pid = wechat_pids[0]

            try:
                self.current_account_pid = pid
            except Exception:
                pass

            self.statusBar().showMessage("正在自动加载联系人数据，请稍候...")

            try:
                def update_progress(current, total, contact):
                    if total > 0:
                        percent = int(current * 100 / total)
                        if contact:
                            self.statusBar().showMessage(f"正在读取联系人数据 {percent}%: {contact.get('nickname', '')}")
                        else:
                            self.statusBar().showMessage(f"读取联系人完成 {percent}%")
                    QApplication.processEvents()

                resources = get_wechat_resources(pid, update_progress)

                if resources:
                    filtered_contacts = [contact for contact in resources['contacts']
                                        if contact.get('nickname') and contact.get('nickname').strip()]

                    filtered_friends = [friend for friend in resources['friends']
                                      if friend.get('nickname') and friend.get('nickname').strip()]

                    filtered_groups = [group for group in resources['groups']
                                     if group.get('nickname') and group.get('nickname').strip()
                                     and '@chatroom' in group.get('wxid', '')]

                    self.all_contacts = filtered_contacts

                    self.friend_tree.clear()
                    self.friend_count_label.setText(f"好友总数: {len(filtered_friends)}")

                    prev_sorting = self.friend_tree.isSortingEnabled()
                    self.friend_tree.setSortingEnabled(False)
                    self.friend_tree.setUpdatesEnabled(False)

                    for i, friend in enumerate(filtered_friends, 1):
                        phone = friend.get("phone", "")
                        if phone == "未知":
                            phone = ""

                        item = QTreeWidgetItem([
                            str(i),
                            friend["nickname"],
                            friend["wxid"],
                            friend.get("remarks", ""),
                            friend.get("tag", ""),
                            phone
                        ])
                        self.friend_tree.addTopLevelItem(item)
                        if i % 200 == 0:
                            QApplication.processEvents()

                    self.friend_tree.setUpdatesEnabled(True)
                    self.friend_tree.setSortingEnabled(prev_sorting)

                    self.group_tree.clear()
                    self.group_count_label.setText(f"群总数: {len(filtered_groups)}")

                    for i, group in enumerate(filtered_groups, 1):
                        item = QTreeWidgetItem([
                            str(i),
                            group["nickname"],
                            group["wxid"]
                        ])
                        self.group_tree.addTopLevelItem(item)

                    self.statusBar().showMessage("正在后台加载群成员信息...", 5000)

                    def fetch_group_members():
                        try:
                            all_members = get_all_group_members(pid, filtered_groups)

                            if all_members:
                                with_duplicates = len(self.all_contacts)

                                existing_wxids = set(c.get("wxid", "") for c in self.all_contacts if isinstance(c, dict))

                                filtered_members = [
                                    m for m in all_members
                                    if isinstance(m, dict)
                                    and m.get('nickname') and str(m.get('nickname')).strip()
                                ]

                                for member in filtered_members:
                                    wxid = member.get("wxid", "")
                                    if wxid and wxid not in existing_wxids:
                                        self.all_contacts.append(member)
                                        existing_wxids.add(wxid)

                            else:
                                pass
                        except Exception as e:
                            pass
                    threading.Thread(target=fetch_group_members, daemon=True).start()

                self.statusBar().showMessage("联系人数据加载完成!", 5000)

            except Exception as e:
                self.statusBar().showMessage(f"加载联系人数据失败: {str(e)}", 5000)

        except Exception as e:
            pass

    def show_load_history_dialog(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("加载历史消息")
            dialog.setMinimumWidth(400)
            layout = QVBoxLayout(dialog)

            date_layout = QFormLayout()

            from_date_layout = QHBoxLayout()
            from_date_label = QLabel("开始日期:")
            from_date = QDateEdit(dialog)
            from_date.setCalendarPopup(True)

            from PySide6.QtCore import QDate
            current_date = datetime.now()
            first_day = QDate(current_date.year, current_date.month, 1)
            from_date.setDate(first_day)

            from_date_layout.addWidget(from_date_label)
            from_date_layout.addWidget(from_date)
            date_layout.addRow("", from_date_layout)

            to_date_layout = QHBoxLayout()
            to_date_label = QLabel("结束日期:")
            to_date = QDateEdit(dialog)
            to_date.setCalendarPopup(True)

            today = QDate(current_date.year, current_date.month, current_date.day)
            to_date.setDate(today)
            to_date_layout.addWidget(to_date_label)
            to_date_layout.addWidget(to_date)
            date_layout.addRow("", to_date_layout)

            layout.addLayout(date_layout)

            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                start_date = from_date.date()
                end_date = to_date.date()

                start_dt = datetime(start_date.year(), start_date.month(), start_date.day(), 0, 0, 0)
                end_dt = datetime(end_date.year(), end_date.month(), end_date.day(), 23, 59, 59)

                start_timestamp = int(start_dt.timestamp())
                end_timestamp = int(end_dt.timestamp())

                self.load_saved_messages(start_timestamp, end_timestamp)

        except Exception as e:
            pass
    def load_saved_messages(self, start_timestamp=None, end_timestamp=None):
        try:
            messages = self.data_manager.load_messages()
            if not messages:
                self.statusBar().showMessage("未找到历史消息", 5000)
                return

            messages.sort(key=lambda m: int(m.get('timestamp', 0)))

            filtered_messages = []
            for message in messages:
                timestamp = int(message['timestamp'])

                if start_timestamp is None and end_timestamp is None:
                    if hasattr(self, 'startup_timestamp') and timestamp >= self.startup_timestamp:
                        filtered_messages.append(message)
                else:
                    if start_timestamp is not None and timestamp < start_timestamp:
                        continue
                    if end_timestamp is not None and timestamp > end_timestamp:
                        continue
                    filtered_messages.append(message)

            if not filtered_messages:
                time_range = ""
                if start_timestamp and end_timestamp:
                    start_date = datetime.fromtimestamp(start_timestamp).strftime('%Y-%m-%d')
                    end_date = datetime.fromtimestamp(end_timestamp).strftime('%Y-%m-%d')
                    time_range = f"在 {start_date} 至 {end_date} 时间段内"
                self.statusBar().showMessage(f"未找到{time_range}的历史消息", 5000)
                return

            count = len(filtered_messages)
            reply = QMessageBox.question(
                self,
                "确认加载",
                f"找到 {count} 条历史消息，是否加载？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )

            if reply != QMessageBox.StandardButton.Yes:
                return

            for message in filtered_messages:
                timestamp = int(message['timestamp'])
                time_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                wxid = message.get('wxid', '')
                content = message.get('content', '')
                account_info = message.get('account', {})
                account_name = account_info.get('nickname', '未知账号')

                parsed_content = parse_special_message(content)
                if parsed_content:
                    content = parsed_content

                is_group_message = "@chatroom" in wxid and "member_id" in message
                member_id = message.get("member_id", "")

                contact_info = None
                if hasattr(self, 'all_contacts') and self.all_contacts:
                    for contact in self.all_contacts:
                        if contact.get("wxid") == wxid:
                            contact_info = contact
                            break

                sender_name = wxid
                if is_group_message:
                    if contact_info:
                       sender_name = contact_info.get("nickname", wxid)

                    member_info = None
                    if hasattr(self, 'all_contacts') and self.all_contacts:
                        for contact in self.all_contacts:
                            if contact.get("wxid") == member_id:
                                member_info = contact
                                break

                    if member_info:
                        member_name = member_info.get("nickname", member_id)
                    content = f"[{member_name}]: {content}"
                else:
                    if contact_info:
                        remarks = contact_info.get("remarks", "")
                        if remarks and remarks.strip():
                           sender_name = remarks
                        else:
                           sender_name = contact_info.get("nickname", wxid)

                message_data = {
                    'self_nickname': account_name,
                    'sender_nickname': sender_name,
                    'sender_wxid': wxid,
                    'content': content,
                    'receive_time': time_str
                }

                self.add_message_to_auto_reply_history(message_data)

            self.statusBar().showMessage(f"已加载 {len(filtered_messages)} 条历史消息", 5000)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载历史消息失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def load_all_accounts_data(self):
        try:
            wechat_info = SimpleWeChatInfo()
            accounts = wechat_info.run()

            if not accounts:
                self.statusBar().showMessage("未找到已登录的微信账号", 5000)
                return

            loaded_count = 0
            for account in accounts:
                wxid = account.get('wxid')
                if wxid:
                    account_data = self.data_manager.load_account_data(wxid)
                    if account_data:
                        self.update_contacts_from_data(account_data)
                        loaded_count += 1
                    else:
                        self.fetch_and_save_account_data(account)

            if loaded_count > 0:
                self.statusBar().showMessage(f"成功加载 {loaded_count} 个微信账号的联系人数据", 5000)
            else:
                self.statusBar().showMessage("未找到缓存的联系人数据，将从内存获取", 5000)

        except Exception as e:
            self.statusBar().showMessage(f"加载微信账号数据失败: {str(e)}", 5000)

    def update_contacts_from_data(self, account_data):
        try:
            if not hasattr(self, 'all_contacts'):
                self.all_contacts = []

            contacts = account_data.get('contacts', [])
            for contact in contacts:
                if contact not in self.all_contacts:
                    self.all_contacts.append(contact)

            self.friend_tree.clear()
            friends = account_data.get('friends', [])
            self.friend_count_label.setText(f"好友总数: {len(friends)}")

            for i, friend in enumerate(friends, 1):
                phone = friend.get("phone", "")
                if phone == "未知":
                    phone = ""

                item = QTreeWidgetItem([
                    str(i),
                    friend["nickname"],
                    friend["wxid"],
                    friend.get("remarks", ""),
                    friend.get("tag", ""),
                    phone
                ])
                self.friend_tree.addTopLevelItem(item)

            self.group_tree.clear()
            groups = account_data.get('groups', [])
            self.group_count_label.setText(f"群总数: {len(groups)}")

            for i, group in enumerate(groups, 1):
                item = QTreeWidgetItem([
                    str(i),
                    group["nickname"],
                    group["wxid"]
                ])
                self.group_tree.addTopLevelItem(item)

        except Exception as e:
            pass

    def fetch_and_save_account_data(self, account):
        try:
            pid = account.get('pid')
            if not pid:
                return

            self.statusBar().showMessage(f"正在从内存获取账号 {account.get('nickname')} 的联系人数据...")

            def update_progress(current, total, contact):
                if total > 0:
                    percent = int(current * 100 / total)
                    if contact:
                        self.statusBar().showMessage(f"正在读取联系人数据 {percent}%: {contact.get('nickname', '')}")
                    else:
                        self.statusBar().showMessage(f"读取联系人完成 {percent}%")
                QApplication.processEvents()

            resources = get_wechat_resources(pid, update_progress)

            if resources:
                filtered_contacts = [contact for contact in resources['contacts']
                                    if contact.get('nickname') and contact.get('nickname').strip()]

                filtered_friends = [friend for friend in resources['friends']
                                  if friend.get('nickname') and friend.get('nickname').strip()]

                filtered_groups = [group for group in resources['groups']
                                 if group.get('nickname') and group.get('nickname').strip()
                                 and '@chatroom' in group.get('wxid', '')]

                self.data_manager.save_account_data(
                    account,
                    filtered_contacts,
                    filtered_friends,
                    filtered_groups
                )

                account_data = {
                    'account_info': account,
                    'contacts': filtered_contacts,
                    'friends': filtered_friends,
                    'groups': filtered_groups
                }
                self.update_contacts_from_data(account_data)

                self.statusBar().showMessage(f"成功获取并缓存账号 {account.get('nickname')} 的联系人数据", 5000)
            else:
                self.statusBar().showMessage(f"获取账号 {account.get('nickname')} 的联系人数据失败", 5000)
        except Exception as e:
            self.statusBar().showMessage(f"加载联系人数据失败: {str(e)}", 5000)

    def save_rules_data(self):
        if hasattr(self, '_is_saving') and self._is_saving:
            return

        try:
            self._is_saving = True

            config_dir = "config"
            os.makedirs(config_dir, exist_ok=True)
            temp_file = os.path.join(config_dir, 'auto_reply_rules.json.tmp')
            target_file = os.path.join(config_dir, 'auto_reply_rules.json')

            rules = []
            keywords_set = set()

            for row in range(self.rules_table.rowCount()):
                checkbox_widget = self.rules_table.cellWidget(row, 1)
                enabled = False
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox:
                        enabled = checkbox.isChecked()

                keyword = self.rules_table.item(row, 2)
                reply = self.rules_table.item(row, 3)

                if keyword and reply:
                    keyword_text = keyword.text()

                    if keyword_text in keywords_set:
                        rules = [r for r in rules if r['keyword'] != keyword_text]

                    keywords_set.add(keyword_text)

                    rule = {
                        'keyword': keyword_text,
                        'reply': reply.text(),
                        'enabled': enabled
                    }
                    rules.append(rule)

            yuanbao_enabled = getattr(self, 'yuanbao_reply_switch', None) and self.yuanbao_reply_switch.isChecked()
            model_enabled = getattr(self, 'model_reply_switch', None) and self.model_reply_switch.isChecked()

            settings = {
                'rule_reply_enabled': self.rule_reply_switch.isChecked(),
                'reply_friend_enabled': self.reply_friend_switch.isChecked(),
                'reply_group_enabled': self.reply_group_switch.isChecked(),
                'specific_friend_enabled': self.specific_friend_switch.isChecked(),
                'specific_group_enabled': self.specific_group_switch.isChecked(),
                'ai_reply_enabled': self.ai_reply_switch.isChecked(),
                'yuanbao_reply_enabled': yuanbao_enabled,
                'model_reply_enabled': model_enabled,
                'new_friend_reply_enabled': getattr(self, 'new_friend_reply_switch', None) and self.new_friend_reply_switch.isChecked(),
                'fuzzy_match_enabled': self.fuzzy_match_switch.isChecked(),
                'exact_match_enabled': self.exact_match_switch.isChecked(),
                'min_interval': self.min_interval.text(),
                'max_interval': self.max_interval.text()
            }

            try:
                from aizhuli_combined import AIManager
                ai_manager = AIManager()

                if yuanbao_enabled:
                    ai_manager.auto_reply_settings['ai_reply_mode'] = 1
                elif model_enabled:
                    ai_manager.auto_reply_settings['ai_reply_mode'] = 0
            except Exception as e:
                pass
            new_friend_reply_content = ""
            for rule in rules:
                if rule['keyword'] == "我通过了你的朋友验证请求":
                    new_friend_reply_content = rule['reply']
                    break

            if new_friend_reply_content:
                settings['new_friend_reply_content'] = new_friend_reply_content

            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'rules': rules,
                    'settings': settings,
                    'last_update': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }, f, ensure_ascii=False, indent=2)

            if os.path.exists(target_file):
                os.replace(temp_file, target_file)
            else:
                os.rename(temp_file, target_file)

            self.statusBar().showMessage(f"规则数据已保存，共 {len(rules)} 条规则", 3000)

        except Exception as e:
            pass
        finally:
            self._is_saving = False

    def load_rules_data(self):
        try:
            self.rules_table.setRowCount(0)

            config_dir = "config"
            os.makedirs(config_dir, exist_ok=True)
            rules_file = os.path.join(config_dir, 'auto_reply_rules.json')

            if not os.path.exists(rules_file):
                with open(rules_file, 'w', encoding='utf-8') as f:
                    json.dump({
                        'rules': [],
                        'settings': {
                            'rule_reply_enabled': False,
                            'reply_friend_enabled': True,
                            'reply_group_enabled': True,
                            'specific_friend_enabled': False,
                            'specific_group_enabled': False,
                            'ai_reply_enabled': False,
                            'new_friend_reply_enabled': False,
                            'fuzzy_match_enabled': True,
                            'exact_match_enabled': False,
                            'min_interval': '1',
                            'max_interval': '5'
                        },
                        'last_update': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }, f, ensure_ascii=False, indent=2)
                return

            try:
                with open(rules_file, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    if not content:
                        with open(rules_file, 'w', encoding='utf-8') as f:
                            json.dump({
                                'rules': [],
                                'settings': {
                                    'rule_reply_enabled': False,
                                    'reply_friend_enabled': True,
                                    'reply_group_enabled': True,
                                    'specific_friend_enabled': False,
                                    'specific_group_enabled': False,
                                    'ai_reply_enabled': False,
                                    'new_friend_reply_enabled': False,
                                    'fuzzy_match_enabled': True,
                                    'exact_match_enabled': False,
                                    'min_interval': '1',
                                    'max_interval': '5'
                                },
                                'last_update': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }, f, ensure_ascii=False, indent=2)
                        return

                    data = json.loads(content)

                    rules = data.get('rules', [])

                    unique_rules = []
                    keywords_set = set()

                    for rule in rules:
                        keyword = rule.get('keyword', '')
                        if keyword in keywords_set:
                            unique_rules = [r for r in unique_rules if r.get('keyword') != keyword]

                        keywords_set.add(keyword)
                        unique_rules.append(rule)

                    for rule in unique_rules:
                        keyword = rule.get('keyword', '')
                        reply = rule.get('reply', '')
                        enabled = rule.get('enabled', True)
                        self.add_rule_to_table(keyword, reply, enabled)

                    settings = data.get('settings', {})
                    if settings:
                        self.rule_reply_switch.setChecked(settings.get('rule_reply_enabled', False))
                        self.reply_friend_switch.setChecked(settings.get('reply_friend_enabled', True))
                        self.reply_group_switch.setChecked(settings.get('reply_group_enabled', True))
                        self.specific_friend_switch.setChecked(settings.get('specific_friend_enabled', False))
                        self.specific_group_switch.setChecked(settings.get('specific_group_enabled', False))

                        self.ai_reply_switch.stateChanged.disconnect()

                        ai_enabled = settings.get('ai_reply_enabled', False)
                        self.ai_reply_switch.setChecked(ai_enabled)

                        self.ai_reply_switch.stateChanged.connect(self.on_ai_reply_switch)

                        try:
                            from aizhuli_combined import AIManager
                            ai_manager = AIManager()
                            ai_reply_mode = ai_manager.auto_reply_settings.get('ai_reply_mode', 1)

                            if hasattr(self, 'yuanbao_reply_switch'):
                                if not ai_enabled:
                                    self.yuanbao_reply_switch.setChecked(False)
                                    self.yuanbao_reply_switch.setEnabled(False)
                                else:
                                    yuanbao_enabled = settings.get('yuanbao_reply_enabled', ai_reply_mode == 1)
                                    self.yuanbao_reply_switch.setChecked(yuanbao_enabled)
                                    self.yuanbao_reply_switch.setEnabled(True)

                            if hasattr(self, 'model_reply_switch'):
                                if not ai_enabled:
                                    self.model_reply_switch.setChecked(False)
                                    self.model_reply_switch.setEnabled(False)
                                else:
                                    model_enabled = settings.get('model_reply_enabled', ai_reply_mode == 0)
                                    self.model_reply_switch.setChecked(model_enabled)
                                    self.model_reply_switch.setEnabled(True)

                            pass
                        except Exception as e:
                            if hasattr(self, 'yuanbao_reply_switch'):
                                self.yuanbao_reply_switch.setChecked(False)
                                self.yuanbao_reply_switch.setEnabled(False)
                            if hasattr(self, 'model_reply_switch'):
                                self.model_reply_switch.setChecked(False)
                                self.model_reply_switch.setEnabled(False)
                        if hasattr(self, 'new_friend_reply_switch'):
                            self.new_friend_reply_switch.setChecked(settings.get('new_friend_reply_enabled', False))

                        self.fuzzy_match_switch.setChecked(settings.get('fuzzy_match_enabled', True))
                        self.exact_match_switch.setChecked(settings.get('exact_match_enabled', False))
                        self.min_interval.setText(settings.get('min_interval', '1'))
                        self.max_interval.setText(settings.get('max_interval', '5'))

                    self.statusBar().showMessage(f"已加载 {len(unique_rules)} 条规则", 3000)

            except json.JSONDecodeError:
                with open(rules_file, 'w', encoding='utf-8') as f:
                    json.dump({
                        'rules': [],
                        'settings': {
                            'rule_reply_enabled': False,
                            'reply_friend_enabled': True,
                            'reply_group_enabled': True,
                            'specific_friend_enabled': False,
                            'specific_group_enabled': False,
                            'ai_reply_enabled': False,
                            'new_friend_reply_enabled': False,
                            'fuzzy_match_enabled': True,
                            'exact_match_enabled': False,
                            'min_interval': '2',
                            'max_interval': '5'
                        },
                        'last_update': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }, f, ensure_ascii=False, indent=2)
                self.statusBar().showMessage("规则配置文件格式错误，已重新初始化", 3000)

        except Exception as e:
            error_msg = f"加载规则数据失败: {str(e)}"
            QMessageBox.warning(self, "错误", error_msg, QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def on_rule_reply_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked
            self.statusBar().showMessage(f"规则回复功能已{'启用' if enabled else '禁用'}", 3000)
            self.mark_data_changed()
        except Exception as e:
            pass
    def on_reply_friend_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked
            self.statusBar().showMessage(f"回复好友消息功能已{'启用' if enabled else '禁用'}", 3000)
            self.mark_data_changed()
        except Exception as e:
            pass
    def on_reply_group_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked
            self.statusBar().showMessage(f"回复群消息功能已{'启用' if enabled else '禁用'}", 3000)
            self.mark_data_changed()
        except Exception as e:
            pass
    def on_specific_friend_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked
            self.statusBar().showMessage(f"指定好友回复功能已{'启用' if enabled else '禁用'}", 3000)
            if enabled:
                self.notebook.setCurrentIndex(0)
                QMessageBox.information(self, "提示", "请在主界面联系人列表中右键选择'指定回复' 加入指定好友列表")

            self.update_specify_reply_ui_visibility()
            self.mark_data_changed()
        except Exception as e:
            pass
    def on_specific_group_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked
            self.statusBar().showMessage(f"指定群回复功能已{'启用' if enabled else '禁用'}", 3000)
            if enabled:
                self.notebook.setCurrentIndex(0)
                QMessageBox.information(self, "提示", "请在主界面群列表中右键选择'指定回复' 加入指定群列表")

            self.update_specify_reply_ui_visibility()
            self.mark_data_changed()
        except Exception as e:
            pass
    def on_click_specify_friend(self):
        try:
            self.notebook.setCurrentIndex(0)
            QMessageBox.information(self, "提示", "请在主界面联系人列表中右键选择'指定回复' 加入指定好友列表")
        except Exception as e:
            pass
    def on_click_specify_group(self):
        try:
            self.notebook.setCurrentIndex(0)
            QMessageBox.information(self, "提示", "请在主界面群列表中右键选择'指定回复' 加入指定群列表")
        except Exception as e:
            pass

    def update_specific_selected_lists(self):
        try:
            if hasattr(self, 'specific_friend_list'):
                self.specific_friend_list.clear()
                for wxid in sorted(self.specific_friend_wxids):
                    name = wxid
                    try:
                        for c in getattr(self, 'all_contacts', []) or []:
                            if c.get('wxid') == wxid:
                                name = c.get('remarks') or c.get('nickname') or wxid
                                break
                    except Exception:
                        pass
                    item = QListWidgetItem(f"{name} ({wxid})")
                    item.setData(Qt.ItemDataRole.UserRole, wxid)
                    self.specific_friend_list.addItem(item)

            if hasattr(self, 'specific_group_list'):
                self.specific_group_list.clear()
                for wxid in sorted(self.specific_group_wxids):
                    name = wxid
                    try:
                        for c in getattr(self, 'all_contacts', []) or []:
                            if c.get('wxid') == wxid:
                                name = c.get('nickname') or wxid
                                break
                    except Exception:
                        pass
                    item = QListWidgetItem(f"{name} ({wxid})")
                    item.setData(Qt.ItemDataRole.UserRole, wxid)
                    self.specific_group_list.addItem(item)

            self.update_specify_reply_ui_visibility()
        except Exception as e:
            pass

    def remove_selected_specific_friend(self, item):
        try:
            wxid = item.data(Qt.ItemDataRole.UserRole)
            if wxid in self.specific_friend_wxids:
                self.specific_friend_wxids.remove(wxid)
                self.update_specific_selected_lists()
                self.mark_data_changed()
        except Exception as e:
            pass

    def remove_selected_specific_group(self, item):
        try:
            wxid = item.data(Qt.ItemDataRole.UserRole)
            if wxid in self.specific_group_wxids:
                self.specific_group_wxids.remove(wxid)
                self.update_specific_selected_lists()
                self.mark_data_changed()
        except Exception as e:
            pass

    def clear_specific_friends(self):
        try:
            self.specific_friend_wxids.clear()
            self.update_specific_selected_lists()
            self.mark_data_changed()

            self.update_specify_reply_ui_visibility()
        except Exception as e:
            pass

    def clear_specific_groups(self):
        try:
            self.specific_group_wxids.clear()
            self.update_specific_selected_lists()
            self.mark_data_changed()

            self.update_specify_reply_ui_visibility()
        except Exception as e:
            pass

    def update_specify_reply_ui_visibility(self):
        try:
            friend_enabled = False
            group_enabled = False
            try:
                friend_enabled = (hasattr(self, 'specific_friend_switch') and self.specific_friend_switch.isChecked())
            except Exception:
                pass
            try:
                group_enabled = (hasattr(self, 'specific_group_switch') and self.specific_group_switch.isChecked())
            except Exception:
                pass

            friend_has_items = False
            group_has_items = False
            try:
                if hasattr(self, 'specific_friend_list') and self.specific_friend_list:
                    friend_has_items = self.specific_friend_list.count() > 0
            except Exception:
                pass
            try:
                if hasattr(self, 'specific_group_list') and self.specific_group_list:
                    group_has_items = self.specific_group_list.count() > 0
            except Exception:
                pass

            try:
                if hasattr(self, 'specific_friend_list') and self.specific_friend_list:
                    visible = friend_enabled or friend_has_items
                    self.specific_friend_list.setVisible(visible)

                    parent_w = self.specific_friend_list.parent()
                    if parent_w and parent_w is not self:
                        parent_w.setVisible(visible)

                if hasattr(self, 'clear_specific_friend_btn'):
                    self.clear_specific_friend_btn.setVisible(friend_enabled or friend_has_items)

                friend_selector_names = [
                    'select_specific_friend_btn',
                    'specify_friend_btn',
                    'btn_select_specific_friend',
                    'btn_specify_friend',
                    'choose_specific_friend_btn'
                ]
                for _name in friend_selector_names:
                    try:
                        if hasattr(self, _name) and getattr(self, _name) is not None:
                            getattr(self, _name).setEnabled(friend_enabled)
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                if hasattr(self, 'specific_group_list') and self.specific_group_list:
                    visible = group_enabled or group_has_items
                    self.specific_group_list.setVisible(visible)
                    parent_w = self.specific_group_list.parent()
                    if parent_w and parent_w is not self:
                        parent_w.setVisible(visible)
                if hasattr(self, 'clear_specific_group_btn'):
                    self.clear_specific_group_btn.setVisible(group_enabled or group_has_items)

                group_selector_names = [
                    'select_specific_group_btn',
                    'specify_group_btn',
                    'btn_select_specific_group',
                    'btn_specify_group',
                    'choose_specific_group_btn'
                ]
                for _name in group_selector_names:
                    try:
                        if hasattr(self, _name) and getattr(self, _name) is not None:
                            getattr(self, _name).setEnabled(group_enabled)
                    except Exception:
                        pass
            except Exception:
                pass
        except Exception:
            pass

    def on_ai_reply_switch(self, state):
        enabled = state == Qt.CheckState.Checked.value

        if enabled:
            dialog = QDialog(self)
            dialog.setWindowTitle("选择AI回复类型")
            dialog.setModal(True)
            dialog.resize(300, 150)

            layout = QVBoxLayout(dialog)

            label = QLabel("请选择AI回复类型（必须选择其中一种）：")
            layout.addWidget(label)

            button_layout = QVBoxLayout()

            yuanbao_btn = QPushButton("元宝客服回复")
            yuanbao_btn.clicked.connect(lambda: self.select_ai_reply_type("yuanbao", dialog))
            button_layout.addWidget(yuanbao_btn)

            model_btn = QPushButton("大模型回复")
            model_btn.clicked.connect(lambda: self.select_ai_reply_type("model", dialog))
            button_layout.addWidget(model_btn)

            layout.addLayout(button_layout)

            cancel_btn = QPushButton("取消")
            cancel_btn.clicked.connect(lambda: self.cancel_ai_reply_selection(dialog))
            layout.addWidget(cancel_btn)

            dialog.exec()
        else:
            if hasattr(self, 'yuanbao_reply_switch'):
                self.yuanbao_reply_switch.setChecked(False)
                self.yuanbao_reply_switch.setEnabled(False)
            if hasattr(self, 'model_reply_switch'):
                self.model_reply_switch.setChecked(False)
                self.model_reply_switch.setEnabled(False)

            try:
                self._update_settings_in_rules_file({'ai_reply_enabled': False, 'yuanbao_reply_enabled': False, 'model_reply_enabled': False})
            except Exception:
                pass

        if hasattr(self, 'statusBar'):
            self.statusBar().showMessage(f"AI回复功能已{'启用' if enabled else '禁用'}", 3000)

    def select_ai_reply_type(self, reply_type, dialog):
        if reply_type == "yuanbao":
            if hasattr(self, 'yuanbao_reply_switch'):
                self.yuanbao_reply_switch.setChecked(True)
                self.yuanbao_reply_switch.setEnabled(True)
            if hasattr(self, 'model_reply_switch'):
                self.model_reply_switch.setChecked(False)
                self.model_reply_switch.setEnabled(True)
        elif reply_type == "model":
            if hasattr(self, 'model_reply_switch'):
                self.model_reply_switch.setChecked(True)
                self.model_reply_switch.setEnabled(True)
            if hasattr(self, 'yuanbao_reply_switch'):
                self.yuanbao_reply_switch.setChecked(False)
                self.yuanbao_reply_switch.setEnabled(True)
        dialog.accept()

    def cancel_ai_reply_selection(self, dialog):
        if hasattr(self, 'ai_reply_switch'):
            self.ai_reply_switch.setChecked(False)
        dialog.reject()

    def on_yuanbao_reply_switch(self, state):
        enabled = state == Qt.CheckState.Checked.value


        if hasattr(self, 'model_reply_switch') and enabled:
            self.model_reply_switch.setChecked(False)
            self.model_reply_switch.setEnabled(True)

        try:
            if hasattr(self, 'ai_reply_switch') and self.ai_reply_switch.isChecked():
                if hasattr(self, 'yuanbao_reply_switch'):
                    self.yuanbao_reply_switch.setEnabled(True)
                if hasattr(self, 'model_reply_switch'):
                    self.model_reply_switch.setEnabled(True)
                if not enabled:
                    if hasattr(self, 'model_reply_switch') and not self.model_reply_switch.isChecked():
                        self.model_reply_switch.setChecked(True)
        except Exception:
            pass

        try:
            from aizhuli_combined import AIManager
            ai_manager = AIManager()
            ai_manager.auto_reply_settings['ai_reply_mode'] = 1 if enabled else 0
        except Exception as e:
            pass
        try:
            ai_enabled = getattr(self, 'ai_reply_switch', None) and self.ai_reply_switch.isChecked()
            self._update_settings_in_rules_file({
                'ai_reply_enabled': bool(ai_enabled),
                'yuanbao_reply_enabled': bool(enabled),
                'model_reply_enabled': not enabled if ai_enabled else False,
                'ai_reply_mode': 1 if enabled else 0
            })
        except Exception:
            pass

        if hasattr(self, 'statusBar'):
            self.statusBar().showMessage(f"元宝客服回复功能已{'启用' if enabled else '禁用'}", 3000)

    def on_model_reply_switch(self, state):
        enabled = state == Qt.CheckState.Checked.value


        if hasattr(self, 'yuanbao_reply_switch') and enabled:
            self.yuanbao_reply_switch.setChecked(False)
            if hasattr(self, 'ai_reply_switch') and self.ai_reply_switch.isChecked():
                self.yuanbao_reply_switch.setEnabled(True)
            else:
                self.yuanbao_reply_switch.setEnabled(False)

        try:
            if hasattr(self, 'ai_reply_switch') and self.ai_reply_switch.isChecked():
                if hasattr(self, 'yuanbao_reply_switch'):
                    self.yuanbao_reply_switch.setEnabled(True)
                if hasattr(self, 'model_reply_switch'):
                    self.model_reply_switch.setEnabled(True)
                if not enabled:
                    if hasattr(self, 'yuanbao_reply_switch') and not self.yuanbao_reply_switch.isChecked():
                        self.yuanbao_reply_switch.setChecked(True)
        except Exception:
            pass

        try:
            from aizhuli_combined import AIManager
            ai_manager = AIManager()
            ai_manager.auto_reply_settings['ai_reply_mode'] = 0 if enabled else 1
        except Exception as e:
            pass
        try:
            ai_enabled = getattr(self, 'ai_reply_switch', None) and self.ai_reply_switch.isChecked()
            self._update_settings_in_rules_file({
                'ai_reply_enabled': bool(ai_enabled),
                'yuanbao_reply_enabled': not enabled if ai_enabled else False,
                'model_reply_enabled': bool(enabled),
                'ai_reply_mode': 0 if enabled else 1
            })
        except Exception:
            pass

        if hasattr(self, 'statusBar'):
            self.statusBar().showMessage(f"大模型回复功能已{'启用' if enabled else '禁用'}", 3000)

    def on_new_friend_reply_switch(self, state):
        try:
            enabled = state == Qt.CheckState.Checked

            if enabled:
                dialog = QDialog(self)
                dialog.setWindowTitle("设置新好友回复内容")
                dialog.setMinimumWidth(400)
                layout = QVBoxLayout(dialog)

                layout.addWidget(QLabel("请输入新好友通过验证后的自动回复内容:"))

                content_edit = QLineEdit()
                content_edit.setText("我通过了你的朋友验证请求，现在我们可以开始聊天了")
                content_edit.setPlaceholderText("请填写新好友通过后的回复内容...")
                layout.addWidget(content_edit)

                button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)

                if dialog.exec() == QDialog.DialogCode.Accepted:
                    reply_content = content_edit.text()
                    if not reply_content:
                        reply_content = "我通过了你的朋友验证请求，现在我们可以开始聊天了"

                    rule_exists = False
                    for row in range(self.rules_table.rowCount()):
                        keyword_item = self.rules_table.item(row, 2)
                        if keyword_item and keyword_item.text() == "我通过了你的朋友验证请求":
                            rule_exists = True

                            reply_item = self.rules_table.item(row, 3)
                            if reply_item:
                                reply_item.setText(reply_content)
                            break

                    if not rule_exists:
                        self.add_rule_to_table("我通过了你的朋友验证请求", reply_content, True)

                    QMessageBox.information(self, "新好友回复规则",
                                         "已添加/更新新好友回复规则，触发关键词为：我通过了你的朋友验证请求",
                                         QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                else:
                    self.new_friend_reply_switch.setChecked(False)
                    return

            self.mark_data_changed()
        except Exception as e:
            pass

    def on_fuzzy_match_switch(self, _):
        try:
            enabled = self.fuzzy_match_switch.isChecked()

            if enabled:
                self.exact_match_switch.setChecked(False)

            self.statusBar().showMessage(f"模糊匹配功能已{'启用' if enabled else '禁用'}", 3000)

            self.mark_data_changed()
        except Exception as e:
            pass

    def on_exact_match_switch(self, _):
        try:
            enabled = self.exact_match_switch.isChecked()

            if enabled:
                self.fuzzy_match_switch.setChecked(False)

            self.statusBar().showMessage(f"精准匹配功能已{'启用' if enabled else '禁用'}", 3000)

            self.mark_data_changed()
        except Exception as e:
            pass

    def mark_data_changed(self):
        self.data_changed = True

        self.data_save_timer.start(2000)

    def show_context_menu(self, pos):
        menu = QMenu(self)

        actions = [
            ("全部选择", self.select_all_rules),
            ("反向选择", self.invert_selection),
            ("勾选选中", self.check_selected),
            ("删除选中", self.delete_selected_rules),
            ("清空列表", self.clear_rules)
        ]

        for text, slot in actions:
            action = menu.addAction(text)
            action.triggered.connect(slot)

        menu.exec(self.rules_table.viewport().mapToGlobal(pos))

    def select_all_rules(self):
        try:
            for row in range(self.rules_table.rowCount()):
                checkbox = self.rules_table.cellWidget(row, 1).findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
            self.statusBar().showMessage("已全部选择规则", 3000)
        except Exception as e:
            pass
    def invert_selection(self):
        try:
            for row in range(self.rules_table.rowCount()):
                checkbox = self.rules_table.cellWidget(row, 1).findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(not checkbox.isChecked())
            self.statusBar().showMessage("已反向选择规则", 3000)
        except Exception as e:
            pass
    def check_selected(self):
        try:
            selected_rows = set(item.row() for item in self.rules_table.selectedItems())
            for row in selected_rows:
                checkbox = self.rules_table.cellWidget(row, 1).findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
            self.statusBar().showMessage(f"已勾选 {len(selected_rows)} 条选中规则", 3000)
        except Exception as e:
            pass
    def delete_selected_rules(self):
        try:
            checked_rows = []
            for row in range(self.rules_table.rowCount()):
                cell_widget = self.rules_table.cellWidget(row, 1)
                if cell_widget:
                    checkbox = cell_widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        checked_rows.append(row)

            target_rows = checked_rows if checked_rows else list(set(item.row() for item in self.rules_table.selectedItems()))

            if not target_rows:
                return

            target_rows = sorted(set(target_rows), reverse=True)

            reply = QMessageBox.question(
                self,
                "确认删除",
                f"确定要删除选中的 {len(target_rows)} 条规则吗？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                for row in target_rows:
                    self.rules_table.removeRow(row)

                self.update_row_numbers()

                self.save_rules_data()

                self.statusBar().showMessage(f"已删除 {len(target_rows)} 条规则", 3000)

        except Exception as e:
            pass
    def clear_rules(self):
        try:
            if self.rules_table.rowCount() == 0:
                return

            reply = QMessageBox.question(
                self,
                "确认清空",
                "确定要清空所有规则吗？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                self.rules_table.setRowCount(0)
                self.save_rules_data()
                self.statusBar().showMessage("已清空所有规则", 3000)

        except Exception as e:
            pass
    def update_row_numbers(self):
        try:
            for row in range(self.rules_table.rowCount()):
                item = QTableWidgetItem(str(row + 1))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                self.rules_table.setItem(row, 0, item)
        except Exception as e:
            pass
    def on_data_changed(self, _):
        try:
            self.data_save_timer.start(1000)
        except Exception as e:
            pass
    def add_reply_rule(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("添加回复规则")
            dialog.setMinimumWidth(400)
            layout = QVBoxLayout(dialog)

            keyword_layout = QHBoxLayout()
            keyword_layout.addWidget(QLabel("触发关键词:"))
            keyword_edit = QLineEdit()
            keyword_edit.setPlaceholderText("输入触发关键词...")
            keyword_layout.addWidget(keyword_edit)
            layout.addLayout(keyword_layout)

            reply_layout = QVBoxLayout()
            reply_layout.addWidget(QLabel("回复内容:"))
            reply_edit = QTextEdit()
            reply_edit.setPlaceholderText("输入回复内容...")
            reply_layout.addWidget(reply_edit)
            layout.addLayout(reply_layout)

            btn_layout = QHBoxLayout()
            ok_btn = QPushButton("确定")
            cancel_btn = QPushButton("取消")

            ok_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            btn_layout.addWidget(ok_btn)
            btn_layout.addWidget(cancel_btn)
            layout.addLayout(btn_layout)

            dialog.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)
            dialog.setWindowModality(Qt.WindowModality.ApplicationModal)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                keyword = keyword_edit.text()
                reply = reply_edit.toPlainText()

                if not keyword or not reply:
                    QMessageBox.warning(self, "输入错误", "关键词和回复内容不能为空！", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                    return

                self.add_rule_to_table(keyword, reply)
                self.statusBar().showMessage(f"已添加自动回复规则: {keyword}", 3000)
                self.save_rules_data()

        except Exception as e:
            pass
    def add_rule_to_table(self, keyword="", reply="", enabled=True):
        try:
            row = self.rules_table.rowCount()
            self.rules_table.insertRow(row)

            number_item = QTableWidgetItem(str(row + 1))
            number_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            number_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.rules_table.setItem(row, 0, number_item)

            checkbox_widget = self.create_rule_checkbox(row, enabled)
            self.rules_table.setCellWidget(row, 1, checkbox_widget)

            keyword_item = QTableWidgetItem(str(keyword))
            keyword_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.rules_table.setItem(row, 2, keyword_item)

            reply_item = QTableWidgetItem(str(reply))
            reply_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.rules_table.setItem(row, 3, reply_item)

        except Exception as e:
            pass
    def create_rule_checkbox(self, row, enabled=True):
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        checkbox = QCheckBox()
        checkbox.setChecked(enabled)
        checkbox.stateChanged.connect(lambda state, r=row: self.on_checkbox_changed(r, state))
        layout.addWidget(checkbox)

        return container

    def on_checkbox_changed(self, *_):
        try:
            self.save_rules_data()
        except Exception as e:
            pass
    def import_rules(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "导入规则",
                "",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            if file_path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active

                data = []
                headers = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # 表头行
                        headers = [str(cell) if cell else '' for cell in row]
                    else:  # 数据行
                        row_data = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx < len(headers):
                                row_data[headers[col_idx]] = str(cell) if cell else ''
                        if row_data:
                            data.append(row_data)

                required_columns = ['关键词', '回复内容']
                if not all(col in headers for col in required_columns):
                    QMessageBox.warning(self, "导入失败", "Excel文件必须包含'关键词'和'回复内容'列", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                    return

                processed_data = []
                for row_data in data:
                    keyword = str(row_data.get('关键词', '')).strip()
                    reply = str(row_data.get('回复内容', '')).strip()
                    enabled = bool(row_data.get('启用状态', True)) if '启用状态' in row_data else True

                    if keyword and reply:
                        processed_data.append({
                            '关键词': keyword,
                            '回复内容': reply,
                            '启用状态': enabled
                        })

                before_drop = len(processed_data)
                unique_data = {}
                for item in processed_data:
                    if item['关键词'] not in unique_data:
                        unique_data[item['关键词']] = item
                processed_data = list(unique_data.values())
                dup_in_file = before_drop - len(processed_data)

                existing_keywords = set()
                try:
                    for row in range(self.rules_table.rowCount()):
                        item = self.rules_table.item(row, 2)
                        if item:
                            existing_keywords.add(item.text().strip())
                except Exception:
                    pass

                final_data = [item for item in processed_data if item['关键词'] not in existing_keywords]
                dup_exists = len(processed_data) - len(final_data)

                imported_count = 0
                for item in final_data:
                    self.add_rule_to_table(item['关键词'], item['回复内容'], item['启用状态'])
                    imported_count += 1

                try:
                    self.save_rules_data()
                except Exception:
                    pass
                skipped_total = dup_in_file + dup_exists
                self.statusBar().showMessage(
                    f"已从{file_path}导入{imported_count}条规则，跳过重复{skipped_total}条（文件内重复{dup_in_file}，与现有重复{dup_exists}）",
                    5000
                )
                QMessageBox.information(
                    self,
                    "导入完成",
                    f"成功导入{imported_count}条规则\n跳过重复{skipped_total}条\n（文件内重复{dup_in_file}，与现有重复{dup_exists}）",
                    QMessageBox.StandardButton.Ok,
                    QMessageBox.StandardButton.Ok
                )
            else:
                QMessageBox.warning(self, "导入失败", "仅支持Excel格式(.xlsx)的规则导入", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"导入规则失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def export_rules(self):
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出规则",
                f"自动回复规则_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            rules_data = []
            for row in range(self.rules_table.rowCount()):
                checkbox_widget = self.rules_table.cellWidget(row, 1)
                enabled = False
                if checkbox_widget:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox:
                        enabled = checkbox.isChecked()

                keyword = self.rules_table.item(row, 2)
                reply = self.rules_table.item(row, 3)

                if keyword and reply:
                    rule_data = {
                        '关键词': keyword.text(),
                        '回复内容': reply.text(),
                        '启用状态': enabled
                    }
                    rules_data.append(rule_data)

            settings_data = [{
                '设置项': '规则回复启用',
                '状态': self.rule_reply_switch.isChecked()
            }, {
                '设置项': '好友回复启用',
                '状态': self.reply_friend_switch.isChecked()
            }, {
                '设置项': '群组回复启用',
                '状态': self.reply_group_switch.isChecked()
            }, {
                '设置项': '指定好友启用',
                '状态': self.specific_friend_switch.isChecked()
            }, {
                '设置项': '指定群组启用',
                '状态': self.specific_group_switch.isChecked()
            }, {
                '设置项': 'AI回复启用',
                '状态': self.ai_reply_switch.isChecked()
            }, {
                '设置项': '元宝回复启用',
                '状态': getattr(self, 'yuanbao_reply_switch', None) and self.yuanbao_reply_switch.isChecked()
            }, {
                '设置项': '模型回复启用',
                '状态': getattr(self, 'model_reply_switch', None) and self.model_reply_switch.isChecked()
            }, {
                '设置项': '新好友回复启用',
                '状态': getattr(self, 'new_friend_reply_switch', None) and self.new_friend_reply_switch.isChecked()
            }, {
                '设置项': '模糊匹配启用',
                '状态': self.fuzzy_match_switch.isChecked()
            }, {
                '设置项': '精确匹配启用',
                '状态': self.exact_match_switch.isChecked()
            }, {
                '设置项': '最小间隔',
                '状态': self.min_interval.text()
            }, {
                '设置项': '最大间隔',
                '状态': self.max_interval.text()
            }]

            data = {
                "自动回复规则": rules_data,
                "功能设置": settings_data
            }

            from openpyxl import Workbook
            wb = Workbook()

            wb.remove(wb.active)

            for sheet_name, sheet_data in data.items():
                if sheet_name and sheet_data:
                    ws = wb.create_sheet(title=sheet_name)

                    if sheet_data:
                        headers = list(sheet_data[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)

                        for row_idx, row_data in enumerate(sheet_data, 2):
                            for col_idx, header in enumerate(headers, 1):
                                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

            wb.save(file_path)

            self.statusBar().showMessage(f"已导出规则到: {file_path}", 3000)
            QMessageBox.information(self, "导出成功", "规则导出成功！", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出规则失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def clear_reply_history(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("清空接收消息")
            layout = QVBoxLayout(dialog)

            layout.addWidget(QLabel("确定要清空所有回复接收消息？"))

            reset_counter_checkbox = QCheckBox("重置序号计数器（否则序号将继续递增）")
            reset_counter_checkbox.setChecked(True)
            layout.addWidget(reset_counter_checkbox)

            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.auto_reply_history_table.setRowCount(0)
                self.auto_reply_history_data.clear()

                if reset_counter_checkbox.isChecked():
                    self.auto_reply_message_counter = 0
                    self.statusBar().showMessage("已清空自动回复接收消息并重置序号", 3000)
                else:
                    self.statusBar().showMessage("已清空自动回复接收消息，序号将继续递增", 3000)

        except Exception as e:
            pass
    def add_message_to_auto_reply_history(self, message_data):
        try:
            pass

            if not hasattr(self, 'auto_reply_history_table') or self.auto_reply_history_table is None:
                return

            row = self.auto_reply_history_table.rowCount()
            self.auto_reply_history_table.insertRow(row)

            self.auto_reply_message_counter += 1

            is_group_message = "@chatroom" in message_data.get('sender_wxid', '')

            id_column_value = message_data.get('sender_wxid', '')
            if is_group_message and 'member_name' in message_data:
                id_column_value = message_data.get('member_name', '')

            columns = [
                (0, str(self.auto_reply_message_counter)),
                (1, message_data.get('self_nickname', '')),
                (2, message_data.get('sender_nickname', '')),
                (3, id_column_value),
                (4, message_data.get('content', '')),
                (5, message_data.get('receive_time', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            ]

            for col, value in columns:
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.auto_reply_history_table.setItem(row, col, item)

            self.auto_reply_history_table.scrollToBottom()

            self.auto_reply_history_data.append(message_data)
            if is_group_message and 'member_name' in message_data:
                pass
            msg_source = message_data.get('sender_nickname', '')
            if is_group_message and 'member_name' in message_data:
                msg_source += f" ({message_data.get('member_name', '')})"
            self.statusBar().showMessage(f"收到来自 {msg_source} 的消息", 3000)

        except Exception as e:
            pass
    def process_message_for_auto_reply(self, message):
        try:
            timestamp = message.get("timestamp", int(time.time()))
            time_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')

            wxid = message.get("wxid", "")
            content = message.get("content", "")

            if wxid == "wxid_wi_1d142z0zdj03":
                pass

                try:
                    import psutil
                    current_pid = None
                    for proc in psutil.process_iter(['pid', 'name']):
                        if proc.info['name'].lower() == 'wechat.exe':
                            current_pid = proc.info['pid']
                            break

                    if not current_pid:
                        pass
                        return

                    found_sender, found_chatroom = get_last_yuanbao_sender()

                    if not found_sender:
                        pass
                        return

                    yuanbao_content = message.get('content', '')

                    pass
                    if found_chatroom:
                        pass
                    else:
                        pass

                    if found_sender:
                        pass

                        if found_chatroom and '@chatroom' in found_chatroom:
                            pass

                            send_result = send_message_simple(current_pid, found_chatroom, yuanbao_content)

                            if send_result:
                                pass
                            else:
                                self._handle_send_failure_simple(current_pid, found_chatroom, yuanbao_content, is_group=True)
                        elif '@chatroom' in found_sender:
                            pass

                            send_result = send_message_simple(current_pid, found_sender, yuanbao_content)

                            if send_result:
                                pass
                            else:
                                self._handle_send_failure_simple(current_pid, found_sender, yuanbao_content, is_group=True)
                        else:
                            pass

                            send_result = send_message_simple(current_pid, found_sender, yuanbao_content)

                            if send_result:
                                pass
                            else:
                                self._handle_send_failure_simple(current_pid, found_sender, yuanbao_content, is_group=False)
                    else:
                        pass
                except Exception as e:
                    pass
                return

            content = message.get("content", "")
            original_content = content
            account_info = message.get("account", {})
            account_name = account_info.get("nickname", "未知账号")

            parsed_content = parse_special_message(content)
            if parsed_content:
                content = parsed_content

            is_group_message = "@chatroom" in wxid
            member_id = message.get("member_id", "")

            is_at_me = False

            contact_info = None
            if hasattr(self, 'all_contacts') and self.all_contacts:
                for contact in self.all_contacts:
                    if contact.get("wxid") == wxid:
                        contact_info = contact
                        break

            sender_name = wxid
            if contact_info:
                if contact_info.get("remarks") and contact_info.get("remarks").strip():
                    sender_name = contact_info.get("remarks")
                else:
                    sender_name = contact_info.get("nickname", wxid)

            if is_group_message and member_id:
                member_info = None
                if hasattr(self, 'all_contacts') and self.all_contacts:
                    for contact in self.all_contacts:
                        if contact.get("wxid") == member_id:
                            member_info = contact
                            break

                member_name = member_id
                if member_info:
                    member_name = member_info.get("nickname", member_id)

            if is_group_message:
                self_nickname = account_info.get('nickname', '')
                self_wxid = account_info.get('wxid', '')
                if "<atuserlist>" in content:
                    is_at_me = True
                elif self_nickname and f"@{self_nickname}" in content:
                    is_at_me = True
                    content = content.replace(f"@{self_nickname}", "").strip()
                elif self_wxid and f"@{self_wxid}" in content:
                    is_at_me = True
                    content = content.replace(f"@{self_wxid}", "").strip()

            message_data = {
                'self_nickname': account_name,
                'sender_nickname': sender_name,
                'sender_wxid': wxid,
                'content': content,
                'original_content': original_content,
                'receive_time': time_str,
                'account': account_info,
                'is_at_me': is_at_me if is_group_message else False
            }

            if is_group_message and member_id:
                member_name = member_id
                if member_info:
                    member_name = member_info.get("nickname", member_id)
                message_data['member_name'] = member_name
                message_data['member_id'] = member_id

            self.process_auto_reply(message_data)

            self.statusBar().showMessage(f"收到新消息: {content[:20]}...", 3000)

        except Exception as e:
            pass

    def format_ai_response(self, response, settings):
        try:
            rules_enabled = settings.get('rules_enabled', True)
            prefix = settings.get('reply_prefix', '')
            if rules_enabled and prefix:
                response = f"{prefix} {response}"

            token_limit = settings.get('model_token_limit', 0)
            if token_limit > 0 and len(response) > token_limit:
                response = response[:token_limit-3] + "..."

            return response

        except Exception as e:
            return response

    def send_delayed_reply(self, receiver_wxid, response, reply_type="ai_reply", pid=None):
        try:
            try:
                min_delay = max(0, int(self.min_interval.text()))
                max_delay = max(min_delay, int(self.max_interval.text()))
            except ValueError:
                min_delay = 1
                max_delay = 5

            if min_delay == max_delay:
                delay = min_delay
            else:
                import random
                delay = random.randint(min_delay, max_delay)



            if pid:
                pass

            QTimer.singleShot(
                delay * 1000,
                lambda: self.send_auto_reply_with_type(receiver_wxid, response, reply_type, pid=pid)
            )

        except Exception as e:
            pass
    def send_auto_reply_with_type(self, receiver_wxid, content, reply_type="auto_reply", pid=None):
        try:
            success = self.send_auto_reply(pid, receiver_wxid, content)

            if success:
                self.save_reply_message_to_ini(receiver_wxid, content, reply_type)
            else:
                pass
            return success

        except Exception as e:
            return False

    def save_reply_message_to_ini(self, receiver_wxid, content, reply_type="auto_reply"):
        try:
            import configparser
            from datetime import datetime

            config_dir = "config"
            os.makedirs(config_dir, exist_ok=True)
            messages_file = os.path.join(config_dir, "messages.ini")

            config = configparser.ConfigParser()
            if os.path.exists(messages_file):
                config.read(messages_file, encoding='utf-8')

            existing_sections = [s for s in config.sections() if s.startswith('message_')]
            next_id = len(existing_sections)
            section_name = f"message_{next_id:06d}"

            config[section_name] = {
                'content': content,
                'sender': 'system',
                'receiver': receiver_wxid,
                'timestamp': datetime.now().isoformat(),
                'type': reply_type,
                'direction': 'outgoing'
            }

            with open(messages_file, 'w', encoding='utf-8') as f:
                config.write(f)
        except Exception as e:
            pass
    def send_auto_reply(self, pid, receiver_wxid, content):
        try:
            current_pid = pid if pid else self._get_wechat_pid()
            if not current_pid:
                return False

            if not receiver_wxid or not receiver_wxid.strip():
                return False

            import os
            file_path = content.strip()

            if os.path.exists(file_path) and os.path.isfile(file_path):

                _, file_extension = os.path.splitext(file_path)
                file_extension = file_extension.lower()

                image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp']

                if file_extension in image_extensions:
                    success = send_image_to_wxid(current_pid, receiver_wxid, file_path)

                    if success:
                        self.statusBar().showMessage(f"已发送图片到 {receiver_wxid}", 3000)
                    else:
                        self.statusBar().showMessage(f"发送图片失败", 3000)

                    return success
                else:
                    try:

                        receiver_name = receiver_wxid
                        if hasattr(self, 'all_contacts') and self.all_contacts:
                            for contact in self.all_contacts:
                                if contact.get("wxid") == receiver_wxid:
                                    receiver_name = contact.get("nickname", receiver_wxid)
                                    break

                        dialog = SendMessageDialog(receiver_name, self, receiver_wxid, pid=current_pid)

                        dialog.image_path_edit.setText(file_path)

                        success = dialog.send_image()

                        if success:
                            self.statusBar().showMessage(f"已发送文件到 {receiver_wxid}", 3000)
                            return True
                        else:
                            self.statusBar().showMessage(f"发送文件失败", 3000)
                            return False
                    except Exception as e:
                        self.statusBar().showMessage(f"发送文件失败: {str(e)}", 3000)
                        return False

            is_group_message = "@chatroom" in receiver_wxid

            if is_group_message:
                print(f"发送消息到群聊: {receiver_wxid}，不添加@前缀")

            print(f"准备发送消息 - PID: {current_pid}, 接收者: {receiver_wxid}, 内容: '{content}'")

            print(f"正在发送回复消息到 {receiver_wxid}: '{content}'")
            success = send_message_to_wxid(current_pid, receiver_wxid, content)

            if success:
                self.statusBar().showMessage(f"已发送自动回复消息到 {receiver_wxid}", 3000)

                self.save_reply_message_to_ini(receiver_wxid, content, "auto_reply")
            else:
                self.statusBar().showMessage(f"发送自动回复消息失败", 3000)

            return success
        except Exception as e:
            pass
            return False

    def show_ai_reply_settings(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("大模型回复规则设置")
            dialog.setFixedSize(450, 400)  
            dialog.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)

            layout = QVBoxLayout(dialog)

            model_layout = QFormLayout()

            try:
                from aizhuli_combined import AIManager
                ai_manager = AIManager()
                current_settings = ai_manager.auto_reply_settings
            except:
                current_settings = {}

            rules_layout = QHBoxLayout()
            self.rules_enabled_check = QCheckBox("启用风格/表情/前缀等规则")
            self.rules_enabled_check.setChecked(current_settings.get('rules_enabled', True))
            rules_layout.addWidget(self.rules_enabled_check)
            rules_layout.addStretch()
            layout.addLayout(rules_layout)

            length_layout = QHBoxLayout()
            length_label = QLabel("回复长度:")
            length_label.setFixedWidth(100)
            length_layout.addWidget(length_label)

            self.length_combo = QComboBox()
            length_options = [
                ("不限制长度", "unlimited"),
                ("自定义长度", "custom")
            ]

            for name, value in length_options:
                self.length_combo.addItem(name, value)

            current_length_type = current_settings.get('length_type', 'unlimited')
            for i in range(self.length_combo.count()):
                if self.length_combo.itemData(i) == current_length_type:
                    self.length_combo.setCurrentIndex(i)
                    break
            else:
                self.length_combo.setCurrentIndex(0)  

            length_layout.addWidget(self.length_combo)

            self.length_input = QLineEdit()
            self.length_input.setPlaceholderText("输入目标字数，如：500")
            current_token_limit = current_settings.get('model_token_limit', 500)
            if current_token_limit > 0:
                self.length_input.setText(str(current_token_limit))
            else:
                self.length_input.setText("500")
            self.length_input.setVisible(current_length_type == "custom")  # 根据当前设置显示/隐藏

            def auto_save_custom_length():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    try:
                        token_limit = int(self.length_input.text()) if self.length_input.text() else 500
                        token_limit = max(1, min(10000, token_limit))
                    except ValueError:
                        token_limit = 500

                    ai_manager.auto_reply_settings['model_token_limit'] = token_limit
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            self.length_input.textChanged.connect(auto_save_custom_length)

            length_layout.addWidget(self.length_input)

            layout.addLayout(length_layout)

            style_layout = QHBoxLayout()
            style_label = QLabel("回复风格:")
            style_label.setFixedWidth(100)
            style_layout.addWidget(style_label)

            self.style_combo = QComboBox()
            self.style_combo.addItems([
                "友好亲切", "自定义回复", "专业正式", "幽默风趣", "简洁明了",
                "详细解答", "温暖贴心"
            ])
            current_style = current_settings.get('model_style', '友好亲切')
            if current_style in [self.style_combo.itemText(i) for i in range(self.style_combo.count())]:
                self.style_combo.setCurrentText(current_style)
            style_layout.addWidget(self.style_combo)
            layout.addLayout(style_layout)

            custom_style_layout = QVBoxLayout()
            self.custom_style_label = QLabel("自定义风格:")
            self.custom_style_label.setStyleSheet("font-weight: bold; margin: 0px; padding: 0px;")
            self.custom_style_label.setMaximumHeight(20)
            custom_style_layout.addWidget(self.custom_style_label)
            custom_style_layout.setSpacing(2) 

            self.custom_style_input = QTextEdit()
            self.custom_style_input.setPlaceholderText("请输入自定义回复风格的详细描述，例如：\n\n• 正式商务风格，使用专业的语言和格式\n• 轻松友好风格，使用口语化表达\n• 教学风格，详细解释和举例说明")
            self.custom_style_input.setMaximumHeight(120)
            self.custom_style_input.setMinimumHeight(60)
            self.custom_style_input.setText(current_settings.get('custom_style', ''))

            def auto_save_custom_style():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    custom_style_text = self.custom_style_input.toPlainText()
                    ai_manager.auto_reply_settings['custom_style'] = custom_style_text
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            self.custom_style_input.textChanged.connect(auto_save_custom_style)

            custom_style_layout.addWidget(self.custom_style_input)

            layout.addLayout(custom_style_layout)

            is_custom_initial = current_style == "自定义回复"
            self.custom_style_label.setVisible(is_custom_initial)
            self.custom_style_input.setVisible(is_custom_initial)

            emoji_layout = QHBoxLayout()
            emoji_label = QLabel("表情设置:")
            emoji_label.setFixedWidth(100)
            emoji_layout.addWidget(emoji_label)
            self.include_emoji_combo = QComboBox()
            self.include_emoji_combo.addItems(["不添加表情", "添加适量表情", "丰富表情表达"])
            current_emoji = current_settings.get('include_emoji', True)
            if current_emoji:
                self.include_emoji_combo.setCurrentIndex(1)  # 默认选择"添加适量表情"
            else:
                self.include_emoji_combo.setCurrentIndex(0)  # 不添加表情
            emoji_layout.addWidget(self.include_emoji_combo)
            layout.addLayout(emoji_layout)

            params_layout = QHBoxLayout()

            creativity_label = QLabel("创意度:")
            creativity_label.setFixedWidth(80)
            params_layout.addWidget(creativity_label)
            self.temperature_input = QLineEdit(str(current_settings.get('temperature', 0.7)))
            self.temperature_input.setPlaceholderText("0.1-1.0")
            self.temperature_input.setFixedWidth(60)

            def auto_save_temperature():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    try:
                        temperature_value = float(self.temperature_input.text()) if self.temperature_input.text() else 0.7
                        temperature_value = max(0.1, min(1.0, temperature_value))
                    except ValueError:
                        temperature_value = 0.7

                    ai_manager.auto_reply_settings['temperature'] = temperature_value
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            self.temperature_input.textChanged.connect(auto_save_temperature)

            params_layout.addWidget(self.temperature_input)

            params_layout.addStretch()

            layout.addLayout(params_layout)

            prefix_layout = QHBoxLayout()
            prefix_label = QLabel("回复前缀:")
            prefix_label.setFixedWidth(100)
            prefix_layout.addWidget(prefix_label)
            self.reply_prefix_input = QLineEdit()
            self.reply_prefix_input.setText(current_settings.get('reply_prefix', ''))
            self.reply_prefix_input.setPlaceholderText("例如：[AI助手] 或留空")

            def auto_save_reply_prefix():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    prefix_text = self.reply_prefix_input.text()
                    ai_manager.auto_reply_settings['reply_prefix'] = prefix_text
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            self.reply_prefix_input.textChanged.connect(auto_save_reply_prefix)

            prefix_layout.addWidget(self.reply_prefix_input)
            layout.addLayout(prefix_layout)

            def on_style_changed():
                is_custom = self.style_combo.currentText() == "自定义回复"
                self.custom_style_label.setVisible(is_custom)
                self.custom_style_input.setVisible(is_custom)

            def on_length_changed():
                is_custom_length = self.length_combo.currentData() == "custom"
                self.length_input.setVisible(is_custom_length)

            def auto_save_style():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    style_text = self.style_combo.currentText()
                    ai_manager.auto_reply_settings['model_style'] = style_text
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            def auto_save_emoji():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    emoji_setting = self.include_emoji_combo.currentText()
                    include_emoji = emoji_setting != "不添加表情"
                    ai_manager.auto_reply_settings['include_emoji'] = include_emoji
                    ai_manager.auto_reply_settings['emoji_level'] = emoji_setting
                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            def auto_save_length():
                try:
                    from aizhuli_combined import AIManager
                    ai_manager = AIManager()
                    length_type = self.length_combo.currentData()
                    ai_manager.auto_reply_settings['length_type'] = length_type

                    if length_type == "custom" and hasattr(self, 'length_input'):
                        try:
                            token_limit = int(self.length_input.text()) if self.length_input.text() else 500
                        except ValueError:
                            token_limit = 500
                        ai_manager.auto_reply_settings['model_token_limit'] = token_limit
                    else:
                        ai_manager.auto_reply_settings['model_token_limit'] = 0

                    ai_manager.save_auto_reply_rules()
                except Exception as e:
                    pass
            self.style_combo.currentTextChanged.connect(on_style_changed)
            self.style_combo.currentTextChanged.connect(auto_save_style)

            self.length_combo.currentTextChanged.connect(on_length_changed)
            self.length_combo.currentTextChanged.connect(auto_save_length)

            self.include_emoji_combo.currentTextChanged.connect(auto_save_emoji)

            on_style_changed()
            on_length_changed()

            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.save_ai_reply_settings()

        except Exception as e:
            QMessageBox.warning(self, "错误", f"显示设置对话框失败: {str(e)}",
                              QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def save_ai_reply_settings(self):
        try:
            from aizhuli_combined import AIManager
            ai_manager = AIManager()

            length_type = self.length_combo.currentData()
            if length_type == "custom":
                try:
                    token_limit = int(self.length_input.text()) if self.length_input.text() else 500
                except ValueError:
                    token_limit = 500
            else:
                token_limit = 0  # 不限制长度

            emoji_setting = self.include_emoji_combo.currentText()
            include_emoji = emoji_setting != "不添加表情"

            ai_manager.auto_reply_settings.update({
                'model_token_limit': token_limit,
                'length_type': length_type,
                'model_style': self.style_combo.currentText(),
                'custom_style': self.custom_style_input.toPlainText(),
                'temperature': float(self.temperature_input.text()) if self.temperature_input.text() else 0.7,
                'include_emoji': include_emoji,
                'emoji_level': emoji_setting,
                'reply_prefix': self.reply_prefix_input.text(),
                'rules_enabled': self.rules_enabled_check.isChecked()
            })

            try:
                ai_manager.save_auto_reply_rules()
            except Exception as _e:
                pass
            ai_rules = None

            try:
                if hasattr(self, 'ai_rules_table') and self.ai_rules_table is not None:
                    tmp_rules = []
                    for row in range(self.ai_rules_table.rowCount()):
                        checkbox = self.ai_rules_table.cellWidget(row, 0)
                        priority_item = self.ai_rules_table.item(row, 1)
                        pattern_item = self.ai_rules_table.item(row, 2)
                        match_combo = self.ai_rules_table.cellWidget(row, 3)
                        template_item = self.ai_rules_table.item(row, 4)

                        if all([checkbox, priority_item, pattern_item, match_combo, template_item]):
                            rule = {
                                'enabled': checkbox.isChecked(),
                                'priority': int(priority_item.text()),
                                'pattern': pattern_item.text(),
                                'match_type': match_combo.currentText(),
                                'template': template_item.text()
                            }
                            tmp_rules.append(rule)
                    ai_rules = tmp_rules
            except Exception:
                ai_rules = None

            config_dir = "config"
            os.makedirs(config_dir, exist_ok=True)
            rules_file = os.path.join(config_dir, "auto_reply_rules.json")

            existing_data = {}
            if os.path.exists(rules_file):
                try:
                    with open(rules_file, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                except:
                    existing_data = {}

            existing_settings = existing_data.get('settings', {}) if isinstance(existing_data, dict) else {}
            try:
                if isinstance(existing_settings, dict):
                    existing_settings.update(ai_manager.auto_reply_settings)
            except Exception:
                pass

            existing_data['settings'] = existing_settings
            if ai_rules is not None:
                existing_data['ai_rules'] = ai_rules
            existing_data['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            existing_data['version'] = '1.0'

            with open(rules_file, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)

            ai_manager.save_auto_reply_rules()

            QMessageBox.information(self, "成功", "AI回复设置已保存。",
                                 QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            try:
                self._cleanup_obsolete_ai_config()
            except Exception:
                pass

        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存设置失败: {str(e)}",
                              QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def show_doc_training_dialog(self):
        """显示上传文档资料大模型训练回复话术对话框"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("上传文档资料大模型训练回复话术")
            dialog.setFixedSize(500, 400)
            dialog.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)

            layout = QVBoxLayout(dialog)

            info_label = QLabel("此功能正在开发中，敬请期待！")
            info_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2196F3; margin: 10px;")
            info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(info_label)

            desc_text = QTextEdit()
            desc_text.setReadOnly(True)
            desc_text.setPlainText("""功能规划：

1. 文档上传功能
   • 支持多种格式：PDF、Word、TXT、Excel等
   • 批量上传多个文档
   • 文档内容自动解析和提取

2. 智能训练功能
   • 基于上传文档自动生成回复话术
   • 支持不同场景的话术分类
   • 个性化回复风格学习

3. 回复优化功能
   • 根据对话历史优化回复质量
   • 支持A/B测试不同回复策略
   • 实时效果反馈和调整

4. 管理功能
   • 话术库管理和分类
   • 训练数据版本控制
   • 效果统计和分析报告

此功能将大大提升AI回复的专业性和准确性，
让您的客服回复更加智能和个性化！""")
            desc_text.setStyleSheet("font-size: 12px; line-height: 1.5;")
            layout.addWidget(desc_text)

            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)

            dialog.exec()

        except Exception as e:
            QMessageBox.warning(self, "错误", f"显示对话框失败: {str(e)}",
                              QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def _update_settings_in_rules_file(self, updates: dict):
        """将部分设置项写入 config/auto_reply_rules.json 的 settings 字段中，保持与 load_rules_data 一致。
        updates: 仅包含需要更新的键值对，例如 {'ai_reply_mode': 0, 'model_reply_enabled': True}
        """
        try:
            config_dir = "config"
            os.makedirs(config_dir, exist_ok=True)
            rules_file = os.path.join(config_dir, 'auto_reply_rules.json')

            data = {}
            if os.path.exists(rules_file):
                try:
                    with open(rules_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                except Exception:
                    data = {}

            if not isinstance(data, dict):
                data = {}

            settings = data.get('settings', {})
            if not isinstance(settings, dict):
                settings = {}

            settings.update(updates or {})

            data['settings'] = settings
            data['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            with open(rules_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            try:
                self._cleanup_obsolete_ai_config()
            except Exception:
                pass
        except Exception as e:
            pass
    def export_reply_history(self):
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出接收消息",
                f"自动回复记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                data = []
                headers = ["序号", "本微信昵称", "好友/群昵称", "好友ID", "消息内容", "接收时间"]

                for row in range(self.auto_reply_history_table.rowCount()):
                    row_data = {}

                    for col, header in enumerate(headers):
                        item = self.auto_reply_history_table.item(row, col)
                        row_data[header] = item.text() if item else ""

                    data.append(row_data)

                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active

                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

                wb.save(file_path)

                self.statusBar().showMessage(f"已导出回复接收消息到: {file_path}", 3000)
                QMessageBox.information(self, "导出成功", "接收消息导出成功！", QMessageBox.StandardButton.Ok)

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出接收消息失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def process_auto_reply(self, message_data):
        try:
            sender_wxid = message_data.get('sender_wxid', '')

            if sender_wxid == "wxid_wi_1d142z0zdj03":
                return

            if hasattr(self, 'rule_reply_switch') and self.rule_reply_switch.isChecked():
                self.check_and_auto_reply(message_data)

        except Exception as e:
            pass
    def on_message_received(self, message):
        try:
            print("收到微信消息:", message)

            try:
                self.process_message_for_auto_reply(message)
            except Exception as e:
                pass
            self.data_manager.save_message(message)

            timestamp = int(message.get('timestamp', time.time()))
            time_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            wxid = message.get('wxid', '')

            if wxid == "wxid_wi_1d142z0zdj03":
                return

            content = message.get('content', '')
            original_content = content
            account_info = message.get('account', {})
            account_name = account_info.get('nickname', '未知账号')

            is_at_me = False

            print(f"解析消息: 发送者={wxid}, 内容={content}, 时间={time_str}")

            parsed_content = parse_special_message(content)
            if parsed_content:
                content = parsed_content
                print(f"解析XML内容: {content}")

            try:
                text_to_check = content or original_content or ''
                if "操作过于频繁，请稍后再试" in text_to_check:
                    print("检测到频繁操作提示，自动停止添加好友流程")
                    if not hasattr(self, 'rate_limit_triggered'):
                        self.rate_limit_triggered = False
                    if not self.rate_limit_triggered:
                        self.rate_limit_triggered = True
                        if getattr(self, 'is_running', False):
                            self.stop_process()
                            try:
                                self.add_friend_status.setText("检测到微信限制：操作过于频繁，已自动停止")
                            except Exception:
                                pass
                            try:
                                QMessageBox.warning(self, "已自动停止", "检测到微信提示：操作过于频繁，请稍后再试。系统已自动停止添加好友任务。")
                            except Exception:
                                pass
            except Exception:
                pass

            is_group_message = "@chatroom" in wxid
            member_id = message.get("member_id", "")

            is_at_me = False

            contact_info = None
            if hasattr(self, 'all_contacts') and self.all_contacts:
                for contact in self.all_contacts:
                    if contact.get("wxid") == wxid:
                        contact_info = contact
                        break

            sender_name = wxid
            if is_group_message:
                if contact_info:
                    sender_name = contact_info.get("nickname", wxid)

                member_info = None
                if hasattr(self, 'all_contacts') and self.all_contacts:
                    for contact in self.all_contacts:
                        if contact.get("wxid") == member_id:
                            member_info = contact
                            break

                self_nickname = account_info.get('nickname', '')
                self_wxid = account_info.get('wxid', '')

                is_at_me = False
                if self_nickname and f"@{self_nickname}" in content:
                    is_at_me = True

                    content = content.replace(f"@{self_nickname}", "").strip()
                    print(f"群消息包含@我，原内容: '{original_content}'，过滤后内容: '{content}'")

                elif self_wxid and f"@{self_wxid}" in content:
                    is_at_me = True

                    content = content.replace(f"@{self_wxid}", "").strip()
                    print(f"群消息包含@我，原内容: '{original_content}'，过滤后内容: '{content}'")

                if member_info:
                    member_name = member_info.get("nickname", member_id)
            else:
                if contact_info:
                    remarks = contact_info.get("remarks", "")
                    if remarks and remarks.strip():
                        sender_name = remarks
                    else:
                        sender_name = contact_info.get("nickname", wxid)

            message_data = {
                'self_nickname': account_name,
                'sender_nickname': sender_name,
                'sender_wxid': wxid,
                'content': content,
                'original_content': original_content,
                'receive_time': time_str,
                'account': account_info,
                'is_at_me': is_at_me if is_group_message else False
            }

            if is_group_message and member_id:
                member_name = member_id
                if member_info:
                    member_name = member_info.get("nickname", member_id)
                message_data['member_name'] = member_name
                message_data['member_id'] = member_id

            self.add_message_to_auto_reply_history(message_data)

            try:
                self.handle_auto_remark_on_acceptance(message_data)
            except Exception as e:
                print(f"自动备注处理异常: {e}")
            self.statusBar().showMessage(f"收到新消息: {content[:20]}...", 3000)

        except Exception as e:
            pass
    def check_zombie_fans(self):
        try:
            pid = self._get_wechat_pid()
            if not pid:
                QMessageBox.warning(self, "提示", "未找到已登录的微信账号", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            QMessageBox.information(self, "功能开发中", "僵尸粉检测功能正在开发中，敬请期待！", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            self.auto_fetch_contacts()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"检测僵尸粉失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            print(traceback.format_exc())

    def start_message_monitoring(self):
        try:
            self.monitor_manager.start_monitor_all()

            self.statusBar().showMessage("已自动启动消息监听，持续监听中...", 5000)
            self.monitor_check_timer = QTimer()
            self.monitor_check_timer.timeout.connect(self.check_monitor_status)
            self.monitor_check_timer.start(60000)
        except Exception as e:
            pass
    def check_monitor_status(self):
        try:
            if not self.monitor_manager.is_running:
                self.monitor_manager.start_monitor_all()
        except Exception as e:
            pass
    def export_contacts(self):
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出联系人数据",
                "微信联系人数据.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            if not hasattr(self, 'all_contacts') or not self.all_contacts:
                QMessageBox.warning(self, "导出失败", "没有联系人数据，请先获取联系人信息", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            friends_data = []
            groups_data = []
            all_members_data = []

            for contact in self.all_contacts:
                if "@chatroom" in contact.get("wxid", ""):
                    group = {
                        "群名称": contact.get("nickname", ""),
                        "群ID": contact.get("wxid", "")
                    }
                    groups_data.append(group)
                else:
                    friend = {
                        "昵称": contact.get("nickname", ""),
                        "微信ID": contact.get("wxid", ""),
                        "备注": contact.get("remarks", ""),
                        "标签": contact.get("tag", ""),
                        "手机号码": contact.get("phone", ""),
                    }
                    friends_data.append(friend)

            if groups_data:
                self.statusBar().showMessage("正在获取所有群的成员信息...", 0)
                QApplication.processEvents()

                try:
                    pid = self._get_wechat_pid()
                    if pid:
                        for i, group in enumerate(groups_data, 1):
                            group_id = group["群ID"]
                            group_name = group["群名称"]

                            def progress_callback(current, total, member):
                                if member:
                                    pass
                                else:
                                    pass
                            self.statusBar().showMessage(
                                f"正在获取群 [{i}/{len(groups_data)}] {group_name} 的成员信息...", 0
                            )
                            QApplication.processEvents()

                            members = get_group_members(pid, group_id, progress_callback)

                            for member in members:
                                member_data = {
                                    "群名称": group_name,
                                    "群ID": group_id,
                                    "群成员昵称": member.get("nickname", ""),
                                    "群成员ID": member.get("wxid", "")
                                }
                                all_members_data.append(member_data)

                        self.statusBar().showMessage(
                            f"成功获取 {len(groups_data)} 个群共 {len(all_members_data)} 个群成员信息", 3000
                        )
                    else:
                        self.statusBar().showMessage("未找到微信进程，跳过群成员获取", 3000)

                except Exception as e:
                    self.statusBar().showMessage("获取群成员信息失败，仅导出好友和群组信息", 3000)
            else:
                self.statusBar().showMessage("没有群组数据，跳过群成员获取", 3000)

            data = {
                "好友列表": friends_data,
                "群列表": groups_data,
                "群成员列表": all_members_data
            }

            self.export_to_excel(file_path, data)

            self.statusBar().showMessage(f"已导出 {len(friends_data)} 个好友、{len(groups_data)} 个群组和 {len(all_members_data)} 个群成员信息到: {file_path}", 5000)
            QMessageBox.information(self, "导出成功", f"已导出 {len(friends_data)} 个好友、{len(groups_data)} 个群组和 {len(all_members_data)} 个群成员信息！", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出联系人数据失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

    def export_to_excel(self, file_path, data):
        try:
            from openpyxl import Workbook
            wb = Workbook()

            wb.remove(wb.active)

            for sheet_name, sheet_data in data.items():
                if sheet_name and sheet_data:
                    ws = wb.create_sheet(title=sheet_name)

                    if sheet_data:
                        headers = list(sheet_data[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)

                        for row_idx, row_data in enumerate(sheet_data, 2):
                            for col_idx, header in enumerate(headers, 1):
                                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

            wb.save(file_path)
        except Exception as e:
            raise Exception(f"Excel导出失败: {str(e)}")

    def export_to_csv(self, file_path, data):
        try:
            if data:
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    fieldnames = list(data[0].keys())
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
        except Exception as e:
            raise Exception(f"CSV导出失败: {str(e)}")

    def start_monitoring(self):
        try:
            pids = detect_wechat_processes()
            if not pids:
                self.add_friend_status.setText("未找到微信进程，请确保微信已启动")
                return

            try:
                for pid in list(self.contact_monitors.keys()):
                    if pid not in pids:
                        mon = self.contact_monitors.pop(pid, None)
                        if mon and mon.is_active():
                            try:
                                mon.stop()
                            except Exception:
                                pass
            except Exception:
                pass

            started = 0
            for pid in pids:
                mon = self.contact_monitors.get(pid)
                if mon is None:
                    mon = ContactInfoMonitor(pid)
                    self.contact_monitors[pid] = mon
                if not mon.is_active():
                    mon.set_callback(self.on_contact_info)
                    mon.start()
                    started += 1

            self.add_friend_status.setText(f"监听已启动，共{len(pids)}个微信，新增启动{started}个。等待搜索...")

        except Exception as e:
            self.add_friend_status.setText(f"监听启动失败: {str(e)}")

    def on_contact_info(self, info):
        try:
            phone = info.get('phone', '')
            v3 = info.get('v3', '')
            nickname = info.get('nickname', '')

            if not phone and not v3 and not nickname:
                return

            for row in range(self.add_friend_table.rowCount()):
                table_phone = self.add_friend_table.item(row, 1).text() if self.add_friend_table.item(row, 1) else ""

                if table_phone and table_phone == phone:
                    if v3:
                        self.add_friend_table.setItem(row, 4, QTableWidgetItem(v3))
                    if nickname:
                        self.add_friend_table.setItem(row, 5, QTableWidgetItem(nickname))

                    # 使用QTimer延迟保存，确保在主线程中执行
                    if QThread.currentThread() == QApplication.instance().thread():
                        QTimer.singleShot(100, self.save_add_friend_data)
                    else:
                        # 如果不在主线程，直接保存
                        self.save_add_friend_data()
                    return

        except Exception as e:
            print(f"联系人信息处理失败: {e}")
    def import_from_xls(self):
        if self.add_friend_table.rowCount() > 0:
            reply = QMessageBox.question(
                self,
                "确认导入",
                "当前列表中有数据，是否需要保存？\n选择'是'将导出数据后再导入，选择'否'将直接导入。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )

            if reply == QMessageBox.StandardButton.Yes:
                self.export_data()

            self.add_friend_table.setRowCount(0)

        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xls *.xlsx)")

        if not file_path:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            data = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # 表头行
                    headers = [str(cell) if cell else '' for cell in row]
                else:  # 数据行
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            row_data[headers[col_idx]] = str(cell) if cell else ''
                    if row_data:
                        data.append(row_data)

            column_mapping = {
                '手机号': ['手机号', '手机', '电话', '电话号码'],
                '招呼语': ['招呼语', '招呼消息', '验证消息', '验证信息'],
                '备注': ['备注', 'remark', '备注名', '备注信息']
            }

            found_columns = {}
            for target, possible_names in column_mapping.items():
                for col in headers:
                    if any(name in str(col) for name in possible_names):
                        found_columns[target] = col
                        break

            phone_numbers = []
            greetings = {}
            remarks = {}

            if '手机号' in found_columns:
                phone_col = found_columns['手机号']
                greeting_col = found_columns.get('招呼语', None)
                remark_col = found_columns.get('备注', None)

                for row_data in data:
                    phone = str(row_data.get(phone_col, '')).strip()
                    if phone.isdigit() and len(phone) >= 5:
                        phone_numbers.append(phone)

                        if greeting_col and row_data.get(greeting_col):
                            greeting = str(row_data.get(greeting_col, '')).strip()
                            if greeting:
                                greetings[phone] = greeting

                        if remark_col and row_data.get(remark_col):
                            remark_text = str(row_data.get(remark_col, '')).strip()
                            if remark_text:
                                remarks[phone] = remark_text
            else:
                for column in headers:
                    for row_data in data:
                        value = row_data.get(column, '')
                        if value:
                            str_value = str(value).strip()
                            if str_value.isdigit() and len(str_value) >= 5:
                                phone_numbers.append(str_value)

            if not phone_numbers:
                self.add_friend_status.setText("未在Excel文件中找到有效手机号")
                return

            unique_phones = []
            seen_phones = set()
            for phone in phone_numbers:
                if phone not in seen_phones:
                    unique_phones.append(phone)
                    seen_phones.add(phone)

            default_greeting = "您好，我想添加您为好友"
            for phone in unique_phones:
                greeting = greetings.get(phone, default_greeting)
                remark_text = remarks.get(phone, "")
                self.add_phone_to_table(phone, greeting, remark=remark_text)

            self.save_add_friend_data()

            self.add_friend_status.setText(f"已从Excel导入 {len(phone_numbers)} 个手机号")

        except Exception as e:
            self.add_friend_status.setText(f"导入失败: {str(e)}")
        import traceback
    def add_phone_to_table(self, phone, greeting="您好，我想添加您为好友", status="等待中", remark=""):
        for row in range(self.add_friend_table.rowCount()):
            if self.add_friend_table.item(row, 1) and self.add_friend_table.item(row, 1).text() == phone:
                return

        row = self.add_friend_table.rowCount()
        self.add_friend_table.insertRow(row)

        for i in range(self.add_friend_table.rowCount()):
            self.add_friend_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))

        self.add_friend_table.setItem(row, 1, QTableWidgetItem(phone))
        self.add_friend_table.setItem(row, 2, QTableWidgetItem(greeting))
        self.add_friend_table.setItem(row, 3, QTableWidgetItem(status))
        self.add_friend_table.setItem(row, 4, QTableWidgetItem(""))
        self.add_friend_table.setItem(row, 5, QTableWidgetItem(""))
        self.add_friend_table.setItem(row, 6, QTableWidgetItem(remark))

    def clear_table(self):
        if self.add_friend_table.rowCount() > 0:
            reply = QMessageBox.question(
                self,
                "确认清空",
                "当前列表中有数据，是否需要保存？\n选择'是'将导出数据后再清空，选择'否'将直接清空。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )

            if reply == QMessageBox.StandardButton.Yes:
                self.export_data()

        self.add_friend_table.setRowCount(0)
        self.save_add_friend_data()
        self.add_friend_status.setText("列表已清空")

    def start_process(self):
        if self.add_friend_table.rowCount() == 0:
            self.add_friend_status.setText("请先导入手机号")
            return

        wechat_info = SimpleWeChatInfo()
        accounts = wechat_info.run()

        if not accounts:
            self.add_friend_status.setText("未找到已登录的微信账号")
            return

        dialog = WeChatAccountSelectionDialog(accounts, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        selected_accounts = dialog.get_selected_accounts()
        if not selected_accounts:
            self.add_friend_status.setText("请至少选择一个微信账号")
            return

        self.selected_wechat_accounts = selected_accounts
        self.current_account_index = 0

        try:
            min_delay = max(1, int(self.min_delay_input.text()))
            max_delay = max(min_delay, int(self.max_delay_input.text()))
        except ValueError:
            min_delay = 10
            max_delay = 60
            self.min_delay_input.setText(str(min_delay))
            self.max_delay_input.setText(str(max_delay))

        self.is_running = True
        self.is_paused = False
        self.rate_limit_triggered = False

        first_waiting_index = self._find_next_waiting_index(0)
        if first_waiting_index == -1:
            self.is_running = False
            self.add_friend_status.setText("没有'等待中'的手机号")
            self.start_button.setEnabled(True)
            self.pause_button.setEnabled(False)
            self.resume_button.setEnabled(False)
            self.stop_button.setEnabled(False)
            self.import_button.setEnabled(True)
            self.clear_button.setEnabled(True)
            return
        self.current_index = first_waiting_index

        self.current_account_index = self.current_index % len(self.selected_wechat_accounts)

        self.initial_no_delay = True

        self.start_button.setEnabled(False)
        self.pause_button.setEnabled(True)
        self.resume_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.import_button.setEnabled(False)
        self.clear_button.setEnabled(False)

        self.add_friend_status.setText(f"开始处理... 使用账号: {selected_accounts[0]['nickname']}")
        self.process_next_phone(min_delay, max_delay)

    def pause_process(self):
        self.is_paused = True
        self.pause_button.setEnabled(False)
        self.resume_button.setEnabled(True)
        self.add_friend_status.setText("已暂停")

    def resume_process(self):
        self.is_paused = False
        self.pause_button.setEnabled(True)
        self.resume_button.setEnabled(False)
        self.add_friend_status.setText("继续处理...")

        try:
            min_delay = max(1, int(self.min_delay_input.text()))
            max_delay = max(min_delay, int(self.max_delay_input.text()))
        except ValueError:
            min_delay = 10
            max_delay = 60

        self.process_next_phone(min_delay, max_delay)

    def stop_process(self):
        self.is_running = False
        self.is_paused = False

        self.start_button.setEnabled(True)
        self.pause_button.setEnabled(False)
        self.resume_button.setEnabled(False)
        self.stop_button.setEnabled(False)
        self.import_button.setEnabled(True)
        self.clear_button.setEnabled(True)

        self.add_friend_status.setText("已结束")

    def process_next_phone(self, min_delay, max_delay):
        if not self.is_running or self.is_paused:
            return

        total_rows = self.add_friend_table.rowCount()
        if self.current_index >= total_rows:
            self.stop_process()
            self.add_friend_status.setText("所有手机号处理完成")
            return

        def get_status_at(index: int) -> str:
            return self.add_friend_table.item(index, 3).text() if (self.add_friend_table.item(index, 3)) else ""
        if get_status_at(self.current_index) != "等待中":
            next_waiting_index = self._find_next_waiting_index(self.current_index + 1)
            if next_waiting_index == -1:
                self.stop_process()
                self.add_friend_status.setText("所有手机号处理完成")
                return
            self.current_index = next_waiting_index

        if self.current_index >= total_rows:
            self.stop_process()
            self.add_friend_status.setText("所有手机号处理完成")
            return

        phone = self.add_friend_table.item(self.current_index, 1).text()
        greeting = self.add_friend_table.item(self.current_index, 2).text()

        current_account = self.selected_wechat_accounts[self.current_account_index]

        self.add_friend_table.setItem(self.current_index, 3, QTableWidgetItem(f"搜索中({current_account['nickname']})"))
        self.add_friend_status.setText(f"正在搜索: {phone} (账号: {current_account['nickname']})")

        try:
            wechat_pid = current_account['pid']
            if not wechat_pid:
                self.add_friend_table.setItem(self.current_index, 3, QTableWidgetItem("失败：未找到微信进程"))
                self.schedule_next_search(min_delay, max_delay)
                return

            handle = OpenProcess(0x1F0FFF, False, wechat_pid)
            if not handle:
                self.add_friend_table.setItem(self.current_index, 3, QTableWidgetItem("失败：无法打开进程"))
                self.schedule_next_search(min_delay, max_delay)
                return

            wechat_base = get_wechat_base(wechat_pid)
            if not wechat_base:
                self.add_friend_table.setItem(self.current_index, 3, QTableWidgetItem("失败：无法获取微信基址"))
                CloseHandle(handle)
                self.schedule_next_search(min_delay, max_delay)
                return

            v3_info = self.add_friend_table.item(self.current_index, 4).text() if self.add_friend_table.item(self.current_index, 4) else ""
            is_group_member = v3_info and v3_info.startswith("wxid_")

            if is_group_member:
                # 确保在主线程中调度
                if QThread.currentThread() == QApplication.instance().thread():
                    QTimer.singleShot(1000, lambda: self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, None))
                else:
                    # 如果不在主线程，立即执行
                    self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, None)
            else:
                monitor = None
                try:
                    monitor = ContactInfoMonitor(wechat_pid)
                    monitor.set_callback(self.on_contact_info)
                    monitor.start()

                    result = add_friend_by_phone(handle, wechat_base, phone)

                    # 确保在主线程中调度
                    if QThread.currentThread() == QApplication.instance().thread():
                        QTimer.singleShot(3000, lambda: self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, monitor))
                    else:
                        # 如果不在主线程，立即执行
                        self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, monitor)
                except Exception as monitor_error:
                    print(f"监控器创建失败: {monitor_error}")
                    if monitor:
                        try:
                            monitor.stop()
                        except:
                            pass
                    # 继续执行，但不使用监控器，确保在主线程中调度
                    if QThread.currentThread() == QApplication.instance().thread():
                        QTimer.singleShot(3000, lambda: self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, None))
                    else:
                        # 如果不在主线程，立即执行
                        self.check_and_add_friend(handle, wechat_base, self.current_index, min_delay, max_delay, None)

        except Exception as e:
            self.add_friend_table.setItem(self.current_index, 3, QTableWidgetItem(f"失败: {str(e)[:20]}"))
            self.schedule_next_search(min_delay, max_delay)

    def check_and_add_friend(self, handle, wechat_base, row_index, min_delay, max_delay, monitor=None):
        if not self.is_running or self.is_paused:
            try:
                CloseHandle(handle)
            except:
                pass
            if monitor:
                try:
                    monitor.stop()
                except Exception as e:
                    print(f"停止监控器失败: {e}")
            return

        try:
            current_account = self.selected_wechat_accounts[self.current_account_index]

            v3_info = self.add_friend_table.item(row_index, 4).text() if self.add_friend_table.item(row_index, 4) else ""
            phone = self.add_friend_table.item(row_index, 1).text() if self.add_friend_table.item(row_index, 1) else ""

            if not v3_info and phone:
                self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"无微信号({current_account['nickname']})"))
                try:
                    CloseHandle(handle)
                except:
                    pass
                if monitor:
                    try:
                        monitor.stop()
                    except Exception as e:
                        print(f"停止监控器失败: {e}")
                self.save_add_friend_data()
                self.schedule_next_search(min_delay, max_delay)
                return

            try:
                CloseHandle(handle)
            except:
                pass

            is_group_member = v3_info and v3_info.startswith("wxid_")

            phone = self.add_friend_table.item(row_index, 1).text() if self.add_friend_table.item(row_index, 1) else ""
            is_phone_search = bool(phone)

            if is_group_member:
                wechat_pid = current_account['pid']
                if wechat_pid:
                    self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加中({current_account['nickname']})"))
                    greeting = self.add_friend_table.item(row_index, 2).text() if self.add_friend_table.item(row_index, 2) else "您好，我想添加您为好友"

                    result = add_wechat_friend(wechat_pid, v3_info, greeting)

                    if result:
                        self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加成功({current_account['nickname']})"))
                    else:
                        self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加失败({current_account['nickname']})"))
                else:
                    self.add_friend_table.setItem(row_index, 3, QTableWidgetItem("失败：未找到微信进程"))
            elif is_phone_search:
                if v3_info and v3_info.startswith("v3_"):
                    wechat_pid = current_account['pid']
                    if wechat_pid:
                        self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加中({current_account['nickname']})"))
                        greeting = self.add_friend_table.item(row_index, 2).text() if self.add_friend_table.item(row_index, 2) else "您好，我想添加您为好友"

                        result = add_wechat_friend(wechat_pid, v3_info, greeting)

                        if result:
                            self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加成功({current_account['nickname']})"))
                        else:
                            self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"添加失败({current_account['nickname']})"))
                    else:
                        self.add_friend_table.setItem(row_index, 3, QTableWidgetItem("失败：未找到微信进程"))
                elif v3_info and not v3_info.startswith("v3_"):
                    self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"已是好友({current_account['nickname']})"))
                else:
                    self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"无微信号({current_account['nickname']})"))
            else:
                self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"无微信号({current_account['nickname']})"))

            self.save_add_friend_data()

            if monitor:
                try:
                    monitor.stop()
                except Exception as e:
                    print(f"停止监控器失败: {e}")

            self.schedule_next_search(min_delay, max_delay)

        except Exception as e:
            print(f"添加好友过程异常: {e}")
            self.add_friend_table.setItem(row_index, 3, QTableWidgetItem(f"失败: {str(e)[:20]}"))
            try:
                CloseHandle(handle)
            except:
                pass
            if monitor:
                try:
                    monitor.stop()
                except Exception as monitor_error:
                    print(f"异常处理中停止监控器失败: {monitor_error}")
            self.schedule_next_search(min_delay, max_delay)

    def schedule_next_search(self, min_delay, max_delay):
        # 确保在主线程中执行
        def _schedule_in_main_thread():
            if not self.is_running or self.is_paused:
                return

            next_waiting_index = self._find_next_waiting_index(self.current_index + 1)
            if next_waiting_index == -1:
                self.stop_process()
                self.add_friend_status.setText("所有手机号处理完成")
                return
            self.current_index = next_waiting_index
            self.current_account_index = self.current_index % len(self.selected_wechat_accounts)

            import random
            if getattr(self, 'initial_no_delay', False):
                delay_minutes = 0
                self.initial_no_delay = False
            else:
                delay_minutes = random.randint(min_delay, max_delay)
            current_account = self.selected_wechat_accounts[self.current_account_index]
            if delay_minutes > 0:
                self.add_friend_status.setText(f"等待 {delay_minutes} 分钟后，账号 {current_account['nickname']} 处理第 {self.current_index + 1} 个手机号...")
            else:
                self.add_friend_status.setText(f"立即由账号 {current_account['nickname']} 处理第 {self.current_index + 1} 个手机号...")

            # 使用QTimer.singleShot在主线程中调度
            QTimer.singleShot(delay_minutes * 60 * 1000, lambda: self.process_next_phone(min_delay, max_delay))
        
        # 如果当前在主线程，直接执行；否则使用QTimer调度到主线程
        if QThread.currentThread() == QApplication.instance().thread():
            _schedule_in_main_thread()
        else:
            # 使用QTimer.singleShot确保在主线程执行
            QTimer.singleShot(0, _schedule_in_main_thread)

    def _find_next_waiting_index(self, start_index: int) -> int:
        total_rows = self.add_friend_table.rowCount()
        for idx in range(max(0, start_index), total_rows):
            item = self.add_friend_table.item(idx, 3)
            status_text = item.text() if item else ""
            if status_text == "等待中":
                return idx
        return -1

    def save_add_friend_data(self):
        try:
            data = []
            for row in range(self.add_friend_table.rowCount()):
                row_data = {
                    "phone": self.add_friend_table.item(row, 1).text() if self.add_friend_table.item(row, 1) else "",
                    "greeting": self.add_friend_table.item(row, 2).text() if self.add_friend_table.item(row, 2) else "",
                    "status": self.add_friend_table.item(row, 3).text() if self.add_friend_table.item(row, 3) else "",
                    "v3": self.add_friend_table.item(row, 4).text() if self.add_friend_table.item(row, 4) else "",
                    "nickname": self.add_friend_table.item(row, 5).text() if self.add_friend_table.item(row, 5) else "",
                    "remark": self.add_friend_table.item(row, 6).text() if self.add_friend_table.item(row, 6) else ""
                }
                data.append(row_data)

            import json
            with open(self.data_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存数据失败: {e}")
            # 尝试备份保存
            try:
                backup_file = self.data_file + ".backup"
                with open(backup_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"数据已备份到: {backup_file}")
            except:
                print("备份保存也失败")
    def load_add_friend_data(self):
        try:
            if not os.path.exists(self.data_file):
                return

            import json
            with open(self.data_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            for item in data:
                row = self.add_friend_table.rowCount()
                self.add_friend_table.insertRow(row)

                self.add_friend_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))

                self.add_friend_table.setItem(row, 1, QTableWidgetItem(item.get("phone", "")))
                self.add_friend_table.setItem(row, 2, QTableWidgetItem(item.get("greeting", "")))
                self.add_friend_table.setItem(row, 3, QTableWidgetItem(item.get("status", "")))
                self.add_friend_table.setItem(row, 4, QTableWidgetItem(item.get("v3", "")))
                self.add_friend_table.setItem(row, 5, QTableWidgetItem(item.get("nickname", "")))
                self.add_friend_table.setItem(row, 6, QTableWidgetItem(item.get("remark", "")))

                if "monitor_phone" in item and not item.get("v3", "") and item.get("monitor_phone", ""):
                    self.add_friend_table.setItem(row, 4, QTableWidgetItem(item.get("monitor_phone", "")))
        except Exception as e:
            pass
    def export_data(self):
        try:
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getSaveFileName(
                self,
                "导出数据",
                f"微信好友数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel文件 (*.xlsx);;CSV文件 (*.csv)"
            )

            if not file_path:
                return

            data = []
            for row in range(self.add_friend_table.rowCount()):
                row_data = {
                    "序号": self.add_friend_table.item(row, 0).text() if self.add_friend_table.item(row, 0) else "",
                    "手机号": self.add_friend_table.item(row, 1).text() if self.add_friend_table.item(row, 1) else "",
                    "招呼语": self.add_friend_table.item(row, 2).text() if self.add_friend_table.item(row, 2) else "",
                    "状态情况": self.add_friend_table.item(row, 3).text() if self.add_friend_table.item(row, 3) else "",
                    "v3信息": self.add_friend_table.item(row, 4).text() if self.add_friend_table.item(row, 4) else "",
                    "联系人昵称": self.add_friend_table.item(row, 5).text() if self.add_friend_table.item(row, 5) else ""
                }
                data.append(row_data)

            if file_path.endswith('.xlsx'):
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active

                if data:
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

                wb.save(file_path)
            elif file_path.endswith('.csv'):
                import csv
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    if data:
                        fieldnames = list(data[0].keys())
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(data)
            else:
                if not (file_path.endswith('.xlsx') or file_path.endswith('.csv')):
                    file_path += '.xlsx'
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active

                if data:
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)

                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

                wb.save(file_path)

            self.add_friend_status.setText(f"成功导出 {len(data)} 条数据到: {file_path}")
            QMessageBox.information(self, "导出成功", f"成功导出 {len(data)} 条数据！")

        except Exception as e:
            self.add_friend_status.setText(f"导出失败: {str(e)}")
            QMessageBox.warning(self, "导出失败", f"导出失败: {str(e)}")
    def add_group_members(self):
        self.notebook.setCurrentIndex(0)

        QMessageBox.information(self, "提示", "请在主界面的群列表中右键点击群聊，选择'添加群成员'功能", QMessageBox.StandardButton.Ok)

    def add_group_members_from_context(self, group_id, group_name):
        try:
            if self.add_friend_table.rowCount() > 0:
                reply = QMessageBox.question(
                    self,
                    "确认添加群成员",
                    "当前列表中有数据，是否需要保存？\n选择'是'将导出数据后再添加群成员，选择'否'将直接添加群成员。",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )

                if reply == QMessageBox.StandardButton.Yes:
                    self.export_data()

                self.add_friend_table.setRowCount(0)

            pid = self._get_wechat_pid()
            if not pid:
                QMessageBox.warning(self, "获取失败", "未找到微信进程，请确保微信已启动", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            def progress_callback(current, total, member):
                pass

            members = get_group_members(pid, group_id, progress_callback)

            if not members:
                QMessageBox.warning(self, "获取失败", f"未能获取到群 {group_name} 的成员信息", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)
                return

            existing_friends = set()
            for row in range(self.friend_tree.topLevelItemCount()):
                item = self.friend_tree.topLevelItem(row)
                if item:
                    friend_wxid = item.text(2)
                    if friend_wxid:
                        existing_friends.add(friend_wxid)

            current_wxid = None
            for row in range(self.account_tree.topLevelItemCount()):
                item = self.account_tree.topLevelItem(row)
                if item:
                    current_wxid = item.text(2)
                    break

            self.notebook.setCurrentIndex(1)

            existing_table_wxids = set()
            for _row in range(self.add_friend_table.rowCount()):
                _item = self.add_friend_table.item(_row, 4)
                if _item:
                    _v = _item.text()
                    if _v:
                        existing_table_wxids.add(_v)

            to_add = []
            skipped_count = 0
            for member in members:
                wxid = member.get("wxid", "")
                if not wxid:
                    skipped_count += 1
                    continue
                if current_wxid and wxid == current_wxid:
                    skipped_count += 1
                    continue
                if wxid in existing_friends:
                    skipped_count += 1
                    continue
                if wxid in existing_table_wxids:
                    skipped_count += 1
                    continue
                to_add.append({
                    "wxid": wxid,
                    "nickname": member.get("nickname", "无昵称")
                })

            added_count = len(to_add)

            if added_count > 0:
                self.add_friend_table.blockSignals(True)
                self.add_friend_table.setUpdatesEnabled(False)
                try:
                    start_row = self.add_friend_table.rowCount()
                    new_total = start_row + added_count
                    self.add_friend_table.setRowCount(new_total)

                    for idx, m in enumerate(to_add):
                        row = start_row + idx
                        self.add_friend_table.setItem(row, 1, QTableWidgetItem(""))
                        self.add_friend_table.setItem(row, 2, QTableWidgetItem("您好，麻烦通过一下"))
                        self.add_friend_table.setItem(row, 3, QTableWidgetItem("等待中"))
                        self.add_friend_table.setItem(row, 4, QTableWidgetItem(m["wxid"]))
                        self.add_friend_table.setItem(row, 5, QTableWidgetItem(m["nickname"]))

                    for i in range(self.add_friend_table.rowCount()):
                        self.add_friend_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                finally:
                    self.add_friend_table.setUpdatesEnabled(True)
                    self.add_friend_table.blockSignals(False)

            self.save_add_friend_data()

            result_message = f"成功添加 {added_count} 个群成员到添加好友列表"
            if skipped_count > 0:
                result_message += f"\n跳过 {skipped_count} 个已是好友、已存在或本微信号的成员"

            self.add_friend_status.setText(result_message)
            QMessageBox.information(self, "添加完成", result_message, QMessageBox.StandardButton.Ok)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"添加群成员失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            print(traceback.format_exc())

    def add_group_member_to_table(self, wxid, nickname, greeting="您好，麻烦通过一下", status="等待中"):
        for row in range(self.add_friend_table.rowCount()):
            existing_v3 = self.add_friend_table.item(row, 4).text() if self.add_friend_table.item(row, 4) else ""
            if existing_v3 == wxid:
                self.add_friend_table.setItem(row, 3, QTableWidgetItem(status))
                return False

        row = self.add_friend_table.rowCount()
        self.add_friend_table.insertRow(row)

        self.add_friend_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))

        self.add_friend_table.setItem(row, 1, QTableWidgetItem(""))
        self.add_friend_table.setItem(row, 2, QTableWidgetItem(greeting))
        self.add_friend_table.setItem(row, 3, QTableWidgetItem(status))
        self.add_friend_table.setItem(row, 4, QTableWidgetItem(wxid))
        self.add_friend_table.setItem(row, 5, QTableWidgetItem(nickname))

        return True

    def handle_scheduled_message_for_friends(self, selected_items):
        try:
            selected_contacts = []
            added_wxids = set()

            for item in selected_items:
                wxid = item.text(2)
                nickname = item.text(1)
                remarks = item.text(3)
                if wxid and wxid not in added_wxids:
                    selected_contacts.append({'wxid': wxid, 'nickname': nickname, 'remarks': remarks})
                    added_wxids.add(wxid)

            if not selected_contacts:
                return

            self.notebook.setCurrentWidget(self.task_tab)
            self.task_tab.add_selected_contacts(selected_contacts)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理定时消息失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            print(traceback.format_exc())

    def handle_scheduled_message_for_groups(self, selected_items):
        try:
            selected_contacts = []
            for item in selected_items:
                wxid = item.text(2)
                nickname = item.text(1)

                contact = {
                    'wxid': wxid,
                    'nickname': nickname
                }
                selected_contacts.append(contact)

            self.notebook.setCurrentWidget(self.task_tab)

            self.task_tab.add_selected_contacts(selected_contacts)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理定时消息失败: {str(e)}", QMessageBox.StandardButton.Ok, QMessageBox.StandardButton.Ok)

            print(traceback.format_exc())

    def check_and_auto_reply(self, message_data):
        try:
            if not getattr(self, 'rule_reply_switch', None) or not self.rule_reply_switch.isChecked():
                return
            content = message_data.get('content', '')
            sender_wxid = message_data.get('sender_wxid', '')
            if not content or not sender_wxid:
                return
            if sender_wxid == "wxid_wi_1d142z0zdj03":
                return
            is_group = '@chatroom' in sender_wxid
            if is_group:
                allowed = False
                if getattr(self, 'reply_group_switch', None) and self.reply_group_switch.isChecked():
                    allowed = True
                elif getattr(self, 'specific_group_switch', None) and self.specific_group_switch.isChecked() and sender_wxid in getattr(self, 'specific_group_wxids', set()):
                    allowed = True
                if not allowed:
                    print("群消息未启用，且不在指定群列表，忽略")
                    return
                if not message_data.get('is_at_me', False):
                    print("收到群消息，但未被@，忽略")
                    return
            else:
                allowed = False
                if getattr(self, 'reply_friend_switch', None) and self.reply_friend_switch.isChecked():
                    allowed = True
                elif getattr(self, 'specific_friend_switch', None) and self.specific_friend_switch.isChecked() and sender_wxid in getattr(self, 'specific_friend_wxids', set()):
                    allowed = True
                if not allowed:
                    print("好友消息未启用，且不在指定好友列表，忽略")
                    return
            account_info = message_data.get('account', {})
            current_pid = account_info.get('pid')
            receiver_wxid = sender_wxid
            matched_rules = []
            try:
                if getattr(self, 'exact_match_switch', None) and self.exact_match_switch.isChecked():
                    for row in range(self.rules_table.rowCount()):
                        checkbox_widget = self.rules_table.cellWidget(row, 1)
                        if not checkbox_widget:
                            continue
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if not checkbox or not checkbox.isChecked():
                            continue
                        keyword_item = self.rules_table.item(row, 2)
                        reply_item = self.rules_table.item(row, 3)
                        if not keyword_item or not reply_item:
                            continue
                        keyword = keyword_item.text() or ''
                        reply = reply_item.text() or ''
                        if content.strip() == keyword.strip():
                            matched_rules.append({'row': row, 'reply': reply})
                if (not matched_rules) and getattr(self, 'fuzzy_match_switch', None) and self.fuzzy_match_switch.isChecked():
                    low_content = content.strip().lower()
                    for row in range(self.rules_table.rowCount()):
                        checkbox_widget = self.rules_table.cellWidget(row, 1)
                        if not checkbox_widget:
                            continue
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if not checkbox or not checkbox.isChecked():
                            continue
                        keyword_item = self.rules_table.item(row, 2)
                        reply_item = self.rules_table.item(row, 3)
                        if not keyword_item or not reply_item:
                            continue
                        keyword = (keyword_item.text() or '').strip()
                        reply = reply_item.text() or ''
                        if keyword and keyword.lower() in low_content:
                            matched_rules.append({'row': row, 'reply': reply})
            except Exception as e:
                pass
            if not matched_rules:
                if getattr(self, 'ai_reply_switch', None) and self.ai_reply_switch.isChecked():
                    try:
                        from aizhuli_combined import AIManager
                        ai_assistant = AIManager()
                        def on_ai_response(response):
                            try:
                                if response:
                                    formatted = self.format_ai_response(response, getattr(ai_assistant, 'auto_reply_settings', {}))
                                    self.send_delayed_reply(receiver_wxid, formatted, reply_type="ai_reply", pid=current_pid)
                            finally:
                                try:
                                    ai_assistant.response_ready.disconnect(on_ai_response)
                                except Exception:
                                    pass
                        ai_assistant.response_ready.connect(on_ai_response)
                        def get_ai_reply():
                            try:
                                import asyncio
                                prompt = content
                                coro = ai_assistant._async_ai_reply(sender_wxid, prompt)
                                try:
                                    asyncio.run(coro)
                                except RuntimeError:
                                    try:
                                        loop = asyncio.get_event_loop()
                                        loop.run_until_complete(coro)
                                    except Exception:
                                        new_loop = asyncio.new_event_loop()
                                        try:
                                            asyncio.set_event_loop(new_loop)
                                            new_loop.run_until_complete(coro)
                                        finally:
                                            new_loop.close()
                            except Exception as e:
                                pass
                        threading.Thread(target=get_ai_reply, daemon=True).start()
                    except Exception as e:
                        pass
                return
            matched_rules.sort(key=lambda x: x['row'])
            try:
                min_delay = max(0, int(self.min_interval.text()))
                max_delay = max(min_delay, int(self.max_interval.text()))
            except Exception:
                min_delay, max_delay = 1, 5
            for i, rule in enumerate(matched_rules):
                if min_delay == max_delay:
                    delay = min_delay
                else:
                    delay = random.randint(min_delay, max_delay)
                print(f"将在 {delay} 秒后发送第 {i+1} 条回复")
                QTimer.singleShot(delay * 1000, lambda r=rule['reply']: self.send_auto_reply(current_pid, receiver_wxid, r))
        except Exception as e:
            pass
    def show_contact_service_dialog(self):
        """显示联系客服对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("联系客服")
        dialog.setFixedSize(400, 150)

        layout = QVBoxLayout(dialog)

        info_label = QLabel("功能错误无法使用，或者增加其他功能，\n请联系：18402038101，电话/微信")
        info_label.setStyleSheet("font-size: 16px; color: #333333; padding: 30px;")
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        button_layout = QHBoxLayout()
        ok_button = QPushButton("确定")
        ok_button.clicked.connect(dialog.accept)
        ok_button.setFixedWidth(80)
        button_layout.addStretch()
        button_layout.addWidget(ok_button)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        dialog.exec()

class WeChatAccountSelectionDialog(QDialog):
    def __init__(self, accounts, parent=None):
        super().__init__(parent)
        self.accounts = accounts
        self.selected_accounts = []
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("选择添加好友的微信账号")
        self.setModal(True)
        self.resize(100, 200)

        layout = QVBoxLayout()

        label = QLabel("请选择要用于添加好友的微信账号：")
        label.setStyleSheet(StyleSheet.LABEL_SECONDARY)
        layout.addWidget(label)

        self.account_table = QTableWidget()
        self.account_table.setColumnCount(2)
        self.account_table.setHorizontalHeaderLabels(["选择", "账号信息"])
        self.account_table.horizontalHeader().setStretchLastSection(True)
        self.account_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.account_table.setAlternatingRowColors(True)

        self.account_table.setStyleSheet(StyleSheet.TABLE)

        self.account_table.setRowCount(len(self.accounts))
        for i, account in enumerate(self.accounts):
            checkbox = QCheckBox()
            checkbox.setStyleSheet(StyleSheet.TABLE_CHECKBOX)

            checkbox.stateChanged.connect(lambda state, row=i: self.on_checkbox_changed(row, state))

            self.account_table.setCellWidget(i, 0, checkbox)

            account_text = f"{account['nickname']} ({account['wxid']})"
            item = QTableWidgetItem(account_text)
            item.setData(Qt.ItemDataRole.UserRole, account)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # 设置为只读
            self.account_table.setItem(i, 1, item)  # 只设置第二列

        header = self.account_table.horizontalHeader()
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        for i in range(self.account_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        header.resizeSection(0, 50)
        for i in range(1, self.account_table.columnCount()):
            header.resizeSection(i, 150)

        layout.addWidget(self.account_table)

        button_layout = QHBoxLayout()

        select_all_btn = QPushButton("全选")
        select_all_btn.setStyleSheet(StyleSheet.BUTTON)
        select_all_btn.clicked.connect(self.select_all)
        button_layout.addWidget(select_all_btn)

        clear_all_btn = QPushButton("全不选")
        clear_all_btn.setStyleSheet(StyleSheet.BUTTON_SECONDARY)
        clear_all_btn.clicked.connect(self.clear_all)
        button_layout.addWidget(clear_all_btn)

        button_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet(StyleSheet.BUTTON)
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet(StyleSheet.BUTTON_SECONDARY)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def on_checkbox_changed(self, row, state):
        """复选框状态变化处理"""
        try:
            print(f"复选框状态变化: 行 {row}, 状态: {state}")
        except Exception as e:
            pass
    def select_all(self):
        """全选所有账号"""
        try:
            print("执行全选操作")
            for row in range(self.account_table.rowCount()):
                checkbox = self.account_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox):
                    checkbox.setChecked(True)
        except Exception as e:
            pass
    def clear_all(self):
        """取消全选"""
        try:
            print("执行全不选操作")
            for row in range(self.account_table.rowCount()):
                checkbox = self.account_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox):
                    checkbox.setChecked(False)
        except Exception as e:
            pass
    def get_selected_accounts(self):
        """获取选中的账号"""
        try:
            selected = []
            for row in range(self.account_table.rowCount()):
                checkbox = self.account_table.cellWidget(row, 0)
                if checkbox and isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                    item = self.account_table.item(row, 1)
                    if item:
                        account = item.data(Qt.ItemDataRole.UserRole)
                        if account:
                            selected.append(account)
            print(f"总共选择了 {len(selected)} 个账号")
            return selected
        except Exception as e:
            return []

if __name__ == "__main__":
    try:
        if not _check_single_instance("Global\\WeChatManagerAppMutex"):
            _show_already_running_message()
            sys.exit(0)

        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv)

        try:
            app.setApplicationName("AI全功能营销系统")
            if hasattr(app, 'setApplicationDisplayName'):
                app.setApplicationDisplayName("AI全功能营销系统")
        except Exception:
            pass

        apply_stylesheet(app)
        window = WeChatManagerApp()
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"程序启动失败: {str(e)}")
        sys.exit(1)
